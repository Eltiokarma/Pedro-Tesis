# ======================================================
# FLOWSHEET UI — block diagram editor for ISBL estimation
# ======================================================
# Toplevel modal con tres paneles:
#
#   [Library]  [Canvas (drag, drop, connect)]  [Properties]
#
# El user:
#   1) selecciona un tipo de equipo en Library y aprieta
#      "+ Add" (o doble-click) → aparece un bloque en el
#      centro del canvas.
#   2) arrastra el bloque a la posición que quiera.
#   3) crea un stream: click derecho sobre un bloque
#      origen → "Connect to..." → click sobre el destino.
#      (Drag de puerto a puerto sería ideal; lo hacemos así
#      en V1 por simplicidad.)
#   4) doble-click sobre un bloque o stream abre dialog
#      para editar S, material, mass flow.
#   5) "Compute": Σ Cp° × n × Lang factor → FCI → ISBL
#      implícito vía los % del proyecto.  Verifica mass
#      balance por bloque (warning si no cierra).
#   6) "Apply ISBL": inyecta el ISBL al df_capital.
#   7) "Save..." / "Open..." persiste a JSON.
#
# Refs: motor en equipment_costs.py (Turton Apx A + Lang).
# ======================================================

import json
import math

from dataclasses import dataclass, field, asdict
from typing import Optional, List, Dict

from tkinter import (
    Toplevel,
    Canvas,
    Frame,
    Label,
    Menu,
    Scrollbar,
    StringVar,
    DoubleVar,
    IntVar,
    END,
    BOTH,
    LEFT,
    RIGHT,
    TOP,
    BOTTOM,
    X,
    Y,
    W,
    E,
    N,
    S,
    NSEW,
    VERTICAL,
    HORIZONTAL,
)
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from tkinter import simpledialog

import equipment_costs as eq
import equipment_ports as ep


# ======================================================
# CONSTANTES DE LAYOUT Y ESTILO
# ======================================================

CANVAS_BG     = "#fafafc"
GRID_COLOR    = "#e8e8ee"
GRID_STEP     = 20

BLOCK_W       = 130
BLOCK_H       = 60
BLOCK_FILL    = "#ffffff"
BLOCK_BORDER  = "#5c6bc0"
BLOCK_BORDER_SEL = "#283593"
BLOCK_TEXT    = "#1a1a1a"
BLOCK_SUB     = "#6c6c70"

STREAM_COLOR     = "#37474f"
STREAM_COLOR_SEL = "#c62828"
STREAM_LABEL_BG  = "#fff8d6"

ZOOM_MIN      = 0.30
ZOOM_MAX      = 3.00
ZOOM_STEP     = 1.15        # factor por tecleo de zoom in/out
CANVAS_W_LOG  = 4000        # tamaño lógico del lienzo (modelo)
CANVAS_H_LOG  = 3000

ROUTING_GAP   = 30          # distancia mínima desde el bloque al codo


# ======================================================
# MODELO DE DATOS
# ======================================================

@dataclass
class Block:
    id:        int
    name:      str               # ej "HX-1"
    eq_type:   str               # nombre del catálogo eq.EQUIPMENT_DATA
    S:         float             # parámetro de tamaño
    n:         int = 1           # cantidad en paralelo
    x:         float = 0.0
    y:         float = 0.0

    # caches del canvas (no se serializan)
    canvas_rect: Optional[int] = field(default=None, repr=False)
    canvas_text: Optional[int] = field(default=None, repr=False)
    canvas_sub:  Optional[int] = field(default=None, repr=False)


@dataclass
class Stream:
    id:        int
    name:      str                  # ej "S-1"
    src:       int                  # id del bloque origen
    dst:       int                  # id del bloque destino
    mass_flow: float = 0.0          # tm/year (signo +)
    role:      str = "internal"     # "internal" | "feed" | "product"
    src_port:  str = ""             # nombre del puerto en src; "" = autoselect
    dst_port:  str = ""             # nombre del puerto en dst; "" = autoselect
    price_usd_per_tm: float = 0.0   # USD/tm — sólo relevante si role∈{feed,product}

    canvas_line:    Optional[int] = field(default=None, repr=False)
    canvas_label:   Optional[int] = field(default=None, repr=False)
    canvas_lbl_bg:  Optional[int] = field(default=None, repr=False)


STREAM_ROLE_COLORS = {
    "internal": "#37474f",
    "feed":     "#2e7d32",
    "product":  "#e65100",
}
STREAM_ROLE_COLORS_SEL = {
    "internal": "#c62828",
    "feed":     "#1b5e20",
    "product":  "#bf360c",
}


@dataclass
class Flowsheet:
    blocks:   Dict[int, Block]   = field(default_factory=dict)
    streams:  Dict[int, Stream]  = field(default_factory=dict)
    _next_id: int = 1

    def new_id(self):
        v = self._next_id
        self._next_id += 1
        return v

    # --- serialization ---
    def to_dict(self):
        return {
            "blocks":   {bid: {k: v for k, v in asdict(b).items() if not k.startswith("canvas_")}
                         for bid, b in self.blocks.items()},
            "streams":  {sid: {k: v for k, v in asdict(s).items() if not k.startswith("canvas_")}
                         for sid, s in self.streams.items()},
            "_next_id": self._next_id,
        }

    @staticmethod
    def from_dict(d):
        fs = Flowsheet()
        for bid, bdict in d.get("blocks", {}).items():
            b = Block(**{k: v for k, v in bdict.items() if k in Block.__annotations__})
            fs.blocks[int(bid)] = b
        for sid, sdict in d.get("streams", {}).items():
            s = Stream(**{k: v for k, v in sdict.items() if k in Stream.__annotations__})
            fs.streams[int(sid)] = s
        fs._next_id = d.get("_next_id", 1)
        return fs


# ======================================================
# DIALOGS DE EDICIÓN
# ======================================================

class BlockEditDialog:
    """Editor del equipo: tag ISA, tamaño S, n unidades en paralelo."""

    def __init__(self, parent, block):
        self.parent = parent
        self.block = block
        self.result = None

        dlg = Toplevel(parent)
        dlg.title(f"Editar equipo — {block.name}")
        dlg.geometry("380x280+300+200")
        dlg.transient(parent)
        dlg.grab_set()
        self.dlg = dlg

        spec = eq.EQUIPMENT_DATA.get(block.eq_type, {})

        frm = ttk.Frame(dlg, padding=14)
        frm.pack(fill=BOTH, expand=True)

        ttk.Label(frm, text="Tag ISA:").grid(row=0, column=0, sticky=W, pady=4)
        ttk.Label(frm, text=block.name, foreground="#555").grid(row=0, column=1, sticky=W, pady=4)

        ttk.Label(frm, text="Tipo de equipo:").grid(row=1, column=0, sticky=W, pady=4)
        ttk.Label(frm, text=block.eq_type, foreground="#555").grid(row=1, column=1, sticky=W, pady=4)

        ttk.Label(frm, text=f"Tamaño S ({spec.get('S_unit', '?')}) *:").grid(row=2, column=0, sticky=W, pady=4)
        self.entry_s = ttk.Entry(frm, justify="right", width=14)
        self.entry_s.insert(0, f"{block.S:g}")
        self.entry_s.grid(row=2, column=1, sticky=W, pady=4)

        s_min = spec.get("S_min")
        s_max = spec.get("S_max")
        if s_min is not None and s_max is not None:
            ttk.Label(
                frm,
                text=f"Rango válido (Turton): [{s_min:g} – {s_max:g}] {spec.get('S_unit', '')}",
                foreground="#888",
                font=("Segoe UI", 8),
            ).grid(row=3, column=0, columnspan=2, sticky=W, pady=(0, 6))

        ttk.Label(frm, text="N° unidades en paralelo:").grid(row=4, column=0, sticky=W, pady=4)
        self.entry_n = ttk.Entry(frm, justify="right", width=14)
        self.entry_n.insert(0, str(block.n))
        self.entry_n.grid(row=4, column=1, sticky=W, pady=4)

        ttk.Label(frm, text="Nombre personalizado:").grid(row=5, column=0, sticky=W, pady=4)
        self.entry_name = ttk.Entry(frm, width=22)
        self.entry_name.insert(0, block.name)
        self.entry_name.grid(row=5, column=1, sticky=W, pady=4)

        btns = ttk.Frame(frm)
        btns.grid(row=6, column=0, columnspan=2, pady=(10, 0), sticky=E)
        ttk.Button(btns, text="Cancelar", command=dlg.destroy).pack(side=LEFT, padx=4)
        ttk.Button(btns, text="OK",       command=self._ok).pack(side=LEFT)

        self.entry_s.focus()
        dlg.wait_window()

    def _ok(self):
        try:
            S = float(self.entry_s.get())
            n = int(self.entry_n.get())
            if S <= 0 or n < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("Inválido", "Se requiere S > 0 y n ≥ 1.")
            return
        nombre = self.entry_name.get().strip() or self.block.name
        self.block.S = S
        self.block.n = n
        self.block.name = nombre
        self.result = "ok"
        self.dlg.destroy()


class StreamEditDialog:
    """Editor de la corriente: nombre, flujo, rol, puertos src/dst."""

    def __init__(self, parent, stream, fs):
        import equipment_ports as ep_mod  # local para evitar shadowing
        self.stream = stream
        self.result = None
        self.fs = fs

        dlg = Toplevel(parent)
        dlg.title(f"Editar corriente — {stream.name}")
        dlg.geometry("440x420+320+200")
        dlg.transient(parent)
        dlg.grab_set()
        self.dlg = dlg

        b_src = fs.blocks[stream.src]
        b_dst = fs.blocks[stream.dst]
        ports_src = list(ep_mod.get_ports(b_src.eq_type).keys())
        ports_dst = list(ep_mod.get_ports(b_dst.eq_type).keys())

        frm = ttk.Frame(dlg, padding=14)
        frm.pack(fill=BOTH, expand=True)

        ttk.Label(frm, text=f"Corriente {stream.name}",
                  font=("Segoe UI", 10, "bold")).grid(
            row=0, column=0, columnspan=2, sticky=W, pady=4)
        ttk.Label(frm, text=f"{b_src.name}  →  {b_dst.name}",
                  foreground="#555").grid(
            row=1, column=0, columnspan=2, sticky=W, pady=2)

        # ---- flujo másico ----
        ttk.Label(frm, text="Flujo másico (tm/año):").grid(row=2, column=0, sticky=W, pady=6)
        self.entry_m = ttk.Entry(frm, justify="right", width=14)
        self.entry_m.insert(0, f"{stream.mass_flow:g}")
        self.entry_m.grid(row=2, column=1, sticky=W, pady=6)

        # ---- rol ----
        ttk.Label(frm, text="Rol:").grid(row=3, column=0, sticky=W, pady=6)
        self.role_var = StringVar(value=stream.role)
        self.role_combo = ttk.Combobox(
            frm, textvariable=self.role_var,
            values=["internal", "feed", "product"],
            state="readonly", width=12,
        )
        self.role_combo.grid(row=3, column=1, sticky=W, pady=6)
        self.role_combo.bind("<<ComboboxSelected>>", lambda e: self._toggle_price())

        ttk.Label(
            frm,
            text="feed:     materia prima externa (entra a costos variables)\n"
                 "product:  producto final (entra a producción anual)\n"
                 "internal: corriente entre bloques (balance de masa)",
            foreground="#888", font=("Segoe UI", 8), justify="left",
        ).grid(row=4, column=0, columnspan=2, sticky=W, pady=(0, 4))

        # ---- precio (solo si feed o product) ----
        self.lbl_price = ttk.Label(frm, text="Precio (USD/tm):")
        self.entry_price = ttk.Entry(frm, justify="right", width=14)
        self.entry_price.insert(0, f"{stream.price_usd_per_tm:g}")
        self.lbl_price.grid(row=5, column=0, sticky=W, pady=6)
        self.entry_price.grid(row=5, column=1, sticky=W, pady=6)
        self._toggle_price()  # mostrar/ocultar según role inicial

        # ---- puertos ISA ----
        ttk.Separator(frm, orient=HORIZONTAL).grid(
            row=6, column=0, columnspan=2, sticky="ew", pady=8)

        ttk.Label(frm, text=f"Puerto en {b_src.name}:").grid(row=7, column=0, sticky=W, pady=4)
        self.src_port_var = StringVar(value=stream.src_port or (ports_src[0] if ports_src else ""))
        self.src_port_combo = ttk.Combobox(
            frm, textvariable=self.src_port_var,
            values=ports_src, state="readonly", width=22,
        )
        self.src_port_combo.grid(row=7, column=1, sticky=W, pady=4)

        ttk.Label(frm, text=f"Puerto en {b_dst.name}:").grid(row=8, column=0, sticky=W, pady=4)
        self.dst_port_var = StringVar(value=stream.dst_port or (ports_dst[0] if ports_dst else ""))
        self.dst_port_combo = ttk.Combobox(
            frm, textvariable=self.dst_port_var,
            values=ports_dst, state="readonly", width=22,
        )
        self.dst_port_combo.grid(row=8, column=1, sticky=W, pady=4)

        # ---- nombre ----
        ttk.Separator(frm, orient=HORIZONTAL).grid(
            row=9, column=0, columnspan=2, sticky="ew", pady=8)

        ttk.Label(frm, text="Nombre:").grid(row=10, column=0, sticky=W, pady=4)
        self.entry_name = ttk.Entry(frm, width=22)
        self.entry_name.insert(0, stream.name)
        self.entry_name.grid(row=10, column=1, sticky=W, pady=4)

        # ---- botones ----
        btns = ttk.Frame(frm)
        btns.grid(row=11, column=0, columnspan=2, pady=(14, 0), sticky=E)
        ttk.Button(btns, text="Cancelar", command=dlg.destroy).pack(side=LEFT, padx=4)
        ttk.Button(btns, text="OK",       command=self._ok).pack(side=LEFT)

        self.entry_m.focus()
        dlg.wait_window()

    def _toggle_price(self):
        """Muestra/oculta el campo Precio según role."""
        role = self.role_var.get()
        if role in ("feed", "product"):
            self.lbl_price.grid()
            self.entry_price.grid()
        else:
            self.lbl_price.grid_remove()
            self.entry_price.grid_remove()

    def _ok(self):
        try:
            m = float(self.entry_m.get())
            if m < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Inválido", "Flujo másico ≥ 0 requerido.")
            return
        role = self.role_var.get() or "internal"
        # precio solo aplica si feed/product
        price = 0.0
        if role in ("feed", "product"):
            try:
                price = float(self.entry_price.get())
                if price < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Inválido", "Precio ≥ 0 requerido.")
                return
        self.stream.mass_flow = m
        self.stream.role = role
        self.stream.price_usd_per_tm = price
        self.stream.src_port = self.src_port_var.get() or ""
        self.stream.dst_port = self.dst_port_var.get() or ""
        nombre = self.entry_name.get().strip() or self.stream.name
        self.stream.name = nombre
        self.result = "ok"
        self.dlg.destroy()


# ======================================================
# VENTANA PRINCIPAL — FlowsheetEditor
# ======================================================

class FlowsheetEditor:
    """Editor de block diagram para estimar ISBL desde un PFD.

    Modos:
      - mode="main": se monta sobre un Tk root (la ventana principal
        de la app).  El botón final es 'Análisis económico →'
        y dispara subprocess.Popen lanzando ANA.py con el ISBL
        ya inyectado.
      - mode="modal": se monta sobre un Toplevel hijo del análisis
        económico existente.  El botón final es 'Apply ISBL' y
        actualiza df_capital in-place (caso legacy: invocado desde
        ANA.py > Tools > Estimate Capital from PFD).
    """

    def __init__(self, container, df_capital=None, on_apply=None, mode="modal"):
        self.container = container
        self.df_capital = df_capital
        self.on_apply = on_apply
        self.mode = mode
        self.win = container        # cualquier widget que acepte .pack como hijo

        self.fs = Flowsheet()
        self.selected_block: Optional[int] = None
        self.selected_stream: Optional[int] = None

        # drag state
        self._drag_block: Optional[int] = None
        self._drag_offx = 0
        self._drag_offy = 0

        # pending connection (after right-click "Connect to...")
        self._connecting_from: Optional[int] = None

        # zoom y pan (modelo en unidades base; canvas se renderiza × zoom)
        self.zoom = 1.0
        self._panning = False

        # título y geometría sólo si el container es ventana
        if hasattr(container, "title"):
            container.title("Diagrama de proceso — diseño en bloques")
        if hasattr(container, "geometry"):
            try:
                container.geometry("1320x780+80+30")
            except Exception:
                pass

        # toolbar arriba
        toolbar = ttk.Frame(container, padding=4)
        toolbar.pack(side=TOP, fill=X)
        ttk.Button(toolbar, text="Nuevo",     command=self.new).pack(side=LEFT, padx=2)
        ttk.Button(toolbar, text="Abrir…",    command=self.open_json).pack(side=LEFT, padx=2)
        ttk.Button(toolbar, text="Guardar…",  command=self.save_json).pack(side=LEFT, padx=2)

        # menu "Ejemplos" para cargar procesos preestablecidos
        examples_btn = ttk.Menubutton(toolbar, text="Ejemplos ▾")
        ex_menu = Menu(examples_btn, tearoff=0)
        ex_menu.add_command(label="HDA — Hidrodealquilación de tolueno",
                            command=lambda: self.load_example("hda"))
        ex_menu.add_command(label="Síntesis de metanol (simplificada)",
                            command=lambda: self.load_example("methanol"))
        ex_menu.add_command(label="Destilación binaria (benceno/tolueno)",
                            command=lambda: self.load_example("distillation"))
        examples_btn.configure(menu=ex_menu)
        examples_btn.pack(side=LEFT, padx=2)

        ttk.Separator(toolbar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=4)
        ttk.Button(toolbar, text="Borrar selección (Supr)",
                   command=self.delete_selected).pack(side=LEFT, padx=2)
        ttk.Separator(toolbar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=4)

        # controles de zoom
        ttk.Button(toolbar, text="−", width=3, command=self.zoom_out).pack(side=LEFT, padx=1)
        self.zoom_label_var = StringVar(value="100%")
        ttk.Button(toolbar, textvariable=self.zoom_label_var, width=6,
                   command=self.zoom_reset).pack(side=LEFT, padx=1)
        ttk.Button(toolbar, text="+", width=3, command=self.zoom_in).pack(side=LEFT, padx=1)
        ttk.Button(toolbar, text="Ajustar", width=8, command=self.zoom_fit).pack(side=LEFT, padx=2)

        ttk.Separator(toolbar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=4)
        ttk.Button(toolbar, text="Calcular", command=self.compute).pack(side=LEFT, padx=2)

        # botón de transición distinto según modo
        if mode == "main":
            ttk.Button(
                toolbar, text="Análisis económico →",
                command=self.launch_analysis,
            ).pack(side=RIGHT, padx=2)
        else:
            ttk.Button(toolbar, text="Aplicar ISBL", command=self.apply_isbl).pack(side=LEFT, padx=2)
            ttk.Button(toolbar, text="Cerrar",       command=self._close).pack(side=RIGHT, padx=2)

        # body
        body = ttk.Frame(container)
        body.pack(side=TOP, fill=BOTH, expand=True)

        # library (left)
        self._build_library(body)

        # canvas (center)
        self._build_canvas(body)

        # properties (right)
        self._build_properties(body)

        # status bar
        self.status_var = StringVar(value="0 equipos · 0 corrientes")
        ttk.Label(
            container, textvariable=self.status_var,
            anchor=W, padding=6,
            relief="sunken",
        ).pack(side=BOTTOM, fill=X)

        # keyboard bindings (bind_all para que funcionen desde cualquier widget)
        container.bind_all("<Delete>",   lambda e: self.delete_selected())
        container.bind_all("<Escape>",   lambda e: self._clear_pending_connection())

        self._draw_grid()
        self._update_status()

    def _close(self):
        try:
            self.win.destroy()
        except Exception:
            pass

    # ==================================================
    # ZOOM Y PAN
    # ==================================================

    def _sc(self, v):
        """Modelo → canvas (multiplicar por zoom)."""
        return v * self.zoom

    def _unsc(self, v):
        """Canvas → modelo."""
        return v / self.zoom

    def _model_xy(self, event):
        """Coord del modelo a partir del event del mouse,
        respetando scroll del canvas y zoom."""
        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)
        return self._unsc(cx), self._unsc(cy)

    def _refresh_zoom_label(self):
        self.zoom_label_var.set(f"{self.zoom * 100:.0f}%")

    def _redraw_all(self):
        """Borra todo y vuelve a dibujar desde el modelo.
        Se usa al cambiar el zoom."""
        # preservar selección
        sel_b = self.selected_block
        sel_s = self.selected_stream

        self.canvas.delete("all")
        self._draw_grid()
        for b in self.fs.blocks.values():
            self._render_block(b)
        for s in self.fs.streams.values():
            self._render_stream(s)

        # reanchear scrollregion al contenido + margen
        max_x = max((b.x + BLOCK_W for b in self.fs.blocks.values()),
                    default=CANVAS_W_LOG)
        max_y = max((b.y + BLOCK_H for b in self.fs.blocks.values()),
                    default=CANVAS_H_LOG)
        sr_w = max(self._sc(max_x) + 200, 800)
        sr_h = max(self._sc(max_y) + 200, 600)
        self.canvas.configure(scrollregion=(0, 0, sr_w, sr_h))

        # reaplicar selección visual
        if sel_b is not None and sel_b in self.fs.blocks:
            self._select_block(sel_b)
        elif sel_s is not None and sel_s in self.fs.streams:
            self._select_stream(sel_s)

        self._refresh_zoom_label()

    def _zoom_at(self, view_x, view_y, factor):
        """Zoom anclado al punto (view_x, view_y) del widget.
        El punto bajo el cursor queda quieto en pantalla."""
        new_zoom = self.zoom * factor
        new_zoom = max(ZOOM_MIN, min(ZOOM_MAX, new_zoom))
        if abs(new_zoom - self.zoom) < 1e-6:
            return

        # punto del modelo bajo el cursor (antes del zoom)
        cx_before = self.canvas.canvasx(view_x)
        cy_before = self.canvas.canvasy(view_y)
        mx = cx_before / self.zoom
        my = cy_before / self.zoom

        self.zoom = new_zoom
        self._redraw_all()

        # mover scroll para que (mx, my) quede en (view_x, view_y)
        cx_after = mx * self.zoom
        cy_after = my * self.zoom
        sr = self.canvas.cget("scrollregion").split()
        try:
            sr_w = float(sr[2])
            sr_h = float(sr[3])
        except (IndexError, ValueError):
            return
        self.canvas.xview_moveto(max(0.0, (cx_after - view_x) / sr_w))
        self.canvas.yview_moveto(max(0.0, (cy_after - view_y) / sr_h))

    def zoom_in(self):
        w = self.canvas.winfo_width()  or 800
        h = self.canvas.winfo_height() or 600
        self._zoom_at(w / 2, h / 2, ZOOM_STEP)

    def zoom_out(self):
        w = self.canvas.winfo_width()  or 800
        h = self.canvas.winfo_height() or 600
        self._zoom_at(w / 2, h / 2, 1 / ZOOM_STEP)

    def zoom_reset(self):
        self.zoom = 1.0
        self._redraw_all()
        self.canvas.xview_moveto(0)
        self.canvas.yview_moveto(0)

    def zoom_fit(self):
        """Ajusta zoom para que todos los bloques entren en la vista."""
        if not self.fs.blocks:
            self.zoom_reset()
            return

        xs = [b.x for b in self.fs.blocks.values()]
        ys = [b.y for b in self.fs.blocks.values()]
        min_x, max_x = min(xs), max(xs) + BLOCK_W
        min_y, max_y = min(ys), max(ys) + BLOCK_H

        margin = 60
        w = (self.canvas.winfo_width()  or 800) - 2 * margin
        h = (self.canvas.winfo_height() or 600) - 2 * margin
        if w <= 0 or h <= 0:
            return

        fx = w / (max_x - min_x + 1)
        fy = h / (max_y - min_y + 1)
        self.zoom = max(ZOOM_MIN, min(ZOOM_MAX, min(fx, fy)))
        self._redraw_all()

        # centrar el contenido en la vista
        sr = self.canvas.cget("scrollregion").split()
        try:
            sr_w = float(sr[2])
            sr_h = float(sr[3])
        except (IndexError, ValueError):
            return
        cx = self._sc((min_x + max_x) / 2)
        cy = self._sc((min_y + max_y) / 2)
        view_w = self.canvas.winfo_width()  or 800
        view_h = self.canvas.winfo_height() or 600
        self.canvas.xview_moveto(max(0.0, (cx - view_w / 2) / sr_w))
        self.canvas.yview_moveto(max(0.0, (cy - view_h / 2) / sr_h))

    def _on_mousewheel_zoom(self, event):
        # Windows: event.delta múltiplo de 120
        factor = ZOOM_STEP if event.delta > 0 else 1 / ZOOM_STEP
        self._zoom_at(event.x, event.y, factor)

    def _on_mousewheel_scroll(self, event):
        self.canvas.yview_scroll(-int(event.delta / 120) or (-1 if event.delta > 0 else 1), "units")

    def _on_mousewheel_hscroll(self, event):
        self.canvas.xview_scroll(-int(event.delta / 120) or (-1 if event.delta > 0 else 1), "units")

    def _on_pan_start(self, event):
        self.canvas.scan_mark(event.x, event.y)
        self.canvas.configure(cursor="fleur")

    def _on_pan_drag(self, event):
        self.canvas.scan_dragto(event.x, event.y, gain=1)

    def _on_pan_end(self, _event):
        self.canvas.configure(cursor="")

    # ==================================================
    # PANELES
    # ==================================================

    def _build_library(self, body):
        frame = ttk.LabelFrame(body, text=" Biblioteca de equipos ", padding=6, width=240)
        frame.pack(side=LEFT, fill=Y, padx=(8, 4), pady=8)
        frame.pack_propagate(False)

        ttk.Label(
            frame,
            text="Seleccioná un equipo y apretá 'Agregar'.\n"
                 "Doble-click en un bloque para editar S/n.\n"
                 "Click derecho en un bloque → Conectar a…",
            foreground="#666", font=("Segoe UI", 8),
            justify="left", wraplength=220,
        ).pack(fill=X, pady=(0, 6))

        # tree con categorías
        tree = ttk.Treeview(frame, show="tree", selectmode="browse", height=18)
        cats = eq.por_categoria()
        for cat, names in cats.items():
            parent = tree.insert("", END, text=cat, open=True, tags=("cat",))
            for nombre in names:
                tree.insert(parent, END, text=nombre, tags=("eq",))
        tree.pack(fill=BOTH, expand=True)
        self.lib_tree = tree

        ttk.Button(frame, text="+ Agregar al diagrama",
                   command=self.add_selected_equipment).pack(fill=X, pady=(6, 0))

        # doble-click también agrega
        tree.bind("<Double-1>", lambda e: self.add_selected_equipment())

    def _build_canvas(self, body):
        center = ttk.Frame(body)
        center.pack(side=LEFT, fill=BOTH, expand=True, padx=4, pady=8)

        # grid layout para canvas + scrollbars
        center.rowconfigure(0, weight=1)
        center.columnconfigure(0, weight=1)

        canvas = Canvas(
            center, bg=CANVAS_BG, highlightthickness=1,
            highlightbackground="#cccccc",
            scrollregion=(0, 0, CANVAS_W_LOG, CANVAS_H_LOG),
        )
        canvas.grid(row=0, column=0, sticky="nsew")
        self.canvas = canvas

        hbar = Scrollbar(center, orient=HORIZONTAL, command=canvas.xview)
        hbar.grid(row=1, column=0, sticky="ew")
        vbar = Scrollbar(center, orient=VERTICAL,   command=canvas.yview)
        vbar.grid(row=0, column=1, sticky="ns")
        canvas.configure(xscrollcommand=hbar.set, yscrollcommand=vbar.set)

        # mouse: izquierdo selecciona/arrastra bloque
        canvas.bind("<Button-1>",       self._on_left_click)
        canvas.bind("<B1-Motion>",      self._on_left_drag)
        canvas.bind("<ButtonRelease-1>",self._on_left_release)
        canvas.bind("<Button-3>",       self._on_right_click)
        canvas.bind("<Double-1>",       self._on_double_click)

        # mouse middle (Button-2): pan
        canvas.bind("<ButtonPress-2>",   self._on_pan_start)
        canvas.bind("<B2-Motion>",       self._on_pan_drag)
        canvas.bind("<ButtonRelease-2>", self._on_pan_end)

        # Space + drag izquierdo también funciona como pan (laptops sin
        # middle button). Mantengo state en self._panning.
        canvas.bind("<KeyPress-space>",   lambda e: setattr(self, "_panning", True))
        canvas.bind("<KeyRelease-space>", lambda e: setattr(self, "_panning", False))
        canvas.focus_set()

        # rueda: zoom con Ctrl, scroll vertical sin Ctrl
        # Linux usa Button-4 / Button-5, Windows/Mac usa MouseWheel
        canvas.bind("<Control-MouseWheel>", self._on_mousewheel_zoom)
        canvas.bind("<Control-Button-4>",   lambda e: self._zoom_at(e.x, e.y, ZOOM_STEP))
        canvas.bind("<Control-Button-5>",   lambda e: self._zoom_at(e.x, e.y, 1 / ZOOM_STEP))
        canvas.bind("<MouseWheel>",         self._on_mousewheel_scroll)
        canvas.bind("<Button-4>",           lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind("<Button-5>",           lambda e: canvas.yview_scroll(1, "units"))
        canvas.bind("<Shift-MouseWheel>",   self._on_mousewheel_hscroll)
        canvas.bind("<Shift-Button-4>",     lambda e: canvas.xview_scroll(-1, "units"))
        canvas.bind("<Shift-Button-5>",     lambda e: canvas.xview_scroll(1, "units"))

    def _build_properties(self, body):
        frame = ttk.LabelFrame(body, text=" Propiedades ", padding=6, width=300)
        frame.pack(side=LEFT, fill=Y, padx=(4, 8), pady=8)
        frame.pack_propagate(False)

        self.prop_var = StringVar(value="(nada seleccionado)")
        self.prop_label = ttk.Label(
            frame, textvariable=self.prop_var,
            justify="left", wraplength=280, anchor="nw",
            font=("Consolas", 9),
        )
        self.prop_label.pack(fill=BOTH, expand=True)

        sep = ttk.Separator(frame, orient=HORIZONTAL)
        sep.pack(fill=X, pady=8)

        self.results_var = StringVar(value="(apretá Calcular para estimar el ISBL)")
        ttk.Label(
            frame, textvariable=self.results_var,
            justify="left", wraplength=280, anchor="nw",
            font=("Consolas", 9), foreground="#1565c0",
        ).pack(fill=X)

    # ==================================================
    # ACCIONES
    # ==================================================

    def new(self):
        if self.fs.blocks and not messagebox.askyesno(
            "Nuevo diagrama", "¿Descartar el diagrama actual y empezar vacío?"
        ):
            return
        self.fs = Flowsheet()
        self.canvas.delete("all")
        self._draw_grid()
        self.selected_block = None
        self.selected_stream = None
        self._update_status()
        self._update_properties()
        self.results_var.set("(apretá Calcular para estimar el ISBL)")

    # ==================================================
    # EXAMPLES — procesos preestablecidos
    # ==================================================

    def load_example(self, key):
        """Carga un proceso preestablecido (sobrescribe el actual).
        Sirve como smoke test del software end-to-end:
        equipos + corrientes con feeds/products + Calcular + ISBL.
        """
        if self.fs.blocks and not messagebox.askyesno(
            "Cargar ejemplo",
            "Esto va a reemplazar el diagrama actual. ¿Continuar?",
        ):
            return

        builders = {
            "hda":          self._example_hda,
            "methanol":     self._example_methanol,
            "distillation": self._example_distillation,
        }
        builder = builders.get(key)
        if builder is None:
            return

        self.fs = Flowsheet()
        builder()
        self._redraw_all()
        self.zoom_fit()
        self.selected_block = None
        self.selected_stream = None
        self._update_status()
        self._update_properties()
        self.results_var.set("(apretá Calcular para estimar el ISBL)")

    def _add_example_block(self, name, eq_type, S, x, y, n=1):
        bid = self.fs.new_id()
        b = Block(id=bid, name=name, eq_type=eq_type, S=S, n=n, x=x, y=y)
        self.fs.blocks[bid] = b
        return bid

    def _add_example_stream(self, src, dst, name, mass_flow=0.0,
                            role="internal", src_port="", dst_port="",
                            price=0.0):
        sid = self.fs.new_id()
        s = Stream(
            id=sid, name=name, src=src, dst=dst,
            mass_flow=mass_flow, role=role,
            src_port=src_port, dst_port=dst_port,
            price_usd_per_tm=price,
        )
        self.fs.streams[sid] = s
        return sid

    def _example_hda(self):
        """HDA — Hidrodealquilación de tolueno (Douglas / Turton).
        Tolueno + H2 → Benceno + CH4. ~75% conversión.
        Tags ISA-5.1: E (HX), F (horno), R (reactor),
        V (separador), T (columna), P (bomba), TK (tanque)."""
        cx = [120, 360, 600, 840, 1080, 1320]
        y_top = 240
        y_bot = 460

        # tren de reacción (fila superior)
        e101 = self._add_example_block("E-101", "Heat exch. — floating head", 250.0, cx[0], y_top)
        f101 = self._add_example_block("F-101", "Fired heater — non-reformer", 5000.0, cx[1], y_top)
        r101 = self._add_example_block("R-101", "Reactor — jacketed non-agit.",  25.0, cx[2], y_top)
        e102 = self._add_example_block("E-102", "Heat exch. — air cooler",      180.0, cx[3], y_top)
        v101 = self._add_example_block("V-101", "Vessel — vertical",             20.0, cx[4], y_top)

        # separación (fila inferior)
        t101 = self._add_example_block("T-101", "Tower (column shell)",          45.0, cx[2], y_bot)
        e104 = self._add_example_block("E-104", "Heat exch. — kettle reboiler", 150.0, cx[3], y_bot)
        e103 = self._add_example_block("E-103", "Heat exch. — floating head",   160.0, cx[1], y_bot)
        p101 = self._add_example_block("P-101", "Pump — centrifugal",            15.0, cx[0], y_bot)
        tk1  = self._add_example_block("TK-101","Storage tank — cone roof",     500.0, cx[5], y_bot)
        tk2  = self._add_example_block("TK-102","Storage tank — cone roof",     100.0, cx[5], y_top)

        # corrientes (puertos específicos por tipo de equipo)
        # Precios USD/tm (referencia mercado 2024):
        #   tolueno feed  ~650, benceno ~1050, H2 ~2000
        self._add_example_stream(p101, e101, "S-1",  11000, role="feed",
                                 src_port="descarga",  dst_port="tube_in",
                                 price=650.0)
        self._add_example_stream(e101, f101, "S-2",
                                 src_port="tube_out",  dst_port="proceso_in")
        self._add_example_stream(f101, r101, "S-3",
                                 src_port="proceso_out", dst_port="alimentacion")
        self._add_example_stream(r101, e102, "S-4",
                                 src_port="producto",  dst_port="proceso_in")
        self._add_example_stream(e102, v101, "S-5",
                                 src_port="proceso_out", dst_port="alimentacion")
        self._add_example_stream(v101, t101, "S-6",
                                 src_port="liquido",   dst_port="alimentacion")
        self._add_example_stream(t101, e104, "S-7",
                                 src_port="liquido_fondo", dst_port="liq_in")
        self._add_example_stream(t101, e103, "S-8",
                                 src_port="vapor_tope", dst_port="tube_in")
        self._add_example_stream(e103, p101, "S-9-recic",   # reciclo
                                 src_port="shell_out", dst_port="succion")
        self._add_example_stream(e103, tk1,  "S-benceno", 8500, role="product",
                                 src_port="tube_out",  dst_port="entrada",
                                 price=1050.0)
        self._add_example_stream(v101, tk2,  "S-purga-H2", 350, role="product",
                                 src_port="vapor",     dst_port="entrada",
                                 price=2000.0)

    def _example_methanol(self):
        """Síntesis de metanol simplificada.
        Syngas (CO + 2H2) → CH3OH. Reactor + flash + columna.

        Layout: 2 filas alineadas para que el routing salga limpio.
          fila 1: K-101 → E-101 → R-101 → E-102 → V-101
          fila 2: TK-MeOH ← E-103 ← T-101 → E-104 → TK-agua
        El destilado va a TK-MeOH (izquierda), el agua de fondo va a TK-agua (derecha),
        sin que ninguna corriente cruce equipos.
        """
        cx = [80, 280, 480, 680, 880]
        y_top = 220
        y_bot = 500

        k101 = self._add_example_block("K-101", "Compressor — centrifugal",      800.0, cx[0], y_top)
        e101 = self._add_example_block("E-101", "Heat exch. — floating head",    220.0, cx[1], y_top)
        r101 = self._add_example_block("R-101", "Reactor — jacketed non-agit.",   30.0, cx[2], y_top)
        e102 = self._add_example_block("E-102", "Heat exch. — air cooler",       200.0, cx[3], y_top)
        v101 = self._add_example_block("V-101", "Vessel — vertical",              15.0, cx[4], y_top)

        # fila inferior: tanque-MeOH y tanque-agua en los extremos
        tk1  = self._add_example_block("TK-101","Storage tank — floating roof",  400.0, cx[0], y_bot)
        e103 = self._add_example_block("E-103", "Heat exch. — floating head",    140.0, cx[1], y_bot)
        t101 = self._add_example_block("T-101", "Tower (column shell)",           35.0, cx[2], y_bot)
        e104 = self._add_example_block("E-104", "Heat exch. — kettle reboiler",  130.0, cx[3], y_bot)
        tk2  = self._add_example_block("TK-102","Storage tank — cone roof",       50.0, cx[4], y_bot)

        # Precios USD/tm (referencia mercado 2024):
        #   syngas ~150, metanol ~430, agua de proceso ~5
        self._add_example_stream(k101, e101, "S-1", 14000, role="feed",
                                 src_port="descarga", dst_port="tube_in",
                                 price=150.0)
        self._add_example_stream(e101, r101, "S-2",
                                 src_port="tube_out", dst_port="alimentacion")
        self._add_example_stream(r101, e102, "S-3",
                                 src_port="producto", dst_port="proceso_in")
        self._add_example_stream(e102, v101, "S-4",
                                 src_port="proceso_out", dst_port="alimentacion")
        self._add_example_stream(v101, t101, "S-5",
                                 src_port="liquido",  dst_port="alimentacion")
        # destilado por el top de T-101 → condensador E-103 → tanque MeOH (izq)
        self._add_example_stream(t101, e103, "S-vap-tope",
                                 src_port="vapor_tope", dst_port="shell_in")
        self._add_example_stream(e103, tk1,  "S-MeOH", 9500, role="product",
                                 src_port="shell_out", dst_port="entrada",
                                 price=430.0)
        # fondo de T-101 → reboiler E-104 → tanque agua (der)
        self._add_example_stream(t101, e104, "S-fondo",
                                 src_port="liquido_fondo", dst_port="liq_in")
        self._add_example_stream(e104, tk2,  "S-agua", 600, role="product",
                                 src_port="cond_out", dst_port="entrada",
                                 price=5.0)

    def _example_distillation(self):
        """Destilación binaria benceno/tolueno (50/50).
        Pre-calentador + columna + condensador + reboiler + bombas."""
        cx = [120, 360, 600, 840, 1080]
        y_top = 220
        y_mid = 380
        y_bot = 540

        tk0  = self._add_example_block("TK-101","Storage tank — cone roof", 200.0, cx[0], y_mid)
        p101 = self._add_example_block("P-101", "Pump — centrifugal",         8.0, cx[1], y_mid)
        e101 = self._add_example_block("E-101", "Heat exch. — floating head",120.0, cx[2], y_mid)
        t101 = self._add_example_block("T-101", "Tower (column shell)",       20.0, cx[3], y_mid)

        e102 = self._add_example_block("E-102", "Heat exch. — floating head", 90.0, cx[3], y_top)
        e103 = self._add_example_block("E-103", "Heat exch. — kettle reboiler", 85.0, cx[3], y_bot)

        tk1  = self._add_example_block("TK-102","Storage tank — cone roof",  150.0, cx[4], y_top)
        tk2  = self._add_example_block("TK-103","Storage tank — cone roof",  150.0, cx[4], y_bot)

        # Precios USD/tm (referencia mercado 2024):
        #   mezcla bz/tol ~850, benceno ~1050, tolueno ~700
        self._add_example_stream(tk0,  p101, "S-1", 10000, role="feed",
                                 src_port="salida",   dst_port="succion",
                                 price=850.0)
        self._add_example_stream(p101, e101, "S-2",
                                 src_port="descarga", dst_port="tube_in")
        self._add_example_stream(e101, t101, "S-3",
                                 src_port="tube_out", dst_port="alimentacion")
        self._add_example_stream(t101, e102, "S-4",
                                 src_port="vapor_tope",     dst_port="tube_in")
        self._add_example_stream(t101, e103, "S-5",
                                 src_port="liquido_fondo",  dst_port="liq_in")
        self._add_example_stream(e102, tk1,  "S-benceno", 5000, role="product",
                                 src_port="tube_out", dst_port="entrada",
                                 price=1050.0)
        self._add_example_stream(e103, tk2,  "S-tolueno", 5000, role="product",
                                 src_port="vap_out",  dst_port="entrada",
                                 price=700.0)

    def open_json(self):
        path = filedialog.askopenfilename(
            title="Abrir diagrama",
            filetypes=[("JSON", "*.json")],
        )
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.fs = Flowsheet.from_dict(data)
        except Exception as e:
            messagebox.showerror("Error al abrir", f"{type(e).__name__}: {e}")
            return
        self.selected_block = None
        self.selected_stream = None
        self._redraw_all()
        self.zoom_fit()
        self._update_status()
        self._update_properties()

    def save_json(self):
        path = filedialog.asksaveasfilename(
            title="Guardar diagrama",
            defaultextension=".json",
            filetypes=[("JSON", "*.json")],
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.fs.to_dict(), f, indent=2)
        except Exception as e:
            messagebox.showerror("Error al guardar", f"{type(e).__name__}: {e}")
            return
        messagebox.showinfo("Guardado", f"Diagrama guardado en:\n{path}")

    def delete_selected(self):
        if self.selected_block is not None:
            self._delete_block(self.selected_block)
            self.selected_block = None
        elif self.selected_stream is not None:
            self._delete_stream(self.selected_stream)
            self.selected_stream = None
        self._update_status()
        self._update_properties()

    def add_selected_equipment(self):
        sel = self.lib_tree.selection()
        if not sel:
            messagebox.showinfo("Seleccioná un equipo",
                                "Hacé click sobre un equipo de la biblioteca.")
            return
        item = sel[0]
        tags = self.lib_tree.item(item, "tags")
        if "eq" not in tags:
            messagebox.showinfo("Elegí un equipo",
                                "Seleccioná un equipo, no una categoría.")
            return
        eq_type = self.lib_tree.item(item, "text")
        self._add_block(eq_type)

    # ==================================================
    # BLOCKS
    # ==================================================

    def _add_block(self, eq_type, x=None, y=None):
        spec = eq.EQUIPMENT_DATA.get(eq_type)
        if not spec:
            return

        bid = self.fs.new_id()
        S_default = (spec["S_min"] + spec["S_max"]) / 2

        # nombre ISA-5.1: prefix por tipo + autoincrement desde 101
        nombre = ep.next_block_name(
            eq_type, [b.name for b in self.fs.blocks.values()]
        )

        if x is None:
            # ubicar en el centro del viewport, evitando overlaps
            x = 200 + (len(self.fs.blocks) % 6) * 160
            y = 100 + ((len(self.fs.blocks) // 6) % 6) * 100

        b = Block(id=bid, name=nombre, eq_type=eq_type, S=S_default, n=1, x=x, y=y)
        self.fs.blocks[bid] = b
        self._render_block(b)
        self._select_block(bid)
        self._update_status()

    def _render_block(self, b):
        # texto principal + sub (S y eq type)
        spec = eq.EQUIPMENT_DATA.get(b.eq_type, {})
        unit = spec.get("S_unit", "")
        cat = spec.get("categoria", "")

        x0 = self._sc(b.x)
        y0 = self._sc(b.y)
        x1 = self._sc(b.x + BLOCK_W)
        y1 = self._sc(b.y + BLOCK_H)

        font_title = ("Segoe UI", max(7, int(10 * self.zoom)), "bold")
        font_sub   = ("Segoe UI", max(6, int(8  * self.zoom)))

        b.canvas_rect = self.canvas.create_rectangle(
            x0, y0, x1, y1,
            fill=BLOCK_FILL, outline=BLOCK_BORDER, width=2,
            tags=("block", f"b{b.id}"),
        )
        b.canvas_text = self.canvas.create_text(
            (x0 + x1) / 2, y0 + self._sc(18),
            text=b.name, fill=BLOCK_TEXT, font=font_title,
            tags=("block", f"b{b.id}"),
        )
        b.canvas_sub = self.canvas.create_text(
            (x0 + x1) / 2, y0 + self._sc(40),
            text=f"{cat}\nS = {b.S:g} {unit}" + (f"  × {b.n}" if b.n > 1 else ""),
            fill=BLOCK_SUB, font=font_sub, justify="center",
            tags=("block", f"b{b.id}"),
        )

        # port markers: pequeños círculos en cada puerto del eq_type
        ports = ep.get_ports(b.eq_type)
        used_out = {t.src_port for t in self.fs.streams.values()
                    if t.src == b.id and t.src_port}
        used_in  = {t.dst_port for t in self.fs.streams.values()
                    if t.dst == b.id and t.dst_port}
        r = max(2.5, 4 * self.zoom)
        for pname, (side, frac) in ports.items():
            if side == "right":
                px, py = b.x + BLOCK_W,         b.y + BLOCK_H * frac
            elif side == "left":
                px, py = b.x,                   b.y + BLOCK_H * frac
            elif side == "top":
                px, py = b.x + BLOCK_W * frac, b.y
            else:
                px, py = b.x + BLOCK_W * frac, b.y + BLOCK_H
            cx, cy = self._sc(px), self._sc(py)
            connected = pname in used_out or pname in used_in
            color = "#1565c0" if connected else "#bbbbbb"
            self.canvas.create_oval(
                cx - r, cy - r, cx + r, cy + r,
                fill=color, outline="#333", width=1,
                tags=("port", f"port_b{b.id}", f"port_b{b.id}_{pname}"),
            )

    def _refresh_block_text(self, b):
        spec = eq.EQUIPMENT_DATA.get(b.eq_type, {})
        unit = spec.get("S_unit", "")
        cat = spec.get("categoria", "")
        self.canvas.itemconfigure(b.canvas_text, text=b.name)
        self.canvas.itemconfigure(
            b.canvas_sub,
            text=f"{cat}\nS = {b.S:g} {unit}" + (f"  × {b.n}" if b.n > 1 else ""),
        )

    def _delete_block(self, bid):
        b = self.fs.blocks.pop(bid, None)
        if b is None:
            return
        # remover sus streams
        for sid in [sid for sid, s in self.fs.streams.items()
                    if s.src == bid or s.dst == bid]:
            self._delete_stream(sid)
        for cid in (b.canvas_rect, b.canvas_text, b.canvas_sub):
            if cid is not None:
                self.canvas.delete(cid)
        # port markers (creados con tag port_b{id})
        self.canvas.delete(f"port_b{bid}")

    def _select_block(self, bid):
        # deselect anteriores
        self._deselect_all_visual()
        self.selected_stream = None
        self.selected_block = bid
        b = self.fs.blocks.get(bid)
        if b and b.canvas_rect is not None:
            self.canvas.itemconfigure(b.canvas_rect,
                                      outline=BLOCK_BORDER_SEL, width=3)
        self._update_properties()

    def _deselect_all_visual(self):
        for b in self.fs.blocks.values():
            if b.canvas_rect is not None:
                self.canvas.itemconfigure(b.canvas_rect,
                                          outline=BLOCK_BORDER, width=2)
        for s in self.fs.streams.values():
            if s.canvas_line is not None:
                color = STREAM_ROLE_COLORS.get(s.role, STREAM_COLOR)
                self.canvas.itemconfigure(s.canvas_line, fill=color, width=2)

    # ==================================================
    # STREAMS
    # ==================================================

    def _add_stream(self, src_id, dst_id, src_port="", dst_port=""):
        if src_id == dst_id:
            return
        # evitar duplicado exacto (mismos bloques + mismos puertos)
        for s in self.fs.streams.values():
            if (s.src == src_id and s.dst == dst_id
                and s.src_port == src_port and s.dst_port == dst_port):
                messagebox.showinfo(
                    "Conexión existente",
                    f"La corriente {s.name} ya conecta esos puertos.",
                )
                return

        b_src = self.fs.blocks.get(src_id)
        b_dst = self.fs.blocks.get(dst_id)

        # autoseleccionar puertos si no se especificaron
        if not src_port and b_src is not None:
            used_out = [t.src_port for t in self.fs.streams.values()
                        if t.src == src_id and t.src_port]
            src_port = ep.autoselect_outlet(b_src.eq_type, used_out)
        if not dst_port and b_dst is not None:
            used_in = [t.dst_port for t in self.fs.streams.values()
                       if t.dst == dst_id and t.dst_port]
            dst_port = ep.autoselect_inlet(b_dst.eq_type, used_in)

        sid = self.fs.new_id()
        nombre = f"S-{len(self.fs.streams) + 1}"
        s = Stream(
            id=sid, name=nombre, src=src_id, dst=dst_id,
            mass_flow=0.0, src_port=src_port, dst_port=dst_port,
        )
        self.fs.streams[sid] = s
        self._render_stream(s)
        self._update_status()

    # ---------- puertos específicos por equipo ----------
    # Cada eq_type tiene un catálogo de puertos en equipment_ports.
    # Si el stream tiene src_port/dst_port nombrados, se usan;
    # si están vacíos, se autoselecciona el primer puerto libre del
    # lado convencional (right para outlet, left para inlet).

    def _resolve_port(self, b, port_name, default_side):
        """Resuelve (port_name, side, x, y) para un bloque y un puerto.
        Si port_name no existe en el catálogo, autoselecciona uno del
        default_side ('right' para salidas, 'left' para entradas)."""
        ports = ep.get_ports(b.eq_type)
        if port_name and port_name in ports:
            side, frac = ports[port_name]
        else:
            # fallback: autoseleccionar uno del default_side
            chosen = None
            for pname, (side, frac) in ports.items():
                if side == default_side:
                    chosen = (pname, side, frac)
                    break
            if chosen is None:
                pname = next(iter(ports))
                side, frac = ports[pname]
            else:
                pname, side, frac = chosen
            port_name = pname
        # coord del puerto en el modelo
        if side == "right":
            x, y = b.x + BLOCK_W,             b.y + BLOCK_H * frac
        elif side == "left":
            x, y = b.x,                       b.y + BLOCK_H * frac
        elif side == "top":
            x, y = b.x + BLOCK_W * frac,     b.y
        elif side == "bottom":
            x, y = b.x + BLOCK_W * frac,     b.y + BLOCK_H
        else:
            x, y = b.x, b.y
        return port_name, side, x, y

    @staticmethod
    def _side_dir(side):
        """Vector unitario hacia afuera del bloque para un lado."""
        return {"right": (1, 0), "left": (-1, 0),
                "top":   (0, -1), "bottom": (0, 1)}.get(side, (1, 0))

    # ---------- detección de colisión con bloques ----------

    @staticmethod
    def _seg_intersects_rect(x1, y1, x2, y2, rx1, ry1, rx2, ry2):
        """Chequea si el segmento (x1,y1)-(x2,y2) intersecta el
        rect (rx1,ry1)-(rx2,ry2).  Asume segmento ortogonal."""
        if abs(x1 - x2) < 1e-6:        # vertical
            if not (rx1 <= x1 <= rx2):
                return False
            ymin, ymax = (y1, y2) if y1 < y2 else (y2, y1)
            return not (ymax < ry1 or ymin > ry2)
        if abs(y1 - y2) < 1e-6:        # horizontal
            if not (ry1 <= y1 <= ry2):
                return False
            xmin, xmax = (x1, x2) if x1 < x2 else (x2, x1)
            return not (xmax < rx1 or xmin > rx2)
        return False                   # no ortogonal: skip

    def _path_crosses_others(self, pts, exclude_ids, shrink=4):
        """¿El polyline cruza algún bloque que NO sea src/dst?
        shrink: encoger el rect para tolerar pasar por el borde."""
        for i in range(0, len(pts) - 2, 2):
            x1, y1, x2, y2 = pts[i], pts[i+1], pts[i+2], pts[i+3]
            for b in self.fs.blocks.values():
                if b.id in exclude_ids:
                    continue
                rx1 = b.x + shrink
                ry1 = b.y + shrink
                rx2 = b.x + BLOCK_W - shrink
                ry2 = b.y + BLOCK_H - shrink
                if self._seg_intersects_rect(x1, y1, x2, y2, rx1, ry1, rx2, ry2):
                    return True
        return False

    # ---------- routing ortogonal ----------

    def _z_route(self, x1, y1, side1, x2, y2, side2, mid):
        """Z-shape ortogonal con el codo en `mid` (x si laterales,
        y si verticales).  Soporta cualquier orientación."""
        gap = ROUTING_GAP
        dx1, dy1 = self._side_dir(side1)
        dx2, dy2 = self._side_dir(side2)
        ex1, ey1 = x1 + dx1 * gap, y1 + dy1 * gap
        ex2, ey2 = x2 + dx2 * gap, y2 + dy2 * gap

        h_sides = ("left", "right")
        v_sides = ("top", "bottom")

        if side1 in h_sides and side2 in h_sides:
            mx = mid
            return [x1, y1, ex1, ey1, mx, ey1, mx, ey2, ex2, ey2, x2, y2]
        if side1 in v_sides and side2 in v_sides:
            my = mid
            return [x1, y1, ex1, ey1, ex1, my, ex2, my, ex2, ey2, x2, y2]
        # perpendiculares: ignoramos mid, L-shape
        if side1 in h_sides:
            return [x1, y1, ex1, ey1, ex2, ey1, ex2, ey2, x2, y2]
        return [x1, y1, ex1, ey1, ex1, ey2, ex2, ey2, x2, y2]

    def _u_route(self, x1, y1, side1, x2, y2, side2, far):
        """U-shape: salida y entrada por el mismo lado, rodeando
        en `far` (x si horizontales, y si verticales)."""
        gap = ROUTING_GAP
        dx1, dy1 = self._side_dir(side1)
        dx2, dy2 = self._side_dir(side2)
        ex1, ey1 = x1 + dx1 * gap, y1 + dy1 * gap
        ex2, ey2 = x2 + dx2 * gap, y2 + dy2 * gap

        if side1 in ("left", "right"):
            return [x1, y1, far, ey1, far, ey2, x2, y2]
        return [x1, y1, ex1, far, ex2, far, x2, y2]

    def _layout_bounds(self):
        """(ymin, ymax, xmin, xmax) del conjunto de bloques."""
        if not self.fs.blocks:
            return 0, 0, 0, 0
        ys1 = [b.y                for b in self.fs.blocks.values()]
        ys2 = [b.y + BLOCK_H     for b in self.fs.blocks.values()]
        xs1 = [b.x                for b in self.fs.blocks.values()]
        xs2 = [b.x + BLOCK_W     for b in self.fs.blocks.values()]
        return min(ys1), max(ys2), min(xs1), max(xs2)

    def _ortho_route(self, x1, y1, side1, x2, y2, side2,
                     src_id=None, dst_id=None):
        """Polyline ortogonal entre dos puertos con avoid-collision:
        prueba varias variantes y devuelve la primera que no cruza
        bloques externos.  Si todas cruzan, devuelve la primaria."""
        gap = ROUTING_GAP
        dx1, dy1 = self._side_dir(side1)
        dx2, dy2 = self._side_dir(side2)
        ex1, ey1 = x1 + dx1 * gap, y1 + dy1 * gap
        ex2, ey2 = x2 + dx2 * gap, y2 + dy2 * gap

        h_sides = ("left", "right")
        v_sides = ("top", "bottom")
        ymin, ymax, xmin, xmax = self._layout_bounds()

        # ---- generar candidatos en orden de preferencia ----
        candidates = []

        # ambos horizontales (opuestos: right↔left)
        if side1 in h_sides and side2 in h_sides:
            same_side = (side1 == side2)
            if not same_side and (
                (side1 == "right" and ex2 >= ex1) or
                (side1 == "left"  and ex1 >= ex2)
            ):
                if abs(ey1 - ey2) < 2:
                    candidates.append([x1, y1, x2, y2])  # recta
                # Z con codo en midpoint x
                candidates.append(self._z_route(x1, y1, side1, x2, y2, side2,
                                                (ex1 + ex2) / 2))
            # variantes por encima y por debajo del layout
            corridor_above = ymin - 40
            corridor_below = ymax + 40
            # Z con codo a "corridor" y → en realidad usamos U que rodea
            # de side1 al side2 pasando por (y = corridor)
            far_x_right = max(ex1, ex2) + gap
            far_x_left  = min(ex1, ex2) - gap
            if side1 == "right":
                # ir derecha → bajar/subir a corridor → entrar dst
                candidates.append([x1, y1, ex1, ey1,
                                   ex1, corridor_above,
                                   ex2, corridor_above,
                                   ex2, ey2, x2, y2])
                candidates.append([x1, y1, ex1, ey1,
                                   ex1, corridor_below,
                                   ex2, corridor_below,
                                   ex2, ey2, x2, y2])
            else:  # both "left" o mixed
                candidates.append([x1, y1, ex1, ey1,
                                   ex1, corridor_above,
                                   ex2, corridor_above,
                                   ex2, ey2, x2, y2])
                candidates.append([x1, y1, ex1, ey1,
                                   ex1, corridor_below,
                                   ex2, corridor_below,
                                   ex2, ey2, x2, y2])
            if same_side:
                candidates.append(self._u_route(x1, y1, side1, x2, y2, side2,
                                                far_x_right if side1 == "right" else far_x_left))

        # ambos verticales
        elif side1 in v_sides and side2 in v_sides:
            opp = ((side1 == "bottom" and side2 == "top" and ey2 >= ey1) or
                   (side1 == "top" and side2 == "bottom" and ey1 >= ey2))
            if opp:
                if abs(ex1 - ex2) < 2:
                    candidates.append([x1, y1, x2, y2])
                candidates.append(self._z_route(x1, y1, side1, x2, y2, side2,
                                                (ey1 + ey2) / 2))
            corridor_right = xmax + 40
            corridor_left  = xmin - 40
            if side1 == "bottom":
                candidates.append([x1, y1, ex1, ey1,
                                   corridor_right, ey1,
                                   corridor_right, ey2,
                                   ex2, ey2, x2, y2])
                candidates.append([x1, y1, ex1, ey1,
                                   corridor_left, ey1,
                                   corridor_left, ey2,
                                   ex2, ey2, x2, y2])
            else:
                candidates.append([x1, y1, ex1, ey1,
                                   corridor_right, ey1,
                                   corridor_right, ey2,
                                   ex2, ey2, x2, y2])
                candidates.append([x1, y1, ex1, ey1,
                                   corridor_left, ey1,
                                   corridor_left, ey2,
                                   ex2, ey2, x2, y2])
            if side1 == side2:
                far_y = ymax + 40 if side1 == "bottom" else ymin - 40
                candidates.append(self._u_route(x1, y1, side1, x2, y2, side2, far_y))

        # perpendiculares → L-shape + variantes
        else:
            if side1 in h_sides:
                candidates.append([x1, y1, ex1, ey1, ex2, ey1, ex2, ey2, x2, y2])
                # variante con codo invertido
                candidates.append([x1, y1, ex1, ey1, ex1, ey2, ex2, ey2, x2, y2])
            else:
                candidates.append([x1, y1, ex1, ey1, ex1, ey2, ex2, ey2, x2, y2])
                candidates.append([x1, y1, ex1, ey1, ex2, ey1, ex2, ey2, x2, y2])

        # ---- seleccionar primer candidato sin colisión ----
        exclude = set()
        if src_id is not None: exclude.add(src_id)
        if dst_id is not None: exclude.add(dst_id)

        for path in candidates:
            if not self._path_crosses_others(path, exclude):
                return path
        # ninguno limpio: devolver el primero
        return candidates[0] if candidates else [x1, y1, x2, y2]

    def _stream_endpoints(self, s):
        """Polyline ortogonal del stream en coords del modelo + sides."""
        b_src = self.fs.blocks.get(s.src)
        b_dst = self.fs.blocks.get(s.dst)
        if b_src is None or b_dst is None:
            return None

        _pn1, side1, x1, y1 = self._resolve_port(b_src, s.src_port, "right")
        _pn2, side2, x2, y2 = self._resolve_port(b_dst, s.dst_port, "left")
        return self._ortho_route(x1, y1, side1, x2, y2, side2,
                                 src_id=s.src, dst_id=s.dst)

    def _label_xy(self, pts):
        """Punto medio del polyline para colocar el label."""
        if len(pts) >= 4:
            # punto medio del segmento horizontal más largo
            best_len = -1
            best = (pts[0], pts[1])
            for i in range(0, len(pts) - 2, 2):
                x1, y1 = pts[i], pts[i+1]
                x2, y2 = pts[i+2], pts[i+3]
                if y1 == y2:
                    L = abs(x2 - x1)
                    if L > best_len:
                        best_len = L
                        best = ((x1 + x2) / 2, y1 - 10)
            return best
        return (pts[0], pts[1])

    def _render_stream(self, s):
        pts = self._stream_endpoints(s)
        if pts is None:
            return
        canvas_pts = [self._sc(v) for v in pts]
        color = STREAM_ROLE_COLORS.get(s.role, STREAM_COLOR)

        s.canvas_line = self.canvas.create_line(
            *canvas_pts,
            fill=color, width=2, arrow="last",
            arrowshape=(12, 14, 5),
            tags=("stream", f"s{s.id}"),
        )
        lx, ly = self._label_xy(pts)
        text = self._stream_label_text(s)
        font = ("Segoe UI", max(6, int(8 * self.zoom)))
        # texto invisible primero para medir bbox, después rectangulo + texto encima
        s.canvas_label = self.canvas.create_text(
            self._sc(lx), self._sc(ly),
            text=text, fill="#222", font=font,
            tags=("stream", f"s{s.id}"),
        )
        bbox = self.canvas.bbox(s.canvas_label)
        if bbox:
            pad = max(2, int(2 * self.zoom))
            s.canvas_lbl_bg = self.canvas.create_rectangle(
                bbox[0] - pad, bbox[1] - 1,
                bbox[2] + pad, bbox[3] + 1,
                fill="#ffffff", outline="",
                tags=("stream", f"s{s.id}"),
            )
            # bring the text to front so the bg doesn't cover it
            self.canvas.tag_lower(s.canvas_lbl_bg, s.canvas_label)

    def _stream_label_text(self, s):
        role_tag = " [feed]" if s.role == "feed" else (" [product]" if s.role == "product" else "")
        flow = f"  {s.mass_flow:g} tm/año" if s.mass_flow else ""
        return s.name + role_tag + flow

    def _refresh_stream(self, s):
        pts = self._stream_endpoints(s)
        if pts is None:
            return
        canvas_pts = [self._sc(v) for v in pts]
        if s.canvas_line is not None:
            self.canvas.coords(s.canvas_line, *canvas_pts)
            color = STREAM_ROLE_COLORS.get(s.role, STREAM_COLOR)
            self.canvas.itemconfigure(s.canvas_line, fill=color)
        if s.canvas_label is not None:
            lx, ly = self._label_xy(pts)
            self.canvas.coords(s.canvas_label, self._sc(lx), self._sc(ly))
            self.canvas.itemconfigure(s.canvas_label, text=self._stream_label_text(s))
            # reajustar fondo
            bbox = self.canvas.bbox(s.canvas_label)
            if bbox and s.canvas_lbl_bg is not None:
                pad = max(2, int(2 * self.zoom))
                self.canvas.coords(
                    s.canvas_lbl_bg,
                    bbox[0] - pad, bbox[1] - 1,
                    bbox[2] + pad, bbox[3] + 1,
                )

    def _delete_stream(self, sid):
        s = self.fs.streams.pop(sid, None)
        if s is None:
            return
        for cid in (s.canvas_line, s.canvas_label, s.canvas_lbl_bg):
            if cid is not None:
                self.canvas.delete(cid)

    def _select_stream(self, sid):
        self._deselect_all_visual()
        self.selected_block = None
        self.selected_stream = sid
        s = self.fs.streams.get(sid)
        if s and s.canvas_line is not None:
            color = STREAM_ROLE_COLORS_SEL.get(s.role, STREAM_COLOR_SEL)
            self.canvas.itemconfigure(s.canvas_line, fill=color, width=3)
        self._update_properties()

    @staticmethod
    def _block_anchor(b, side):
        # legacy: anchor centrado.  Hoy usamos _block_port para
        # distribuir múltiples puertos a lo largo del lado.
        if side == "right":
            return b.x + BLOCK_W, b.y + BLOCK_H / 2
        if side == "left":
            return b.x, b.y + BLOCK_H / 2
        if side == "top":
            return b.x + BLOCK_W / 2, b.y
        if side == "bottom":
            return b.x + BLOCK_W / 2, b.y + BLOCK_H

    # ==================================================
    # CANVAS EVENTS
    # ==================================================

    def _draw_grid(self):
        # grid en coords del CANVAS (ya escaladas), cubre el scrollregion
        sr = self.canvas.cget("scrollregion").split()
        try:
            w = float(sr[2]) if len(sr) >= 4 else CANVAS_W_LOG * self.zoom
            h = float(sr[3]) if len(sr) >= 4 else CANVAS_H_LOG * self.zoom
        except ValueError:
            w, h = CANVAS_W_LOG * self.zoom, CANVAS_H_LOG * self.zoom

        step = max(8, int(GRID_STEP * self.zoom))  # no spammear líneas a zoom bajo
        x = 0
        while x <= w:
            self.canvas.create_line(x, 0, x, h, fill=GRID_COLOR, tags=("grid",))
            x += step
        y = 0
        while y <= h:
            self.canvas.create_line(0, y, w, y, fill=GRID_COLOR, tags=("grid",))
            y += step
        self.canvas.tag_lower("grid")

    def _block_under_cursor(self, x, y):
        for b in self.fs.blocks.values():
            if (b.x <= x <= b.x + BLOCK_W and
                b.y <= y <= b.y + BLOCK_H):
                return b
        return None

    def _stream_under_cursor(self, mx, my):
        """mx, my en coords del MODELO."""
        cx, cy = self._sc(mx), self._sc(my)
        items = self.canvas.find_closest(cx, cy, halo=4)
        if not items:
            return None
        for item in items:
            for tag in self.canvas.gettags(item):
                if tag.startswith("s") and tag[1:].isdigit():
                    sid = int(tag[1:])
                    s = self.fs.streams.get(sid)
                    if s is not None:
                        bbox = self.canvas.bbox(item)
                        if bbox and bbox[0] - 6 <= cx <= bbox[2] + 6 and bbox[1] - 6 <= cy <= bbox[3] + 6:
                            return s
        return None

    def _on_left_click(self, event):
        # space+drag = pan
        if self._panning:
            self.canvas.scan_mark(event.x, event.y)
            self.canvas.configure(cursor="fleur")
            return
        mx, my = self._model_xy(event)
        self.canvas.focus_set()

        # pending connection
        if self._connecting_from is not None:
            b_dst = self._block_under_cursor(mx, my)
            if b_dst is not None:
                self._add_stream(self._connecting_from, b_dst.id)
            self._clear_pending_connection()
            return

        b = self._block_under_cursor(mx, my)
        if b is not None:
            self._select_block(b.id)
            self._drag_block = b.id
            self._drag_offx = mx - b.x
            self._drag_offy = my - b.y
            return

        s = self._stream_under_cursor(mx, my)
        if s is not None:
            self._select_stream(s.id)
            return

        self._deselect_all_visual()
        self.selected_block = None
        self.selected_stream = None
        self._update_properties()

    def _on_left_drag(self, event):
        if self._panning:
            self.canvas.scan_dragto(event.x, event.y, gain=1)
            return
        if self._drag_block is None:
            return
        mx, my = self._model_xy(event)
        b = self.fs.blocks.get(self._drag_block)
        if b is None:
            return
        nx = round((mx - self._drag_offx) / GRID_STEP) * GRID_STEP
        ny = round((my - self._drag_offy) / GRID_STEP) * GRID_STEP
        nx = max(0, nx)
        ny = max(0, ny)
        dx_c = self._sc(nx - b.x)
        dy_c = self._sc(ny - b.y)
        b.x = nx
        b.y = ny
        # mueve rect/text/sub + todos los port markers de un saque (tag b{id})
        self.canvas.move(f"b{b.id}", dx_c, dy_c)
        self.canvas.move(f"port_b{b.id}", dx_c, dy_c)
        for s in self.fs.streams.values():
            if s.src == b.id or s.dst == b.id:
                self._refresh_stream(s)

    def _on_left_release(self, event):
        if self._panning:
            self.canvas.configure(cursor="")
        self._drag_block = None

    def _on_right_click(self, event):
        mx, my = self._model_xy(event)
        b = self._block_under_cursor(mx, my)
        if b is None:
            return

        menu = Menu(self.canvas, tearoff=0)
        menu.add_command(label=f"{b.name}", state="disabled")
        menu.add_separator()
        menu.add_command(label="Conectar a…",
                         command=lambda: self._start_connection(b.id))
        menu.add_command(label="Editar propiedades (doble-click)",
                         command=lambda: self._open_block_dialog(b))
        menu.add_command(label="Borrar",
                         command=lambda: self._delete_and_refresh_block(b.id))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _on_double_click(self, event):
        mx, my = self._model_xy(event)
        b = self._block_under_cursor(mx, my)
        if b is not None:
            self._open_block_dialog(b)
            return
        s = self._stream_under_cursor(mx, my)
        if s is not None:
            d = StreamEditDialog(self.win, s, self.fs)
            if d.result == "ok":
                self._refresh_stream(s)
                self._update_properties()

    def _open_block_dialog(self, b):
        d = BlockEditDialog(self.win, b)
        if d.result == "ok":
            self._refresh_block_text(b)
            self._update_properties()

    def _delete_and_refresh_block(self, bid):
        self.selected_block = bid
        self.delete_selected()

    def _start_connection(self, bid):
        self._connecting_from = bid
        b = self.fs.blocks[bid]
        self.status_var.set(
            f"Conectando desde {b.name}…  hacé click en el equipo destino (Esc para cancelar)"
        )

    def _clear_pending_connection(self):
        self._connecting_from = None
        self._update_status()

    # ==================================================
    # PROPERTIES PANEL
    # ==================================================

    def _update_properties(self):
        if self.selected_block is not None:
            b = self.fs.blocks[self.selected_block]
            spec = eq.EQUIPMENT_DATA.get(b.eq_type, {})

            ins = [s for s in self.fs.streams.values() if s.dst == b.id]
            outs = [s for s in self.fs.streams.values() if s.src == b.id]
            in_total  = sum(s.mass_flow for s in ins)
            out_total = sum(s.mass_flow for s in outs)
            bal = "(fuente/sumidero abierto)"
            if ins and outs:
                diff = abs(in_total - out_total)
                rel  = diff / max(in_total, out_total, 1e-9)
                if rel < 0.005:
                    bal = "✓ cierra"
                else:
                    bal = f"✗ {diff:g} tm/año  ({rel*100:.1f}%)"

            txt = (
                f"EQUIPO    {b.name}\n"
                f"Tipo      {b.eq_type}\n"
                f"Cat.      {spec.get('categoria', '?')}\n"
                f"S         {b.S:g} {spec.get('S_unit', '?')}\n"
                f"n         {b.n}\n"
                f"Rango     [{spec.get('S_min', '?')} … {spec.get('S_max', '?')}]\n\n"
                f"Entradas: {len(ins)}  ({in_total:g} tm/año)\n"
                f"Salidas:  {len(outs)} ({out_total:g} tm/año)\n"
                f"Balance:  {bal}"
            )
            self.prop_var.set(txt)
            return

        if self.selected_stream is not None:
            s = self.fs.streams[self.selected_stream]
            src = self.fs.blocks[s.src].name
            dst = self.fs.blocks[s.dst].name
            sp = s.src_port or "(auto)"
            dp = s.dst_port or "(auto)"
            txt = (
                f"CORRIENTE  {s.name}\n"
                f"Desde      {src}  ({sp})\n"
                f"Hacia      {dst}  ({dp})\n"
                f"Rol        {s.role}\n"
                f"Flujo      {s.mass_flow:g} tm/año"
            )
            if s.role in ("feed", "product"):
                price = s.price_usd_per_tm
                total_usd = price * s.mass_flow
                label = "Ingreso" if s.role == "product" else "Costo MP"
                txt += (
                    f"\nPrecio     {price:g} USD/tm"
                    f"\n{label:10s} $ {total_usd:>12,.0f}/año"
                )
            self.prop_var.set(txt)
            return

        self.prop_var.set(
            "(nada seleccionado)\n\n"
            "Hacé click en un equipo o corriente para ver sus propiedades.\n"
            "Click derecho en un equipo → Conectar a… para dibujar flechas."
        )

    def _update_status(self):
        nb = len(self.fs.blocks)
        ns = len(self.fs.streams)
        if self._connecting_from is None:
            self.status_var.set(f"{nb} equipos · {ns} corrientes")

    # ==================================================
    # COMPUTE / APPLY
    # ==================================================

    def compute(self):
        if not self.fs.blocks:
            messagebox.showinfo("Calcular",
                                "Primero agregá al menos un equipo.")
            return

        equipos = [
            {"nombre": b.eq_type, "S": b.S, "n": b.n}
            for b in self.fs.blocks.values()
        ]

        # tipo de planta + año CEPCI vía prompts (v2: UI persistente)
        from tkinter.simpledialog import askstring
        plant_type = askstring(
            "Tipo de planta",
            "Tipo de planta:\n  Fluid processing (default)\n  Solid-fluid processing\n  Solid processing",
            initialvalue="Fluid processing",
        )
        if not plant_type:
            return
        if plant_type not in eq.LANG_FACTORS:
            messagebox.showerror("Inválido", f"Tipo de planta desconocido: {plant_type}")
            return

        year_str = askstring("Año destino (CEPCI)",
                             "Año:", initialvalue="2026")
        try:
            year = int(year_str)
        except (TypeError, ValueError):
            return

        try:
            res = eq.lang_fci(equipos, plant_type=plant_type, year_target=year)
        except ValueError as e:
            messagebox.showerror("Error de cálculo", str(e))
            return

        # balance de masa global por bloque
        warnings_mb = []
        for b in self.fs.blocks.values():
            ins  = [s for s in self.fs.streams.values() if s.dst == b.id]
            outs = [s for s in self.fs.streams.values() if s.src == b.id]
            if not ins or not outs:
                continue  # fuente o sumidero — sin balance esperado
            in_t  = sum(s.mass_flow for s in ins)
            out_t = sum(s.mass_flow for s in outs)
            diff  = abs(in_t - out_t)
            rel   = diff / max(in_t, out_t, 1e-9)
            if rel >= 0.005:
                warnings_mb.append(
                    f"{b.name}: ent={in_t:g}, sal={out_t:g}, Δ={diff:g} tm/año ({rel*100:.1f}%)"
                )

        # ISBL implícito
        if self.df_capital is not None and not self.df_capital.empty:
            try:
                osbl = float(self.df_capital.iloc[1, 2]) / (100.0 if self.df_capital.iloc[1, 2] > 1 else 1.0)
                eng  = float(self.df_capital.iloc[2, 2]) / (100.0 if self.df_capital.iloc[2, 2] > 1 else 1.0)
                cont = float(self.df_capital.iloc[3, 2]) / (100.0 if self.df_capital.iloc[3, 2] > 1 else 1.0)
                isbl_mm = eq.isbl_implicito(res["FCI_MMUSD"], osbl, eng, cont)
            except Exception:
                isbl_mm = None
        else:
            isbl_mm = None

        # feeds y products (para coupling con análisis económico)
        feeds    = [s for s in self.fs.streams.values() if s.role == "feed"]
        products = [s for s in self.fs.streams.values() if s.role == "product"]
        feed_total    = sum(s.mass_flow for s in feeds)
        product_total = sum(s.mass_flow for s in products)

        # mostrar resultados
        out_lines = [
            f"Tipo de planta:  {plant_type}",
            f"Factor Lang:     {res['lang_factor']:.2f}",
            f"Año destino:     {res['year_target']}",
            "",
            f"Σ Cp°:   $ {res['sum_Cp']:>14,.0f}",
            f"FCI:     {res['FCI_MMUSD']:>10.2f} MM USD",
        ]
        if isbl_mm is not None:
            out_lines.append(f"ISBL:    {isbl_mm:>10.2f} MM USD  (back-out)")
        elif self.mode == "main":
            # en modo main, ISBL implícito requiere %OSBL/%ENG/%CONT
            # del análisis económico (defaults Turton)
            isbl_mm = eq.isbl_implicito(res["FCI_MMUSD"], 0.30, 0.10, 0.10)
            out_lines.append(f"ISBL:    {isbl_mm:>10.2f} MM USD  (defaults 30/10/10%)")
        else:
            out_lines.append("ISBL:    (cargá un proyecto)")

        if feeds or products:
            out_lines.append("")
            out_lines.append("─ Flujos másicos ─")
            if products:
                out_lines.append(f"Producción anual: {product_total:g} tm/año  "
                                 f"({len(products)} corriente{'s' if len(products)>1 else ''})")
            if feeds:
                out_lines.append(f"Alimentación:     {feed_total:g} tm/año  "
                                 f"({len(feeds)} corriente{'s' if len(feeds)>1 else ''})")
                for s in feeds:
                    out_lines.append(f"  · {s.name}: {s.mass_flow:g} tm/año")

            # Ingresos / costos materia prima a partir de precios USD/tm
            revenue = sum(s.mass_flow * s.price_usd_per_tm for s in products)
            raw_mp  = sum(s.mass_flow * s.price_usd_per_tm for s in feeds)
            if revenue or raw_mp:
                out_lines.append("")
                out_lines.append("─ Economía aproximada (USD/año) ─")
                if revenue:
                    out_lines.append(f"Ingresos:         $ {revenue:>14,.0f}")
                if raw_mp:
                    out_lines.append(f"Materia prima:    $ {raw_mp:>14,.0f}")
                if revenue and raw_mp:
                    margin = revenue - raw_mp
                    out_lines.append(f"Margen bruto:     $ {margin:>14,.0f}")

        if res["warnings"]:
            out_lines.append("")
            out_lines.append("⚠ Avisos en Cp°:")
            out_lines += [f"  · {w}" for w in res["warnings"]]

        if warnings_mb:
            out_lines.append("")
            out_lines.append("⚠ Balance de masa:")
            out_lines += [f"  · {w}" for w in warnings_mb]
        elif self.fs.streams:
            out_lines.append("")
            out_lines.append("✓ Balance de masa OK en todos los equipos internos.")

        self.results_var.set("\n".join(out_lines))
        self._last_isbl = isbl_mm
        self._last_feed_total    = feed_total
        self._last_product_total = product_total
        self._last_feeds    = feeds
        self._last_products = products

    def apply_isbl(self):
        if not hasattr(self, "_last_isbl") or self._last_isbl is None:
            messagebox.showinfo("Aplicar ISBL", "Primero apretá Calcular.")
            return
        if self.df_capital is None or self.df_capital.empty:
            messagebox.showinfo(
                "Aplicar ISBL",
                "Para aplicar el ISBL al análisis, importá o creá un proyecto primero."
            )
            return
        if not messagebox.askyesno(
            "Aplicar ISBL",
            f"¿Aplicar ISBL = {self._last_isbl:.2f} MM USD al proyecto?"
        ):
            return
        self.df_capital.iat[0, 2] = float(self._last_isbl)
        if self.on_apply is not None:
            self.on_apply()
        messagebox.showinfo("Aplicado", "ISBL actualizado en el proyecto.")

    # ==================================================
    # TRANSICIÓN AL ANÁLISIS ECONÓMICO (modo main)
    # ==================================================

    def launch_analysis(self):
        """Modo main: genera un xlsx temporal con:
          - df_capital: ISBL inyectado desde el diagrama
          - df_fixed:   templates Turton (o lo que traiga el xlsx base)
          - df_variable: filas de los feeds y products del diagrama
                         auto-poblados (marcadas con sufijo '(PFD)')
        y lanza `python ANA.py --import <tmp.xlsx>`.

        El diagrama de bloques en esta ventana queda intacto.
        Subir un xlsx solo aporta datos económicos extra; nunca
        reemplaza el diagrama.
        """
        import subprocess
        import sys
        import os
        import tempfile

        if not self.fs.blocks:
            if not messagebox.askyesno(
                "Sin proceso modelado",
                "El diagrama está vacío. ¿Abrir el análisis económico igual?",
            ):
                return
            feeds, products, isbl = [], [], None
        else:
            # asegurar Compute previo (silencioso con defaults si no hay)
            if not hasattr(self, "_last_isbl") or self._last_isbl is None:
                try:
                    equipos = [
                        {"nombre": b.eq_type, "S": b.S, "n": b.n}
                        for b in self.fs.blocks.values()
                    ]
                    res = eq.lang_fci(equipos, plant_type="Fluid processing", year_target=2026)
                    self._last_isbl = eq.isbl_implicito(res["FCI_MMUSD"], 0.30, 0.10, 0.10)
                    self._last_feeds    = [s for s in self.fs.streams.values() if s.role == "feed"]
                    self._last_products = [s for s in self.fs.streams.values() if s.role == "product"]
                except ValueError as e:
                    messagebox.showerror("Error de cálculo", str(e))
                    return
            isbl = self._last_isbl
            feeds    = getattr(self, "_last_feeds", [])    or []
            products = getattr(self, "_last_products", []) or []

        # opción de xlsx base
        usar_xlsx = messagebox.askyesnocancel(
            "Análisis económico",
            "El diagrama de bloques queda intacto en esta ventana.\n\n"
            "¿Usar un .xlsx base existente para el análisis?\n\n"
            "  Sí        → seleccionás el archivo y le agrego los\n"
            "              feeds/products del diagrama\n"
            "  No        → genero la plantilla Turton + feeds/products\n"
            "  Cancelar  → vuelvo al diagrama",
        )
        if usar_xlsx is None:
            return

        base_xlsx = None
        if usar_xlsx:
            base_xlsx = filedialog.askopenfilename(
                title="Proyecto .xlsx base",
                filetypes=[("Excel", "*.xlsx *.xls")],
            )
            if not base_xlsx:
                return

        # generar xlsx temporal
        try:
            tmp_dir = tempfile.gettempdir()
            tmp_path = os.path.join(tmp_dir, f"ANA_from_PFD_{os.getpid()}.xlsx")
            self._write_project_xlsx(tmp_path, isbl, feeds, products, base_xlsx)
        except Exception as e:
            messagebox.showerror(
                "Falló la generación del xlsx",
                f"{type(e).__name__}: {e}\n\nProbá importar tu xlsx "
                "manualmente desde ANA.py.",
            )
            return

        cmd = [sys.executable, "ANA.py", "--import", tmp_path]

        try:
            cwd = os.path.dirname(os.path.abspath(__file__))
            subprocess.Popen(cmd, cwd=cwd)
        except Exception as e:
            messagebox.showerror("Falló el lanzamiento",
                                 f"{type(e).__name__}: {e}")

    # ==================================================
    # GENERACIÓN DEL XLSX PARA EL ANÁLISIS ECONÓMICO
    # ==================================================

    def _write_project_xlsx(self, path, isbl, feeds, products, base_xlsx=None):
        """Escribe un xlsx con las 3 secciones que ANA.ImportarProyecto
        espera (Capital cols A-C, Fixed cols E-G, Variable cols I-N).

        Si base_xlsx es None → templates Turton + filas del PFD.
        Si base_xlsx existe → reusa ese xlsx, override ISBL, mantiene
        sus filas variables y APENDE las del PFD (marcadas con sufijo
        ' (PFD)' para distinguirlas; si reabrís el análisis varias
        veces, las del PFD se reemplazan, no se duplican)."""
        import pandas as pd
        import templates as tmpl

        if base_xlsx:
            df_capital, df_fixed, df_variable = self._read_project_xlsx(base_xlsx)
        else:
            df_capital  = tmpl.template_capital()
            df_fixed    = tmpl.template_fixed()
            df_variable = pd.DataFrame(columns=[
                "variable operating costs", "units", "time basis",
                "flowrate", "price usd/units", "stream",
            ])

        # inyectar ISBL
        if isbl is not None and not df_capital.empty:
            df_capital.iat[0, 2] = float(isbl)

        # dedupe: borrar filas previas marcadas '(PFD)'
        if "variable operating costs" in df_variable.columns:
            mask = df_variable["variable operating costs"].astype(str).str.contains(
                r"\(PFD\)", regex=True, na=False)
            df_variable = df_variable[~mask].reset_index(drop=True)

        # append filas del PFD (con precio del propio stream)
        new_rows = []
        for s in products:
            new_rows.append({
                "variable operating costs": f"{s.name} (PFD)",
                "units":              "tm",
                "time basis":         "year",
                "flowrate":           float(s.mass_flow),
                "price usd/units":    float(getattr(s, "price_usd_per_tm", 0.0)),
                "stream":             "Key Products",
            })
        for s in feeds:
            new_rows.append({
                "variable operating costs": f"{s.name} (PFD)",
                "units":              "tm",
                "time basis":         "year",
                "flowrate":           float(s.mass_flow),
                "price usd/units":    float(getattr(s, "price_usd_per_tm", 0.0)),
                "stream":             "Raw Materials",
            })
        if new_rows:
            df_variable = pd.concat(
                [df_variable, pd.DataFrame(new_rows)],
                ignore_index=True,
            )

        # escribir xlsx con las 3 secciones lado a lado
        self._write_3sections_xlsx(path, df_capital, df_fixed, df_variable)

    @staticmethod
    def _read_project_xlsx(path):
        """Lee un xlsx del análisis económico y devuelve los 3 dfs
        (capital, fixed, variable) replicando la lógica de
        ANA.ImportarProyecto."""
        import pandas as pd
        df_raw = pd.read_excel(path, header=None)

        def _section(cols):
            df = df_raw.iloc[:, cols].copy()
            df = df.dropna(how="all")
            if df.empty:
                return df
            df.columns = df.iloc[0].tolist()
            df = df.iloc[1:].reset_index(drop=True)
            df = df[df.iloc[:, 0].notna()].reset_index(drop=True)
            return df

        df_capital  = _section([0, 1, 2])
        df_fixed    = _section([4, 5, 6])
        df_variable = _section([8, 9, 10, 11, 12, 13])
        return df_capital, df_fixed, df_variable

    @staticmethod
    def _write_3sections_xlsx(path, df_capital, df_fixed, df_variable):
        """Escribe un xlsx con:
          cols A-C  → Capital Costs
          col  D    → vacía
          cols E-G  → Fixed Operating Costs
          col  H    → vacía
          cols I-N  → Variable Operating Costs
        Formato compatible con ANA.ImportarProyecto."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project"

        def _write_block(df, col_start):
            if df is None or df.empty:
                return
            # header
            for j, name in enumerate(df.columns):
                ws.cell(row=1, column=col_start + j, value=str(name))
            # rows
            for i in range(len(df)):
                for j in range(len(df.columns)):
                    v = df.iat[i, j]
                    try:
                        if v is None:
                            continue
                        import math
                        if isinstance(v, float) and math.isnan(v):
                            continue
                    except Exception:
                        pass
                    ws.cell(row=2 + i, column=col_start + j, value=v)

        _write_block(df_capital,  1)
        _write_block(df_fixed,    5)
        _write_block(df_variable, 9)
        wb.save(path)


# ======================================================
# ENTRY POINT
# ======================================================

def AbrirVentanaFlowsheet(parent, df_capital=None, on_apply=None):
    """Modo legacy: abre el editor como Toplevel hijo del
    análisis económico.  Botón final 'Apply ISBL' actualiza
    df_capital in-place.  Usado desde ANA.py > Tools."""
    top = Toplevel(parent)
    top.transient(parent)
    FlowsheetEditor(top, df_capital=df_capital, on_apply=on_apply, mode="modal")
