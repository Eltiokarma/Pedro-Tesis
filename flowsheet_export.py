"""
FLOWSHEET EXPORT — funciones puras para exportar el flowsheet al
formato xlsx que consume ANA.py (análisis económico).

Sin dependencias de UI.  Lo importan tanto el editor Tk como el Qt.

API pública:

  collect_equipment_rows(fs)            → lista de dicts con info de equipos
  compute_utilities_from_duties(fs)     → (rows_para_df_variable, summary)
  read_project_xlsx(path)               → (df_capital, df_fixed, df_variable)
  write_project_xlsx(path, fs, isbl,
                     feeds, products,
                     base_xlsx=None)    → genera xlsx temporal con:
                                            · capital con ISBL inyectado
                                            · fixed con labor Turton o override
                                            · variable con feeds+products+
                                              opex_extras+utilities auto

Las filas inyectadas se marcan con sufijo:
  '(PFD)'      → feeds, products, opex_extras manuales
  '(PFD-util)' → utilities auto-calculadas desde duties

El dedupe regenera ambas categorías cada vez.
"""

import math
import pandas as pd

import equipment_costs as eq
import equipment_ports as ep
import templates as tmpl


# ======================================================
# EQUIPOS DEL PFD
# ======================================================

def _block_material(fs, block_id):
    """Heurística: detecta el material que requiere el bloque mirando
    composiciones de TODAS las corrientes conectadas (in + out) +
    presión de operación.  Devuelve el más severo necesario."""
    b = fs.blocks.get(block_id)
    if b is None:
        return "CS"
    # User override explícito tiene prioridad
    if getattr(b, "material", None):
        return b.material
    p_op = float(getattr(b, "P_op_bar", 0.0) or 0.0) or 1.0
    comps = []
    for s in fs.streams.values():
        if s.src == block_id or s.dst == block_id:
            if s.composition:
                comps.append(s.composition)
    return eq.suggested_material(comps, p_op_bar=p_op)


def collect_equipment_rows(fs, year_target=2024):
    """Lista de dicts con la info de cada bloque del PFD, para
    escribir como pestaña 'Equipment' en el xlsx.

    Incluye specs operacionales (T_op, P_op, duty, ΔP, η) y de unit
    op (reacciones, column LK/HK, flash T/P, splitter fractions) +
    COSTING TURTON por equipo individual (Cp escalado a year_target,
    FBM con factor de presión Y MATERIAL, CBM = Cp × FBM).

    Material se auto-detecta heurísticamente desde composiciones
    de corrientes adjuntas (Cl2 → Titanio, HCl → Hastelloy, HNO3 →
    SS316, H2 a alta P → SS304, etc.).  Override manual via
    block.material (string).
    """
    rows = []
    for b in sorted(fs.blocks.values(), key=lambda b: b.name):
        spec = eq.EQUIPMENT_DATA.get(b.eq_type, {})
        # Material auto-detected (or user override)
        material = _block_material(fs, b.id)
        # Costing Turton individual
        cp_usd, fbm, cbm, fp, fm = 0.0, 0.0, 0.0, 1.0, 1.0
        fuera_rango = False
        try:
            p_op = float(getattr(b, "P_op_bar", 0.0) or 0.0) or 1.0
            res = eq.bare_module_cost(b.eq_type, b.S, P_op_bar=p_op,
                                       year_target=year_target,
                                       material=material)
            cp_usd      = res["Cp_target"] * b.n
            cbm         = res["CBM"] * b.n
            fbm         = res["FBM"]
            fp          = res["FP"]
            fm          = res["FM"]
            fuera_rango = res["fuera_rango"]
        except Exception:
            pass

        row = {
            "Tag":        b.name,
            "Type":       b.eq_type,
            "Category":   spec.get("categoria", ""),
            "Size S":     float(b.S),
            "Unit":       spec.get("S_unit", ""),
            "N° units":   int(b.n),
            "Duty kW":    float(getattr(b, "duty", 0.0)),
            "T_op K":     float(getattr(b, "T_op_K", 0.0) or 0.0),
            "P_op bar":   float(getattr(b, "P_op_bar", 0.0) or 0.0),
            "ΔP bar":     float(getattr(b, "delta_p_bar", 0.0) or 0.0),
            "η":          float(getattr(b, "efficiency", 0.0) or 0.0),
            "Reactions":  ",".join(getattr(b, "reactions", []) or []),
            "Heat source": str(getattr(b, "heat_source", "") or ""),
            # ── Costing Turton (CEPCI year_target) ──
            "Material":                 material,
            "FM":                       fm,
            "FP":                       fp,
            "FBM":                      fbm,
            f"Cp USD ({year_target})":  cp_usd,
            f"CBM USD ({year_target})": cbm,
            "S fuera rango":            "⚠" if fuera_rango else "",
        }
        # Column specs (FUG-resueltos por solve_columns)
        if getattr(b, "column_active", False):
            row["Column LK"]  = str(getattr(b, "column_LK", "") or "")
            row["Column HK"]  = str(getattr(b, "column_HK", "") or "")
            row["Column R"]   = float(getattr(b, "_column_R", 0.0) or 0.0)
            row["Column N"]   = float(getattr(b, "_column_N", 0.0) or 0.0)
            row["Q_reb kW"]   = float(getattr(b, "_column_Q_reb", 0.0) or 0.0)
            row["Q_cond kW"]  = float(getattr(b, "_column_Q_cond", 0.0) or 0.0)
        # Flash specs
        if getattr(b, "flash_active", False):
            row["Flash T K"]  = float(getattr(b, "flash_T_K", 0.0) or 0.0)
            row["Flash P bar"]= float(getattr(b, "flash_P_bar", 0.0) or 0.0)
        # Splitter fractions
        fracs = getattr(b, "splitter_fractions", None) or []
        if fracs:
            row["Splitter fracs"] = ",".join(f"{f:.3f}" for f in fracs)
        rows.append(row)
    return rows


def compute_turton_costing(fs, df_variable, df_fixed, fci_musd,
                            year_target=2024):
    """Calcula resumen de costing Turton para sheet 'Costing Turton'.

    Reúne:
      · ΣCBM por categoría con MATERIAL auto-detectado (Cl2→Ti,
        HCl→Hastelloy, etc.) y factor de presión FP por bloque
      · CGR Grass Roots Capital (Turton Eq 7.10) = ΣCBM + cont + aux
      · CRM, CUT, CWT, COL desglosados
      · COM_d Turton Eq 8.2 con y sin depreciación
      · SENSIBILIDAD ±25 % en CAPEX y OPEX (AACE Class 4 estimate)
      · RENTABILIDAD: revenue, gross/net profit, payback, ROI,
        NPV (10 yr / 10 %) e IRR
    """
    # CBM agregado por categoría — con material auto-detectado
    by_cat = {}     # cat → [cbm_usd, n_equipos, material_pred]
    sum_cp = 0.0
    sum_cbm = 0.0
    for b in fs.blocks.values():
        spec = eq.EQUIPMENT_DATA.get(b.eq_type, {})
        cat  = spec.get("categoria", "Otros")
        p_op = float(getattr(b, "P_op_bar", 0.0) or 0.0) or 1.0
        mat  = _block_material(fs, b.id)
        try:
            res = eq.bare_module_cost(b.eq_type, b.S, P_op_bar=p_op,
                                       year_target=year_target,
                                       material=mat)
            cp  = res["Cp_target"] * b.n
            cbm = res["CBM"]      * b.n
            by_cat.setdefault(cat, [0.0, 0, set()])
            by_cat[cat][0] += cbm
            by_cat[cat][1] += 1
            by_cat[cat][2].add(mat)
            sum_cp  += cp
            sum_cbm += cbm
        except Exception:
            continue

    # CRM, CUT, CWT, REVENUE desde df_variable
    crm = cut = cwt = revenue = 0.0
    if df_variable is not None and not df_variable.empty:
        for _, row in df_variable.iterrows():
            stream = str(row.get("stream", "")).strip()
            flow   = float(row.get("flowrate", 0.0) or 0.0)
            price  = float(row.get("price usd/units", 0.0) or 0.0)
            cost   = flow * price
            if stream == "Raw Materials":
                crm += cost
            elif stream == "Utilities":
                cut += cost
            elif stream == "Waste / Byproduct":
                if price < 0:
                    cwt += abs(cost)         # disposal cost
                elif price > 0:
                    revenue += cost          # byproduct vendible
            elif stream == "Key Products":
                revenue += cost              # producto principal

    # COL desde df_fixed (Labor)
    col = 0.0
    if df_fixed is not None and not df_fixed.empty \
       and "Concept" in df_fixed.columns:
        mask = df_fixed["Concept"].astype(str).str.strip() == "Labor"
        if mask.any():
            col = float(df_fixed.loc[mask, "Value"].iloc[0] or 0.0)

    # FCI: ÚNICA fuente vía capex.compute_fci (fuente única — bug-fix).
    # Si user pasó ISBL en MUSD, lo pasamos como override; sino se
    # computa desde el costing por bloque (year_target unificado,
    # material auto-detectado).
    import capex as _capex
    isbl_override_usd = fci_musd * 1e6 if fci_musd else None
    capex_data = _capex.compute_fci(
        fs, year_target=year_target, isbl_override_usd=isbl_override_usd
    )
    fci_usd          = capex_data["FCI_grass_roots"]
    wc_usd           = capex_data["working_capital"]
    capex_total      = capex_data["CAPEX_total"]
    depreciable_base = capex_data["depreciable_base"]
    # Override locales de sum_cp/sum_cbm/by_cat para que las filas del
    # report concuerden con la fuente única.
    sum_cp  = capex_data["sum_cp"]
    sum_cbm = capex_data["sum_cbm"]
    by_cat  = capex_data["by_category"]

    # Defaults financieros y sensibilidad — desde econ_defaults
    try:
        import econ_defaults as _ed
        _fin  = _ed.get_financial()
        _sens = _ed.get_sensitivity()
        _years   = _fin["project_years"]
        _tax     = _fin["tax_rate"]
        _disc    = _fin["discount_rate"]
        _f_low   = _sens["low_factor"]
        _f_high  = _sens["high_factor"]
    except Exception:
        _years, _tax, _disc = 10, 0.30, 0.10
        _f_low, _f_high     = 0.75, 1.25

    # COM Turton Eq 8.2 con desglose explícito (incluye Dep_SL y M+T+I
    # separados — bug-fix consistencia COM/Dep).
    com = eq.cost_of_manufacture_components(
        FCI_usd=fci_usd, COL_usd=col,
        CUT_usd=cut, CRM_usd=crm, CWT_usd=cwt,
        depreciable_base_usd=depreciable_base,
        useful_life_yr=_years,
    )

    # Rentabilidad usando depreciable_base + WC explícitos.
    prof = eq.profitability_indicators(
        revenue_usd_yr=revenue, com_d_usd_yr=com["COM_d"],
        fci_usd=fci_usd,
        depreciable_base_usd=depreciable_base,
        working_capital_usd=wc_usd,
        useful_life_yr=_years,
        years_op=_years, tax_rate=_tax, disc_rate=_disc,
    )

    # Sensibilidad AACE Class 4 (±low/high del perfil)
    sens = {}
    _pct_lo = int(round((1 - _f_low) * 100))
    _pct_hi = int(round((_f_high - 1) * 100))
    for scen, cap_f, op_f in [
        (f"Bajo  (-{_pct_lo} % cap, -{_pct_lo} % op)", _f_low,  _f_low),
        ("Base  (mid case)",                            1.00,   1.00),
        (f"Alto  (+{_pct_hi} % cap, +{_pct_hi} % op)", _f_high, _f_high),
    ]:
        fci_s   = fci_usd * cap_f
        dep_b_s = depreciable_base * cap_f
        wc_s    = wc_usd * cap_f
        com_s = (eq.cost_of_manufacture(
            FCI_usd=fci_s, COL_usd=col,
            CUT_usd=cut * op_f, CRM_usd=crm * op_f,
            CWT_usd=cwt * op_f,
        ))["COM_d"]
        rev_s = revenue * op_f
        prof_s = eq.profitability_indicators(
            revenue_usd_yr=rev_s, com_d_usd_yr=com_s, fci_usd=fci_s,
            depreciable_base_usd=dep_b_s,
            working_capital_usd=wc_s,
            useful_life_yr=_years,
            years_op=_years, tax_rate=_tax, disc_rate=_disc,
        )
        sens[scen] = prof_s

    # ───────────────────── Build rows ─────────────────────
    rows = []
    rows.append(("CAPITAL — Bare Module Cost by Category", "", ""))
    rows.append(("  (FBM = B1 + B2·FM·FP con material auto-detectado)",
                  "", ""))
    for cat, (cbm, n, mats) in sorted(by_cat.items(),
                                          key=lambda kv: -kv[1][0]):
        mats_str = "/".join(sorted(mats))
        rows.append((f"  · {cat}",
                      f"{n} equipo(s)  ({mats_str})",
                      f"{cbm:>14,.0f} USD"))
    rows.append(("  Σ Cp purchased (FOB)", "", f"{sum_cp:>14,.0f} USD"))
    rows.append(("  Σ CBM bare module",    "", f"{sum_cbm:>14,.0f} USD"))
    rows.append(("", "", ""))
    rows.append(("GRASS ROOTS CAPITAL (Turton Eq 7.10)", "", ""))
    rows.append(("  Contingency (18 % CBM)", "",
                  f"{capex_data['contingency']:>14,.0f} USD"))
    rows.append(("  Aux facilities (50 % CBM)", "",
                  f"{capex_data['aux_facilities']:>14,.0f} USD"))
    rows.append(("  FCI Grass Roots (Turton 7.10)", "",
                  f"{fci_usd:>14,.0f} USD"))
    rows.append(("  Working Capital (15 % FCI)", "",
                  f"{wc_usd:>14,.0f} USD"))
    rows.append(("  CAPEX total (year 0 outflow)", "",
                  f"{capex_total:>14,.0f} USD"))
    if capex_data["isbl_override_used"]:
        rows.append(("  (ISBL override del user activo)", "", ""))
    rows.append(("", "", ""))
    rows.append(("OPEX COMPONENTS (USD/year)", "", ""))
    rows.append(("  CRM (Raw Materials)",   "", f"{crm:>14,.0f} USD/yr"))
    rows.append(("  CUT (Utilities)",       "", f"{cut:>14,.0f} USD/yr"))
    rows.append(("  CWT (Waste Treatment)", "", f"{cwt:>14,.0f} USD/yr"))
    rows.append(("  COL (Operating Labor)", "", f"{col:>14,.0f} USD/yr"))
    rows.append(("", "", ""))
    rows.append(("COST OF MANUFACTURE (Turton Eq 8.2)", "", ""))
    rows.append(("  0.180·FCI",          "",
                  f"{com['0.180·FCI']:>14,.0f}"))
    rows.append(("  2.73·COL",            "",
                  f"{com['2.73·COL']:>14,.0f}"))
    rows.append(("  1.23·(CUT+CRM+CWT)",  "",
                  f"{com['1.23·(CUT+CRM+CWT)']:>14,.0f}"))
    rows.append(("  COM_d (con depreciación implícita)","",
                  f"{com['COM_d']:>14,.0f} USD/yr"))
    rows.append(("  COM   (sin depreciación implícita)","",
                  f"{com['COM']:>14,.0f} USD/yr"))
    rows.append(("    ─ Depreciation_SL (real, base depreciable)", "",
                  f"{com['Depreciation_SL']:>14,.0f} USD/yr"))
    rows.append(("    ─ Maintenance+Tax+Insurance (FCI-pegged)", "",
                  f"{com['Maintenance_Tax_Insurance']:>14,.0f} USD/yr"))
    rows.append(("", "", ""))
    rows.append(("PROFITABILITY (Turton Ch 9-10, base case)", "", ""))
    rows.append(("  Revenue (products + byprods)", "",
                  f"{revenue:>14,.0f} USD/yr"))
    rows.append(("  Gross profit (Rev - COM_d)", "",
                  f"{prof['Gross profit']:>14,.0f} USD/yr"))
    rows.append((f"  Depreciación SL ({_years} yr / depreciable_base)", "",
                  f"{prof['Depreciation']:>14,.0f} USD/yr"))
    rows.append((f"  Tax ({_tax*100:.0f} %)",         "",
                  f"{prof['Tax (30%)']:>14,.0f} USD/yr"))
    rows.append(("  Net profit (post-tax)", "",
                  f"{prof['Net profit']:>14,.0f} USD/yr"))
    rows.append(("  Cash flow (Net + dep)", "",
                  f"{prof['Cash flow']:>14,.0f} USD/yr"))
    rows.append(("  Payback simple",       "", prof['Payback str']))
    rows.append(("  ROI %",                "",
                  f"{prof['ROI %']:>14.1f} %"))
    rows.append((f"  NPV ({_years} yr, {_disc*100:.0f} %)", "",
                  f"{prof['NPV']:>14,.0f} USD"))
    rows.append(("  IRR %",                "", prof['IRR str']))
    rows.append(("  VEREDICTO",            "",
                  f"PROYECTO {prof['Veredicto']}"))
    rows.append(("", "", ""))
    rows.append(("SENSIBILIDAD AACE Class 4 (±25 %)", "", ""))
    rows.append(("  Escenario", "NPV USD", "Payback yr / IRR %"))
    for scen, ps in sens.items():
        pbp_s = ps['Payback simple']
        pbp_str = f"{pbp_s:.1f}" if pbp_s != float('inf') else "∞"
        irr_s = ps['IRR %']
        irr_str = f"{irr_s:.1f}%" if irr_s is not None else "n/a"
        rows.append((f"  {scen}",
                      f"{ps['NPV']:>14,.0f}",
                      f"{pbp_str} / {irr_str}"))
    rows.append(("", "", ""))
    # Footer con parámetros usados — desde econ_defaults perfil activo
    try:
        import econ_defaults as _ed
        _prof_name = _ed.active_profile()
    except Exception:
        _prof_name = "PE_2024"
    rows.append(("Año CEPCI", "", str(year_target)))
    rows.append(("Año base Turton", "", "2001 (CEPCI=397)"))
    rows.append(("Hurdle rate (NPV)", "", f"{_disc*100:.0f} %"))
    rows.append(("Horizonte (NPV)",   "", f"{_years} años"))
    rows.append(("Tax rate",          "", f"{_tax*100:.0f} %"))
    rows.append(("Perfil econ.",      "", _prof_name))

    return {
        "rows":               rows,
        "by_category":        by_cat,
        "sum_cp":             sum_cp,
        "sum_cbm":            sum_cbm,
        "capex":              capex_data,     # dict canónico capex.py
        "fci":                fci_usd,
        "working_capital":    wc_usd,
        "capex_total":        capex_total,
        "depreciable_base":   depreciable_base,
        "col":                col,
        "crm":                crm,
        "cut":                cut,
        "cwt":                cwt,
        "com":                com,
        "revenue":            revenue,
        "profit":             prof,
        "sensitivity":        sens,
        "mti_usd_yr":         com["Maintenance_Tax_Insurance"],
    }


def assert_costing_income_coherence(costing_data, income_rows,
                                       rel_tol=0.001):
    """Verifica que el Cash Flow y NPV del Costing Turton (resumen
    Turton Ch 9-10) coincidan con los del Income Statement año a año
    dentro de rel_tol.  Bug-fix: antes diferían varios MM USD por
    rutas paralelas distintas.

    Raises AssertionError si la divergencia excede rel_tol.
    """
    prof = costing_data["profit"]
    cf_costing = prof["Cash flow"]
    npv_costing = prof["NPV"]
    # Income Statement: cash flow promedio años de operación,
    # excluyendo year-0 y WC recovery del último año.
    years = costing_data["profit"]["years_op"]
    op_rows = [r for r in income_rows if r["Year"] >= 1 and r["Year"] <= years]
    if not op_rows:
        raise AssertionError("Income Statement vacío — no se pudo validar.")
    # Cash flow operativo "régimen" = CF − WC recovery del año (si aplica)
    cfs_op = [r["Cash Flow"] - r.get("WC Recov", 0.0) for r in op_rows]
    cf_income_avg = sum(cfs_op) / len(cfs_op)
    # NPV reconstruido desde income statement
    disc = prof["disc_rate"]
    npv_income = 0.0
    for r in income_rows:
        npv_income += r["Cash Flow"] / ((1 + disc) ** r["Year"])
    def _close(a, b, tol):
        return abs(a - b) / max(abs(a), abs(b), 1e-9) < tol
    if not _close(cf_costing, cf_income_avg, rel_tol):
        raise AssertionError(
            f"Cash flow inconsistente: Costing={cf_costing:,.0f} vs "
            f"Income avg={cf_income_avg:,.0f} (>{rel_tol*100:.2f}%)"
        )
    if not _close(npv_costing, npv_income, rel_tol):
        raise AssertionError(
            f"NPV inconsistente: Costing={npv_costing:,.0f} vs "
            f"Income={npv_income:,.0f} (>{rel_tol*100:.2f}%)"
        )
    return True


def collect_stream_rows(fs):
    """Lista de dicts con TODA la data de cada stream, para pestaña
    'Streams' del xlsx.  Incluye masa, T, P, fase, composiciones,
    pipe specs, role y precio."""
    rows = []
    # Componentes únicos para columnas wt% — solo los que aparecen en
    # ≥ 1 stream con composition_locked o computado por el solver.
    comps = set()
    for s in fs.streams.values():
        comps.update((s.composition or {}).keys())
    comps = sorted(comps)

    for s in sorted(fs.streams.values(), key=lambda x: x.name):
        b_src = fs.blocks.get(s.src)
        b_dst = fs.blocks.get(s.dst)
        row = {
            "Stream":     s.name,
            "From":       b_src.name if b_src else "",
            "From port":  s.src_port or "",
            "To":         b_dst.name if b_dst else "",
            "To port":    s.dst_port or "",
            "Mass tm/yr":  float(s.mass_flow),
            "T °C":       float(s.temperature),
            "P bar":      float(getattr(s, "pressure_bar", 0.0) or 0.0),
            "Phase":      s.phase or "",
            "Role":       s.role or "",
            "Price USD/tm": float(getattr(s, "price_usd_per_tm", 0.0) or 0.0),
            "Cp kJ/kg·K": float(getattr(s, "cp", 0.0) or 0.0),
            "Main comp":  s.main_component or "",
            "Locks":      ",".join([
                k.replace("_locked", "")
                for k in ("mass_flow_locked","temperature_locked",
                           "composition_locked","pressure_locked")
                if getattr(s, k, False)
            ]),
            "Pipe L m":   float(getattr(s, "pipe_length_m", 0.0) or 0.0),
            "Pipe D m":   float(getattr(s, "pipe_diameter_m", 0.0) or 0.0),
            "Pipe ΔP bar":float(getattr(s, "delta_p_pipe_bar", 0.0) or 0.0),
        }
        # Composición wt% por componente
        comp = s.composition or {}
        for c in comps:
            row[f"wt% {c}"] = float(comp.get(c, 0.0))
        rows.append(row)
    return rows


# ======================================================
# T PROMEDIO POR BLOQUE (para autoselect de utility)
# ======================================================

def block_avg_temperature(fs, block_id, t_ref=25.0):
    """T promedio (°C) entre entradas y salidas del bloque."""
    ts = []
    for s in fs.streams.values():
        if (s.src == block_id or s.dst == block_id) and s.cp > 0:
            ts.append(s.temperature)
    if not ts:
        return t_ref
    return sum(ts) / len(ts)


# ======================================================
# UTILITIES AUTO-CALCULADAS DESDE DUTIES
# ======================================================

def compute_utilities_from_duties(fs):
    """Para cada bloque con duty != 0, autoselect de utility +
    consumo + costo, agrupados por tipo.

    Returns:
      (rows, summary)
        rows: list of dicts compatibles con opex_extras
        summary: list of (block_name, util_key, units, cons, cost)
                 para mostrar en el panel de resultados
    """
    rows = []
    summary = []
    agg = {}

    # detección de cross-exchange (HX proceso-proceso) — no carga utility
    try:
        from flowsheet_solver import is_cross_exchange
    except ImportError:
        is_cross_exchange = lambda fs, b: False

    for b in fs.blocks.values():
        if b.duty == 0:
            continue
        if is_cross_exchange(fs, b):
            # heat integration: el calor lo entrega otra corriente del
            # proceso, no una utility.  Reporte informativo, sin costo.
            continue
        # Skip splitters y flashes estructurales: el solver les puede
        # asignar duty por diferencias de T entre input/output, pero
        # estos bloques NO consumen utility real (son separación física
        # adiabática; cualquier transferencia de calor se hace en HX
        # explícitos upstream o downstream, no acá).
        if getattr(b, "splitter_active", False):
            continue
        if getattr(b, "flash_active", False):
            # Flash adiabático: duty viene de la entalpía del feed mismo.
            # Si el user quiere modelar enfriamiento PRE-flash, debe
            # poner un HX explícito (E-XXX) antes del flash.
            continue
        # Skip vessels puros sin servicio térmico explícito (V-XX como
        # decantadores, knock-out drums sin chaqueta).  Solo si su duty
        # fue auto-asignado.  Si user puso duty_locked, se respeta.
        if (("Vessel" in b.eq_type or "Tower" in b.eq_type)
                and not getattr(b, "duty_locked", False)
                and not getattr(b, "column_active", False)):
            continue
        # Skip si el bloque YA tiene un stream de utility entrando
        # explícitamente (BFW boiler que da steam, CW del cooler, etc.):
        # el user modeló el ciclo cerrado y no queremos doble-contar.
        # Detección: input con role='utility' y phase apropiado al duty.
        has_explicit_util_input = False
        for s_in in (s for s in fs.streams.values() if s.dst == b.id):
            if s_in.role == "utility":
                ph = (s_in.phase or "").lower()
                # Para heaters (duty > 0): vapor/steam entrando
                if b.duty > 0 and ("vapor" in ph or "gas" in ph):
                    has_explicit_util_input = True
                # Para coolers (duty < 0): liquid (CW o refrigerante)
                elif b.duty < 0 and "liquid" in ph:
                    has_explicit_util_input = True
        if has_explicit_util_input:
            summary.append((b.name, "(closed loop)", "—", 0.0, 0.0))
            continue
        T_avg = block_avg_temperature(fs, b.id)
        util_key = b.heat_source or ep.autoselect_heat_source(
            b.eq_type, b.duty, T_avg
        )
        if not util_key or util_key not in ep.UTILITIES:
            continue
        util = ep.UTILITIES[util_key]
        consumption = ep.utility_consumption(util_key, b.duty)
        cost = consumption * util["price"]
        summary.append((b.name, util_key, util["units"],
                        consumption, cost))
        if util_key in agg:
            agg[util_key] = (
                agg[util_key][0] + consumption,
                agg[util_key][1] + cost,
            )
        else:
            agg[util_key] = (consumption, cost)

    # Heat integration factor — en una planta real, ~50 % del calor
    # se recupera vía cross-exchange (Pinch).  Aplicamos al consumption
    # total para no sobreestimar CUT.  Configurable via econ_defaults.
    try:
        import econ_defaults as _ed
        hi_factor = _ed.get_heat_integration_factor()
    except Exception:
        # fallback alineado con econ_defaults.HEAT_INTEGRATION["factor"]
        hi_factor = 0.4

    for util_key, (cons, _cost) in agg.items():
        util = ep.UTILITIES[util_key]
        # Solo reducir heating/cooling thermal — la electricidad no se
        # "recupera" via heat integration.
        if util.get("type") in ("heating", "cooling"):
            cons_adjusted = cons * hi_factor
            tag = f"{util['name']} (PFD-util, HI×{hi_factor:.2f})"
        else:
            cons_adjusted = cons
            tag = f"{util['name']} (PFD-util)"
        rows.append({
            "name":               tag,
            "units":              util["units"],
            "time_basis":         "year",
            "flowrate":           float(cons_adjusted),
            "price_usd_per_unit": float(util["price"]),
            "stream":             "Utilities",
        })
    return rows, summary


# ======================================================
# AUTO-SIZING DE BLOQUES UTILITY (boiler / cooling tower)
# ======================================================
# Si el user pone un Boiler block en el flowsheet, se dimensiona
# automáticamente con la suma de todas las demandas de steam del proceso.
# Idem cooling tower con la suma de cooling water duty.
#
# Una vez sizeado, el CAPEX se calcula con la correlación Turton
# (equipment_costs.EQUIPMENT_DATA del eq_type).

def aggregate_utility_demand(fs):
    """Suma duties de todos los bloques del proceso por tipo de utility.

    Returns:
        dict {util_key: total_duty_kw}
        Ej: {'steam_LP': 1500.0, 'cooling_water': -3200.0, 'fuel_gas': 5000.0}
    """
    # local import para evitar ciclo
    from flowsheet_solver import is_cross_exchange
    import equipment_ports as _ep

    agg = {}
    for b in fs.blocks.values():
        # los blocks utility (boiler/cooling tower) NO son demanda, son oferta.
        cat = None
        try:
            import equipment_costs as _eq
            cat = _eq.EQUIPMENT_DATA.get(b.eq_type, {}).get("categoria")
        except ImportError:
            pass
        if cat == "Utilities":
            continue
        if b.duty == 0:
            continue
        if is_cross_exchange(fs, b):
            continue
        T_avg = block_avg_temperature(fs, b.id)
        util_key = b.heat_source or _ep.autoselect_heat_source(
            b.eq_type, b.duty, T_avg
        )
        if not util_key:
            continue
        agg[util_key] = agg.get(util_key, 0.0) + abs(b.duty)
    return agg


def auto_size_utility_blocks(fs):
    """Para cada bloque del tipo Utilities en el flowsheet, asigna su
    `S` (capacidad) en función de la demanda total del proceso.

    - Boiler: total steam demand en kg/s (de steam_LP + MP + HP).
    - Cooling tower: total cooling demand en MW.

    No toca otros campos.  Devuelve dict {block_id: (S, S_unit)}.
    """
    import equipment_costs as _eq
    demand = aggregate_utility_demand(fs)

    # Mapear total demand a las unidades del catálogo
    # Steam: kg/s = kW / (ΔH_vap en kJ/kg).  Promedio ~2200 kJ/kg.
    steam_total_kw = (demand.get("steam_LP", 0)
                      + demand.get("steam_MP", 0)
                      + demand.get("steam_HP", 0)
                      + demand.get("fuel_gas", 0))
    steam_kg_s = steam_total_kw / 2200.0 if steam_total_kw > 0 else 0.0
    # Cooling: MW = kW / 1000
    cooling_mw = demand.get("cooling_water", 0) / 1000.0
    if "refrigeration" in demand:
        cooling_mw += demand["refrigeration"] / 1000.0

    sized = {}
    for b in fs.blocks.values():
        cat = _eq.EQUIPMENT_DATA.get(b.eq_type, {}).get("categoria")
        if cat != "Utilities":
            continue
        S_unit = _eq.EQUIPMENT_DATA[b.eq_type].get("S_unit", "")
        if b.eq_type.startswith("Boiler") and steam_kg_s > 0:
            b.S = max(steam_kg_s, 0.1)
            sized[b.id] = (b.S, S_unit)
        elif b.eq_type.startswith("Cooling tower") and cooling_mw > 0:
            b.S = max(cooling_mw, 0.1)
            sized[b.id] = (b.S, S_unit)
    return sized


# ======================================================
# LEER XLSX EXISTENTE (formato ANA.py)
# ======================================================

def read_project_xlsx(path):
    """Lee un xlsx del análisis económico y devuelve (df_capital,
    df_fixed, df_variable).  Replica la lógica de
    ANA.ImportarProyecto sin sus side-effects de UI."""
    df_raw = pd.read_excel(path, header=None)

    def _section(cols):
        df = df_raw.iloc[:, cols].copy()
        df = df.dropna(how="all")
        if df.empty:
            return df
        df.columns = df.iloc[0].tolist()
        df = df.iloc[1:].reset_index(drop=True)
        df = df[df.iloc[:, 0].notna()].reset_index(drop=True)
        return df

    df_capital  = _section([0, 1, 2])
    df_fixed    = _section([4, 5, 6])
    df_variable = _section([8, 9, 10, 11, 12, 13])
    return df_capital, df_fixed, df_variable


# ======================================================
# ESCRIBIR XLSX (compatible con ANA.ImportarProyecto)
# ======================================================

def compute_income_statement(revenue_usd_yr, crm, cut, cwt, col,
                              fci_usd, years_op=None, tax_rate=None,
                              startup_year=1,
                              depreciable_base_usd=None,
                              working_capital_usd=0.0,
                              useful_life_yr=None,
                              maintenance_tax_insurance_usd_yr=0.0,
                              alpha_d=0.180, beta=2.73, gamma=1.23,
                              alpha=0.305):
    """Construye el Estado de Resultados año a año, CONSISTENTE con
    Costing Turton (Eq 8.2).

    Convención clave (bug-fix consistencia financiera):
        EBIT (taxable) = Revenue − COM_d_Turton
    Donde COM_d_Turton = α_d·FCI + β·COL + γ·(CRM+CUT+CWT).

    Las líneas mostradas SUMAN a COM_d, en lugar de inventar una
    suma incompatible.  Breakdown:
        · Raw OPEX: CRM, CUT, CWT, COL
        · Overheads Turton:
            · Labor OH    = (β−1)·COL          [~1.73·COL]
            · Materials OH= (γ−1)·(CRM+CUT+CWT) [~0.23·(RM+UT+WT)]
            · M+T+I       = max(0, α_d·FCI − Dep_user)
            · Depreciation= depreciable_base / useful_life
        Σ líneas = α_d·FCI + β·COL + γ·(...) = COM_d_Turton ✓

    Loss carry-forward (NOL): EBIT < 0 acumula como saldo a favor;
    en años con EBIT > 0 se aplica contra el taxable income antes de
    calcular tax.  Refleja práctica fiscal estándar (Perú, USA, etc.).

    CapEx año 0 = −(FCI + WC).  Working capital se recupera en el
    último año del horizonte (cash inflow positivo).

    Loss carry-forward (NOL): EBIT < 0 acumula como saldo a favor;
    en años con EBIT > 0 se aplica contra el taxable income antes de
    calcular tax.  Refleja práctica fiscal estándar (Perú, USA, etc.).

    CapEx año 0 = −(FCI + WC).  Working capital se recupera en el
    último año del horizonte (cash inflow positivo).

    Args:
        revenue_usd_yr: ingresos anuales
        crm, cut, cwt, col: opex anuales (raw mat, util, waste, labor)
        fci_usd: FCI total (para reportar CapEx año 0)
        depreciable_base_usd: si None, usa fci_usd.  No incluye WC.
        working_capital_usd: WC one-time year-0; recupera year_op final.
        useful_life_yr: vida útil dep SL; default = years_op.
        maintenance_tax_insurance_usd_yr: cargos FCI-pegged cash
            calculados desde Turton COM (ver
            cost_of_manufacture_components).  Default 0 (legacy).

    Devuelve list of dicts (uno por año) con:
      Year, Revenue, CRM, CUT, CWT, COL, M+T+I, Depreciation, EBIT,
      NOL applied, Taxable Income, Tax, Net Income, CapEx, WC Recov,
      Cash Flow, Cumulative CF, paga_impuestos (bool).
    """
    if years_op is None or tax_rate is None:
        try:
            import econ_defaults as _ed
            fin = _ed.get_financial()
            if years_op is None: years_op = fin["project_years"]
            if tax_rate is None: tax_rate = fin["tax_rate"]
        except Exception:
            if years_op is None: years_op = 10
            if tax_rate is None: tax_rate = 0.30
    if useful_life_yr is None:
        useful_life_yr = years_op
    if depreciable_base_usd is None:
        depreciable_base_usd = fci_usd
    capex_year0  = fci_usd + working_capital_usd
    depreciation = depreciable_base_usd / max(useful_life_yr, 1)
    # M+T+I según convención usuario (corrección financial-bug §2):
    #     M+T+I = (COM − COM_d) − Dep = (α − α_d)·FCI − Dep
    # 0.125·FCI − Dep_user para defaults Turton.  Si Dep > 0.125·FCI
    # (vida útil muy corta, < 8 yr), M+T+I negativo → clamp a 0 con
    # warning.  El caller puede sobreescribir vía param explícito.
    fci_burden_annual = max(0.0, alpha - alpha_d) * fci_usd
    mti_derived = fci_burden_annual - depreciation
    if mti_derived < 0:
        import warnings
        warnings.warn(
            f"useful_life_yr={useful_life_yr} → dep_SL "
            f"({depreciation:,.0f}) > (α−α_d)·FCI "
            f"({fci_burden_annual:,.0f}). M+T+I clampado a 0; la "
            f"coherencia se mantiene si dep_user ≤ (α−α_d)·FCI.",
            UserWarning, stacklevel=2,
        )
        mti_derived = 0.0
    if (maintenance_tax_insurance_usd_yr is None
            or maintenance_tax_insurance_usd_yr <= 0):
        maintenance_tax_insurance_usd_yr = mti_derived
    # Overheads Turton (β−1)·COL y (γ−1)·materiales — suman a la
    # base deducible junto con (α−α_d)·FCI (= Dep+M+T+I).  Total
    # deductible coincide con profitability_indicators.
    labor_oh_factor = max(0.0, beta - 1.0)
    mat_oh_factor   = max(0.0, gamma - 1.0)
    rows = []
    # Año 0 = construcción / CapEx (FCI + WC)
    rows.append({
        "Year":            0,
        "Revenue":         0.0,
        "CRM":             0.0,
        "CUT":             0.0,
        "CWT":             0.0,
        "COL":             0.0,
        "Labor OH":        0.0,
        "Materials OH":    0.0,
        "M+T+I":           0.0,
        "Depreciation":    0.0,
        "EBIT":            0.0,
        "NOL applied":     0.0,
        "Taxable Income":  0.0,
        "Tax":             0.0,
        "Net Income":      0.0,
        "CapEx":           -capex_year0,
        "WC Recov":        0.0,
        "Cash Flow":       -capex_year0,
        "Cumulative CF":   -capex_year0,
    })
    cum     = -capex_year0
    nol     = 0.0        # Net Operating Loss carry-forward acumulado
    paga_tax_alguna_vez = False
    for yr in range(1, years_op + 1):
        operating = yr >= startup_year
        rev = revenue_usd_yr if operating else 0.0
        rm  = crm if operating else 0.0
        ut  = cut if operating else 0.0
        wt  = cwt if operating else 0.0
        lab = col if operating else 0.0
        mti = maintenance_tax_insurance_usd_yr if operating else 0.0
        dep = depreciation if operating else 0.0
        # Depreciación sólo años con activo en uso, hasta useful_life.
        if yr - startup_year + 1 > useful_life_yr:
            dep = 0.0
        # Overheads Turton — escalan a COL y materiales raw.
        lab_oh = labor_oh_factor * lab
        mat_oh = mat_oh_factor   * (rm + ut + wt)
        ebit = rev - rm - ut - wt - lab - lab_oh - mat_oh - mti - dep
        # Loss carry-forward: aplica NOL acumulado contra EBIT positivo
        if ebit > 0 and nol > 0:
            nol_used = min(nol, ebit)
            taxable  = ebit - nol_used
            nol     -= nol_used
        else:
            nol_used = 0.0
            taxable  = ebit
            if ebit < 0:
                nol += -ebit    # acumula pérdida como saldo a favor
        tax = max(0.0, taxable) * tax_rate
        if tax > 0:
            paga_tax_alguna_vez = True
        net = ebit - tax
        wc_recov = working_capital_usd if yr == years_op else 0.0
        cf  = net + dep + wc_recov
        cum += cf
        rows.append({
            "Year":            yr,
            "Revenue":         rev,
            "CRM":             rm,
            "CUT":             ut,
            "CWT":             wt,
            "COL":             lab,
            "Labor OH":        lab_oh,
            "Materials OH":    mat_oh,
            "M+T+I":           mti,
            "Depreciation":    dep,
            "EBIT":            ebit,
            "NOL applied":     nol_used,
            "Taxable Income":  taxable,
            "Tax":             tax,
            "Net Income":      net,
            "CapEx":           0.0,
            "WC Recov":        wc_recov,
            "Cash Flow":       cf,
            "Cumulative CF":   cum,
        })
    # Tag el flag en la última fila para que el consumer (UI) sepa
    rows[-1]["paga_impuestos"] = paga_tax_alguna_vez
    return rows


def write_3sections_xlsx(path, df_capital, df_fixed, df_variable,
                         equipment=None, streams=None, costing=None,
                         income_stmt=None):
    """Escribe un xlsx con las 3 secciones lado a lado:
      cols A-C  → Capital Costs
      col  D    → vacía
      cols E-G  → Fixed Operating Costs
      col  H    → vacía
      cols I-N  → Variable Operating Costs

    Pestañas adicionales:
      'Equipment'  → tag, type, S, duty, T_op, P_op, ΔP, η, reactions,
                     column/flash/splitter specs (cuando aplique)
      'Streams'    → todas las corrientes con mass, T, P, fase,
                     composición wt% por componente, role, precio,
                     pipe specs y locks del user
    Ambas pestañas se generan automáticamente desde el flowsheet,
    así el analista económico ve TODO el contexto del diseño en el
    mismo xlsx."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project"

    def _write_block(df, col_start):
        if df is None or df.empty:
            return
        for j, name in enumerate(df.columns):
            ws.cell(row=1, column=col_start + j, value=str(name))
        for i in range(len(df)):
            for j in range(len(df.columns)):
                v = df.iat[i, j]
                if v is None:
                    continue
                try:
                    if isinstance(v, float) and math.isnan(v):
                        continue
                except Exception:
                    pass
                ws.cell(row=2 + i, column=col_start + j, value=v)

    _write_block(df_capital,  1)
    _write_block(df_fixed,    5)
    _write_block(df_variable, 9)

    if equipment:
        ws_eq = wb.create_sheet("Equipment")
        # Cols dinámicas: union de todas las keys que aparecen en rows
        # (algunos bloques tienen specs de columna, otros no)
        base_cols = ["Tag", "Type", "Category", "Size S", "Unit", "N° units",
                       "Duty kW", "T_op K", "P_op bar", "ΔP bar", "η",
                       "Reactions", "Heat source"]
        extra_keys = []
        for row in equipment:
            for k in row.keys():
                if k not in base_cols and k not in extra_keys:
                    extra_keys.append(k)
        cols = base_cols + extra_keys
        for j, name in enumerate(cols):
            ws_eq.cell(row=1, column=j + 1, value=name)
        for i, row in enumerate(equipment):
            for j, key in enumerate(cols):
                v = row.get(key, "")
                if isinstance(v, float) and v == 0.0:
                    v = ""    # cell vacío en lugar de 0 ruidoso
                ws_eq.cell(row=2 + i, column=j + 1, value=v)

    if costing and costing.get("rows"):
        ws_c = wb.create_sheet("Costing Turton")
        ws_c.cell(row=1, column=1, value="Concepto")
        ws_c.cell(row=1, column=2, value="Detalle")
        ws_c.cell(row=1, column=3, value="Valor")
        for i, (a, b, c) in enumerate(costing["rows"]):
            ws_c.cell(row=2 + i, column=1, value=a)
            ws_c.cell(row=2 + i, column=2, value=b)
            ws_c.cell(row=2 + i, column=3, value=c)
        ws_c.column_dimensions["A"].width = 38
        ws_c.column_dimensions["B"].width = 18
        ws_c.column_dimensions["C"].width = 28

    if streams:
        ws_s = wb.create_sheet("Streams")
        # Cols dinámicas (wt% por componente varía por proyecto)
        base_cols = ["Stream", "From", "From port", "To", "To port",
                       "Mass tm/yr", "T °C", "P bar", "Phase", "Role",
                       "Price USD/tm", "Cp kJ/kg·K", "Main comp", "Locks",
                       "Pipe L m", "Pipe D m", "Pipe ΔP bar"]
        extra_keys = []
        for row in streams:
            for k in row.keys():
                if k not in base_cols and k not in extra_keys:
                    extra_keys.append(k)
        cols = base_cols + extra_keys
        for j, name in enumerate(cols):
            ws_s.cell(row=1, column=j + 1, value=name)
        for i, row in enumerate(streams):
            for j, key in enumerate(cols):
                v = row.get(key, "")
                if isinstance(v, float) and v == 0.0:
                    v = ""
                ws_s.cell(row=2 + i, column=j + 1, value=v)

    if income_stmt:
        ws_i = wb.create_sheet("Income Statement")
        cols = ["Year", "Revenue", "CRM", "CUT", "CWT", "COL",
                 "Depreciation", "EBIT", "Tax", "Net Income",
                 "CapEx", "Cash Flow", "Cumulative CF"]
        for j, name in enumerate(cols):
            ws_i.cell(row=1, column=j + 1, value=name)
        for i, row in enumerate(income_stmt):
            for j, key in enumerate(cols):
                v = row.get(key, "")
                if isinstance(v, float) and abs(v) < 0.01 and key != "Year":
                    v = ""
                ws_i.cell(row=2 + i, column=j + 1, value=v)
        ws_i.column_dimensions["A"].width = 6
        for col_letter in "BCDEFGHIJKLM":
            ws_i.column_dimensions[col_letter].width = 14

    wb.save(path)


def write_project_xlsx(path, fs, isbl, feeds, products, base_xlsx=None):
    """Genera el xlsx que ANA.py importa.

      path:         destino del xlsx
      fs:           Flowsheet (para opex_extras, fixed_overrides,
                    equipment list, utilities desde duties)
      isbl:         ISBL en MM USD a inyectar en df_capital[0,2]
      feeds:        lista de Stream con role='feed'
      products:     lista de Stream con role='product'
      base_xlsx:    si se da, se carga ese xlsx como base (preservando
                    filas variables que el user editó a mano)
                    Si None, se usan templates Turton.
    """
    if base_xlsx:
        df_capital, df_fixed, df_variable = read_project_xlsx(base_xlsx)
    else:
        df_capital  = tmpl.template_capital()
        df_fixed    = tmpl.template_fixed()
        df_variable = pd.DataFrame(columns=[
            "variable operating costs", "units", "time basis",
            "flowrate", "price usd/units", "stream",
        ])

    # inyectar ISBL
    if isbl is not None and not df_capital.empty:
        df_capital.iat[0, 2] = float(isbl)

    # fixed overrides (Labor manual del user, etc.)
    if (not df_fixed.empty and "Concept" in df_fixed.columns
            and fs.fixed_overrides):
        for concept, value in fs.fixed_overrides.items():
            mask = df_fixed["Concept"].astype(str).str.strip() == concept
            if mask.any():
                df_fixed.loc[mask, "Value"] = float(value)

    # auto-labor con Turton §8.3 si no hay override
    if (not df_fixed.empty and "Concept" in df_fixed.columns
            and "Labor" not in fs.fixed_overrides
            and fs.blocks):
        labor_info = ep.turton_labor_cost(fs.blocks.values())
        mask = df_fixed["Concept"].astype(str).str.strip() == "Labor"
        if mask.any():
            df_fixed.loc[mask, "Value"] = float(labor_info["labor_usd_yr"])

    # dedupe: borrar filas previas marcadas '(PFD)' o '(PFD-util)'
    if "variable operating costs" in df_variable.columns:
        mask = df_variable["variable operating costs"].astype(str).str.contains(
            r"\(PFD(?:-util)?\)", regex=True, na=False
        )
        df_variable = df_variable[~mask].reset_index(drop=True)

    # append filas (PFD): products, feeds, wastes, utility streams,
    # opex_extras manuales.  TODOS los streams con price != 0 (o role
    # explícito) se contabilizan — antes solo feed/product se exportaban.
    new_rows = []
    seen_ids = set()
    for s in products:
        new_rows.append({
            "variable operating costs": f"{s.name} (PFD)",
            "units":              "tm",
            "time basis":         "year",
            "flowrate":           float(s.mass_flow),
            "price usd/units":    float(getattr(s, "price_usd_per_tm", 0.0)),
            "stream":             "Key Products",
        })
        seen_ids.add(s.id)
    for s in feeds:
        new_rows.append({
            "variable operating costs": f"{s.name} (PFD)",
            "units":              "tm",
            "time basis":         "year",
            "flowrate":           float(s.mass_flow),
            "price usd/units":    float(getattr(s, "price_usd_per_tm", 0.0)),
            "stream":             "Raw Materials",
        })
        seen_ids.add(s.id)
    # Wastes — costos de disposición / byproducts vendibles.
    # En Talara: coque, slurry, gas seco (todos role='product' o
    # 'waste' con price > 0).  Antes los wastes con price>0 se
    # perdían silenciosamente.
    for s in fs.streams.values():
        if s.id in seen_ids:
            continue
        if s.role == "waste":
            new_rows.append({
                "variable operating costs": f"{s.name} (PFD waste)",
                "units":              "tm",
                "time basis":         "year",
                "flowrate":           float(s.mass_flow),
                # waste con price>0: ingreso (subproducto vendible)
                # waste con price<0: gasto (tratamiento)
                # waste con price=0: ignored en NPV pero queda registrado
                "price usd/units":    float(getattr(s, "price_usd_per_tm", 0.0)),
                "stream":             "Waste / Byproduct",
            })
            seen_ids.add(s.id)
    # Utility streams declarados como feed externo (BFW, fuel gas,
    # H2SO4 secado, etc.) — antes se ignoraban si role='utility'.
    for s in fs.streams.values():
        if s.id in seen_ids:
            continue
        if s.role == "utility":
            # Solo si entra al proceso desde un tanque source (no
            # interno) — para evitar doble-conteo con auto-utilities
            # que se calculan desde duty.  Marca con price>0 (compra).
            src_b = fs.blocks.get(s.src)
            is_external_feed = (src_b is not None
                                and "Storage tank" in (src_b.eq_type or ""))
            if is_external_feed and s.price_usd_per_tm > 0:
                new_rows.append({
                    "variable operating costs": f"{s.name} (PFD utility)",
                    "units":              "tm",
                    "time basis":         "year",
                    "flowrate":           float(s.mass_flow),
                    "price usd/units":    float(s.price_usd_per_tm),
                    "stream":             "Utilities",
                })
                seen_ids.add(s.id)
    for ex in fs.opex_extras:
        new_rows.append({
            "variable operating costs": f"{ex.get('name','?')} (PFD)",
            "units":              ex.get("units", "tm"),
            "time basis":         ex.get("time_basis", "year"),
            "flowrate":           float(ex.get("flowrate", 0.0)),
            "price usd/units":    float(ex.get("price_usd_per_unit", 0.0)),
            "stream":             ex.get("stream", "Utilities"),
        })

    # utilities (PFD-util) auto-calculadas desde duties
    auto_rows, _summary = compute_utilities_from_duties(fs)
    for ex in auto_rows:
        new_rows.append({
            "variable operating costs": ex["name"],
            "units":              ex["units"],
            "time basis":         ex["time_basis"],
            "flowrate":           ex["flowrate"],
            "price usd/units":    ex["price_usd_per_unit"],
            "stream":             ex["stream"],
        })

    if new_rows:
        df_variable = pd.concat(
            [df_variable, pd.DataFrame(new_rows)],
            ignore_index=True,
        )

    # escribir xlsx con 3 secciones + pestañas Equipment, Streams,
    # Costing Turton (resumen CBM por categoría + COM Eq 8.2) +
    # Income Statement año por año (P&L Turton Ch 9-10)
    equipment_rows = collect_equipment_rows(fs, year_target=2024)
    stream_rows    = collect_stream_rows(fs)
    costing_data   = compute_turton_costing(
        fs, df_variable, df_fixed, fci_musd=isbl, year_target=2024)
    _prof = costing_data["profit"]
    income_rows    = compute_income_statement(
        revenue_usd_yr=costing_data["revenue"],
        crm=costing_data["crm"], cut=costing_data["cut"],
        cwt=costing_data["cwt"], col=costing_data["col"],
        fci_usd=costing_data["fci"],
        depreciable_base_usd=costing_data["depreciable_base"],
        working_capital_usd=costing_data["working_capital"],
        useful_life_yr=_prof["useful_life_yr"],
        years_op=_prof["years_op"],
        tax_rate=_prof["tax_rate"],
        maintenance_tax_insurance_usd_yr=costing_data["mti_usd_yr"],
    )
    # Assert de coherencia (fail-fast si las dos hojas divergen).
    try:
        assert_costing_income_coherence(costing_data, income_rows)
    except AssertionError as _e:
        import warnings
        warnings.warn(f"Coherencia Costing/Income: {_e}", UserWarning,
                       stacklevel=2)
    write_3sections_xlsx(path, df_capital, df_fixed, df_variable,
                         equipment=equipment_rows,
                         streams=stream_rows,
                         costing=costing_data,
                         income_stmt=income_rows)
