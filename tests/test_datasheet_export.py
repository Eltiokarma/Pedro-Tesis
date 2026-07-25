"""tests/test_datasheet_export.py — la ficha sale del programa (tandas 3-4).

Lo que se congela acá:

  1. UN SOLO CONTENIDO — el aplanado `datasheet_rows` es la fuente de las
     dos exportaciones.  Si alguien agrega un campo a una y no a la otra,
     el test de paridad lo caza.
  2. COBERTURA — el libro XLSX tiene una hoja por equipo de CUALQUIER
     ejemplo, incluidos los tipos sin catálogo (el fallback genérico del
     agregador garantiza contenido para los 60 tipos).
  3. EXPORTAR NO MUEVE LA FÍSICA — el golden del ejemplo es idéntico antes
     y después de exportar.  Es la misma restricción que la ficha ya tiene
     en el inspector, y hay que sostenerla en el export.
  4. EL VEREDICTO NO AFIRMA DE MÁS — un tipo sin catálogo explica por qué
     con la razón de ingeniería de equipos_a_pedido.json, no con la frase
     vieja que decía que "ningún fabricante publica" (falsa para varios
     tipos, ver la corrección de la cosecha 2026-07).
"""
import os
import sys
import tempfile
import unittest

_PARENT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PARENT not in sys.path:
    sys.path.insert(0, _PARENT)

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import datasheet as ds
import datasheet_export as dx

try:
    import openpyxl                                  # noqa: F401
    _OPENPYXL = True
except Exception:                                    # pragma: no cover
    _OPENPYXL = False

try:
    import PySide6.QtGui                             # noqa: F401
    _QT = True
except Exception:                                    # pragma: no cover
    _QT = False

_REVS = [
    {"rev": "A", "desc": "Emisión", "date": "24-07", "by": "GN"},
    {"rev": "B", "desc": "+ selector comercial", "date": "25-07", "by": "GN"},
]


def _ejemplo(clave):
    from export_examples import _headless_mocks
    _headless_mocks()
    import examples_registry as reg
    import flowsheet_solver as fsv
    fs = reg.load_example(clave)
    fsv.solve(fs)
    return fs


class TestFilas(unittest.TestCase):
    """El aplanado — la parte Qt-free y la que ambos formatos comparten."""

    def test_secciones_en_orden_y_sin_vacios(self):
        fs = _ejemplo("boiler_ft")
        for b in dx.bloques_de(fs):
            filas = dx.datasheet_rows(ds.datasheet_spec(b, fs))
            self.assertTrue(filas, f"{b.name}: ficha sin filas")
            for seccion, etiqueta, valor, _nota in filas:
                self.assertTrue(seccion.strip(), f"{b.name}: sección vacía")
                self.assertTrue(etiqueta.strip(), f"{b.name}: etiqueta vacía")
                self.assertTrue(str(valor).strip(),
                                f"{b.name}/{etiqueta}: valor vacío")
            # Una sección no puede reaparecer después de cerrarse: los dos
            # renderers dibujan el encabezado al cambiar de sección, así que
            # un orden intercalado imprimiría el mismo título dos veces.
            vistas, previa = [], None
            for seccion, *_ in filas:
                if seccion != previa:
                    self.assertNotIn(seccion, vistas,
                                     f"{b.name}: '{seccion}' intercalada")
                    vistas.append(seccion)
                    previa = seccion

    def test_identidad_siempre_presente(self):
        """El fallback del agregador garantiza identidad para todo tipo."""
        fs = _ejemplo("pfr")
        for b in dx.bloques_de(fs):
            filas = dx.datasheet_rows(ds.datasheet_spec(b, fs))
            tags = [v for s, e, v, _ in filas if s == "Identidad" and e == "Tag"]
            self.assertEqual(tags, [b.name])

    def test_auxiliares_fuera(self):
        """auto_aux es el lazo de servicio que arma el simulador, no un
        equipo que alguien compre: no lleva ficha."""
        fs = _ejemplo("cooling")
        for b in dx.bloques_de(fs):
            self.assertFalse(getattr(b, "auto_aux", False), b.name)

    def test_sin_catalogo_da_razon_no_frase_vieja(self):
        """Hallazgo del export: `verificacion()['mensaje']` seguía con la
        frase que la cosecha ya había corregido en la UI."""
        fs = _ejemplo("pfr")
        pfr = [b for b in fs.blocks.values()
               if not ds.entradas_para(b.eq_type)]
        self.assertTrue(pfr, "el ejemplo debería traer un tipo sin catálogo")
        for b in pfr:
            v = ds.verificacion(b, fs)
            self.assertEqual(v["estado"], "no_aplica")
            self.assertNotIn("ningún fabricante publica", v["mensaje"])
            motivo = ds.motivo_a_pedido(b.eq_type)
            self.assertIn(motivo["titulo"], v["mensaje"])


@unittest.skipUnless(_OPENPYXL, "openpyxl no disponible")
class TestXlsx(unittest.TestCase):

    def test_una_hoja_por_equipo_mas_indice(self):
        import openpyxl
        fs = _ejemplo("boiler_ft")
        equipos = dx.bloques_de(fs)
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "fichas.xlsx")
            n = dx.write_datasheets_xlsx(path, fs, proyecto="boiler_ft")
            self.assertEqual(n, len(equipos))
            wb = openpyxl.load_workbook(path)
            self.assertEqual(wb.sheetnames[0], "Índice")
            self.assertEqual(len(wb.sheetnames), len(equipos) + 1)
            for b in equipos:
                self.assertIn(b.name, wb.sheetnames)
            idx = wb["Índice"]
            self.assertEqual(idx["A3"].value, "Tag")
            tags = {idx.cell(row=r, column=1).value
                    for r in range(4, 4 + len(equipos))}
            self.assertEqual(tags, {b.name for b in equipos})

    def test_nombre_de_hoja_saneado_y_unico(self):
        """Excel rechaza []:*?/\\ y >31 caracteres, y no admite duplicados."""
        usados = set()
        n1 = dx._nombre_hoja("E-101/A:B*C", usados)
        self.assertNotIn("/", n1)
        self.assertNotIn(":", n1)
        largo = dx._nombre_hoja("X" * 60, usados)
        self.assertLessEqual(len(largo), 31)
        gemelo = dx._nombre_hoja("X" * 60, usados)
        self.assertLessEqual(len(gemelo), 31)
        self.assertNotEqual(largo, gemelo)

    def test_exportar_no_mueve_la_fisica(self):
        from export_examples import golden
        import flowsheet_solver as fsv
        fs = _ejemplo("beer")
        g0 = golden(fs, fsv.solve(fs))
        with tempfile.TemporaryDirectory() as d:
            dx.write_datasheets_xlsx(os.path.join(d, "f.xlsx"), fs)
        self.assertEqual(g0, golden(fs, fsv.solve(fs)))


@unittest.skipUnless(_QT, "PySide6 no disponible")
class TestPdf(unittest.TestCase):

    def test_una_pagina_por_equipo_mas_revisiones(self):
        fs = _ejemplo("boiler_ft")
        fs.revisions = [dict(r) for r in _REVS]
        equipos = dx.bloques_de(fs)
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "fichas.pdf")
            n = dx.write_datasheets_pdf(path, fs, proyecto="boiler_ft",
                                        fecha="25-07-2026")
            self.assertEqual(n, len(equipos))
            self.assertGreater(os.path.getsize(path), 1000)
            with open(path, "rb") as f:
                crudo = f.read()
            self.assertTrue(crudo.startswith(b"%PDF"))

    def test_sin_revisiones_no_emite_pagina_de_historial(self):
        """El plano sin historial no dibuja un cuadro vacío; el legajo
        tampoco agrega una página que diría nada."""
        fs = _ejemplo("boiler_ft")
        fs.revisions = []
        with tempfile.TemporaryDirectory() as d:
            sin = os.path.join(d, "sin.pdf")
            con = os.path.join(d, "con.pdf")
            dx.write_datasheets_pdf(sin, fs, fecha="25-07-2026")
            fs.revisions = [dict(r) for r in _REVS]
            dx.write_datasheets_pdf(con, fs, fecha="25-07-2026")
            self.assertLess(os.path.getsize(sin), os.path.getsize(con))

    def test_flowsheet_vacio_falla_claro(self):
        from flowsheet_model import Flowsheet
        with tempfile.TemporaryDirectory() as d:
            with self.assertRaises(ValueError):
                dx.write_datasheets_pdf(os.path.join(d, "x.pdf"), Flowsheet())

    def test_exportar_no_mueve_la_fisica(self):
        from export_examples import golden
        import flowsheet_solver as fsv
        fs = _ejemplo("beer")
        g0 = golden(fs, fsv.solve(fs))
        with tempfile.TemporaryDirectory() as d:
            dx.write_datasheets_pdf(os.path.join(d, "f.pdf"), fs,
                                    fecha="25-07-2026")
        self.assertEqual(g0, golden(fs, fsv.solve(fs)))


@unittest.skipUnless(_QT, "PySide6 no disponible")
class TestAccionesQt(unittest.TestCase):
    """El export tiene que estar ENCHUFADO, no solo existir como módulo."""

    @classmethod
    def setUpClass(cls):
        from PySide6.QtWidgets import QApplication
        cls.app = QApplication.instance() or QApplication([])

    def _ventana(self):
        import flowsheet_qt as fq
        w = fq.FlowsheetMainWindow()
        w.action_load_example("boiler_ft")
        return w

    def test_acciones_en_el_menu_exportar(self):
        w = self._ventana()
        self.assertTrue(callable(w.action_export_fichas_pdf))
        self.assertTrue(callable(w.action_export_fichas_xlsx))
        etiquetas = [a.text() for a in w.findChildren(type(w.menuBar()
                                                           .actions()[0]))]
        texto = " ".join(etiquetas)
        self.assertIn("Fichas técnicas", texto,
                      "las fichas no llegaron al menú Exportar")

    def test_proyecto_sale_del_cuadro_de_titulo(self):
        """Plano y legajo son dos documentos del mismo trabajo: el rótulo
        no se inventa, se toma del cuadro de título del Marco PFD."""
        w = self._ventana()
        pf = w.scene.paper_frame
        self.assertIsNotNone(pf, "el ejemplo debería traer Marco PFD")
        self.assertEqual(w._nombre_proyecto(), pf._project_title)
        self.assertTrue(w._nombre_proyecto())

    def test_sin_marco_no_inventa_nombre(self):
        import flowsheet_qt as fq
        w = fq.FlowsheetMainWindow()
        self.assertEqual(w._nombre_proyecto(), "")


@unittest.skipUnless(_OPENPYXL and _QT, "openpyxl y PySide6 requeridos")
class TestParidad(unittest.TestCase):
    """Los dos formatos son el MISMO documento: si divergen, es un bug."""

    def test_xlsx_contiene_exactamente_las_filas_del_aplanado(self):
        import openpyxl
        fs = _ejemplo("centrifuge")
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "f.xlsx")
            dx.write_datasheets_xlsx(path, fs, proyecto="centrifuge")
            wb = openpyxl.load_workbook(path)
            for b in dx.bloques_de(fs):
                filas = dx.datasheet_rows(ds.datasheet_spec(b, fs))
                ws = wb[b.name]
                # Las etiquetas de la hoja, salteando los encabezados de
                # sección (que van en MAYÚSCULA en la columna A).
                secciones = {s.upper() for s, _e, _v, _n in filas}
                leidas = []
                for fila in ws.iter_rows(min_row=4, max_col=2,
                                         values_only=True):
                    etiqueta, valor = fila[0], fila[1]
                    if etiqueta is None or etiqueta in secciones:
                        continue
                    leidas.append((str(etiqueta), str(valor)))
                esperadas = [(e, str(v)) for _s, e, v, _n in filas]
                self.assertEqual(leidas, esperadas,
                                 f"{b.name}: el XLSX no refleja el aplanado")


if __name__ == "__main__":
    unittest.main(verbosity=2)
