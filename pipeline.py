# ======================================================
# PIPELINE GUI -> MOTOR ECONÓMICO
# ======================================================
# Adapta los DataFrames de la GUI (df_capital, df_fixed,
# engine.df_internal) al formato que esperan CostModel /
# CashFlowModel / ReportGenerator y orquesta el cálculo
# end-to-end.
# ======================================================

import pandas as pd

from flujoflujoclass import (
    CostModel,
    CashFlowModel,
    ReportGenerator,
)

import cepci
import indicators


# ======================================================
# CONSTRUCCIÓN DEL DICT data ESPERADO POR CostModel
# ======================================================

# CostModel.calcular() usa, para consumables y utilities,
# la fórmula coef * price * production.  La GUI no captura
# coeficientes (la columna física donde estaban en el Excel
# antiguo ahora aloja "time basis"), por lo que aquí
# adoptamos la convención: cost = flowrate_SI * price_SI,
# usando coef = flowrate_SI y forzando production = 1
# mediante un key_products dummy con flow = 1 cuando hay
# consumables/utilities.  Para evitar romper la valoración
# de los productos reales, computamos production_factor a
# partir del primer key product real y lo neutralizamos
# multiplicando el coef de cada consumible/utility por su
# valor; ver _construir_data.
# ======================================================


_STREAM_TO_BUCKET = {
    "Key Products": "key_products",
    "By-products": "byproducts",
    "Waste Streams": "byproducts",
    "Raw Materials": "raw_materials",
    "Consumables": "consumables",
    "Utilities": "utilities",
}


def _valor_capital(df_capital, fila):
    return float(df_capital.iloc[fila, 2])


def _valor_fixed(df_fixed, fila):
    return float(df_fixed.iloc[fila, 2])


# ======================================================
# PARSERS POR NOMBRE (más robustos que por posición)
# ======================================================

_CAPITAL_KEYS = {
    "ISBL":     ["isbl", "battery limit"],
    "OSBL_pct": ["osbl", "off-site", "offsite"],
    "ENG_pct":  ["engineering", "eng cost", "eng "],
    "CONT_pct": ["contingency", "cont"],
    "WC_pct":   ["working capital", "working cap", "wc"],
}

_FIXED_KEYS = {
    "labor":                ["labor"],
    "supervision_pct":      ["supervision"],
    "salary_overhead_pct":  ["salary overhead", "direct salary", "salary o"],
    "maintenance_pct":      ["maintenance"],
    "plant_overhead_pct":   ["plant overhead"],
    "tax_insurance_pct":    ["tax & insurance", "tax and ins", "tax ins", "insurance"],
    "interest_pct":         ["interest"],
    "general_expenses_pct": ["general expense", "general "],
    "royalties_pct":        ["royalties", "royalty", "royal"],
}


def _parse_seccion_por_nombre(df, mapeo, fallback_zero=True):
    """Parsea una sección (capital o fixed) por keywords en
    la primera columna.  Tolera orden arbitrario y filas
    faltantes.  Si una key no se encuentra y fallback_zero=
    True, devuelve 0 para esa entry.

    df: DataFrame con [concept, units/basis, value] en cols 0/1/2.
    mapeo: dict {key_modelo: [patrones lowercase]}.
    """
    result = {}

    # Pre-procesar concepts en lowercase
    concepts = [str(df.iloc[i, 0]).strip().lower() for i in range(len(df))]
    valores  = [df.iloc[i, 2] for i in range(len(df))]

    used = set()  # índices ya usados (evita asignar la misma fila a 2 keys)

    for key, patterns in mapeo.items():
        match_idx = None
        # Buscar fila por patrones, prefiere el match más temprano
        for i, c in enumerate(concepts):
            if i in used:
                continue
            if any(p in c for p in patterns):
                match_idx = i
                break

        if match_idx is None:
            if fallback_zero:
                result[key] = 0.0
            else:
                raise ValueError(
                    f"No row found for '{key}' (patterns: {patterns})"
                )
        else:
            v = valores[match_idx]
            if pd.isna(v):
                result[key] = 0.0
            else:
                result[key] = float(v)
            used.add(match_idx)

    return result


def _pct(valor):
    """Convierte porcentaje a fracción: SIEMPRE /100.

    Convención: el Excel y los DataFrames de la GUI usan
    escala 0-100 para todos los porcentajes (OSBL %, ENG %,
    Maintenance %, etc., incluso si son chicos como
    Royalties = 0.5%).

    Antes esta función chequeaba 'si valor <= 1, ya es
    fracción' — pero ese heurístico interpretaba mal el
    caso real (Royalties 0.5% interpretado como 50% →
    inflaba CCOP en ~ Revenue/2)."""
    return float(valor) / 100.0


def _construir_data(
        df_capital,
        df_fixed,
        df_internal,
        factor_cepci=1.0,
):
    """Arma el dict data que consume CostModel.

    factor_cepci: multiplicador opcional para llevar ISBL
    del año de estimación al año objetivo (CEPCI).
    Default 1.0 (sin ajuste).

    Filas asumidas en df_capital (orden estricto):
        0: ISBL (USD)
        1: OSBL  (%)
        2: ENG   (%)
        3: CONT  (%)
        4: WC    (%)

    Filas asumidas en df_fixed (orden estricto):
        0: Labor (USD/yr)
        1: Supervision         (%)
        2: Salary Overhead     (%)
        3: Maintenance         (%)
        4: Plant Overhead      (%)
        5: Tax & Insurance     (%)
        6: Interest            (%)
        7: General Expenses    (%)
        8: Royalties           (%)
    """

    data = {}

    # CAPITAL
    # CAPITAL — parseo por nombre (más robusto que orden fijo)
    cap = _parse_seccion_por_nombre(df_capital, _CAPITAL_KEYS)
    data["ISBL"]     = cap["ISBL"] * factor_cepci
    data["OSBL_pct"] = _pct(cap["OSBL_pct"])
    data["ENG_pct"]  = _pct(cap["ENG_pct"])
    data["CONT_pct"] = _pct(cap["CONT_pct"])
    data["WC_pct"]   = _pct(cap["WC_pct"])

    # FCOP — también por nombre.  Tolera la convención del
    # Excel real (8 filas: sin General Expenses) y la
    # convención Turton estricta (9 filas).
    # labor está en USD/yr crudo; lo convertimos a MMUSD/yr
    # para que sume coherente con maintenance (% de FCI).
    fixed = _parse_seccion_por_nombre(df_fixed, _FIXED_KEYS)

    # labor: si el valor es muy pequeño (<1000) asumimos
    # que ya está en MMUSD/yr; si es grande lo dividimos.
    labor_raw = fixed["labor"]
    labor_mm = labor_raw if labor_raw <= 1000 else labor_raw / 1e6

    data["FCOP_inputs"] = {
        "labor":                labor_mm,
        "supervision_pct":      _pct(fixed["supervision_pct"]),
        "salary_overhead_pct":  _pct(fixed["salary_overhead_pct"]),
        "maintenance_pct":      _pct(fixed["maintenance_pct"]),
        "plant_overhead_pct":   _pct(fixed["plant_overhead_pct"]),
        "tax_insurance_pct":    _pct(fixed["tax_insurance_pct"]),
        "interest_pct":         _pct(fixed["interest_pct"]),
        "general_expenses_pct": _pct(fixed["general_expenses_pct"]),
        "royalties_pct":        _pct(fixed["royalties_pct"]),
    }

    # STREAMS (todos valuados como flow_SI * price_SI)
    buckets = {
        "key_products":  [],
        "byproducts":    [],
        "raw_materials": [],
        "consumables":   [],
        "utilities":     [],
    }

    for _, fila in df_internal.iterrows():

        stream = str(fila["Stream"]).strip()
        bucket = _STREAM_TO_BUCKET.get(stream)

        if bucket is None:
            continue

        flow_si  = float(fila["Flowrate SI"])
        price_si = float(fila["Price SI"])

        item = {
            "concept": str(fila["Variable"]),
            "unit":    "SI",
            "coef":    flow_si,
            "flow":    flow_si,
            "price":   price_si,
        }

        buckets[bucket].append(item)

    data.update(buckets)

    # Neutralizar production en consumables/utilities.
    # CostModel calcula CONS = sum(coef*price*production).
    # Queremos que CONS = sum(flow_SI*price_SI), por lo que
    # forzamos production = 1.  Eso se logra con un key
    # product dummy con flow=1, PERO necesitamos preservar
    # los key products reales para REV.  Solución: dividir
    # los coef de cons/uts por production_real, así
    # production_real * coef = flow_SI.
    if data["key_products"] and (data["consumables"] or data["utilities"]):
        production_real = data["key_products"][0]["flow"]
        if production_real not in (0, 0.0):
            for item in data["consumables"] + data["utilities"]:
                item["coef"] = item["coef"] / production_real

    return data


# ======================================================
# SCHEDULE
# ======================================================

def _parsear_csv_floats(texto, default):
    if texto is None or str(texto).strip() == "":
        return list(default)
    partes = [
        p.strip()
        for p in str(texto).split(",")
        if p.strip() != ""
    ]
    return [float(p) for p in partes]


def construir_params(
        inputs_economicos,
):
    """Construye el dict params que consume CashFlowModel a
    partir de los inputs económicos de la GUI."""

    schedule_raw = {
        "FC":   _parsear_csv_floats(
            inputs_economicos.get("fc_csv"),
            default=[1.0],
        ),
        "VCOP": _parsear_csv_floats(
            inputs_economicos.get("vcop_csv"),
            default=[1.0],
        ),
    }

    # tax_rate y discount_rate aceptan ambas convenciones:
    #   "0.35"  → 35%      "35" → 35%
    #   "0.15"  → 15%      "15" → 15%
    # Si el user escribe >1, asumimos escala 0-100 y dividimos.
    # Si está entre 0 y 1, asumimos fracción y lo dejamos.
    tax_raw   = float(inputs_economicos["tax_rate"])
    disc_raw  = float(inputs_economicos["discount_rate"])
    tasa_imp  = tax_raw  / 100.0 if tax_raw  > 1 else tax_raw
    tasa_int  = disc_raw / 100.0 if disc_raw > 1 else disc_raw

    return {
        "vida":          int(inputs_economicos["project_life"]),
        "tasa_impuesto": tasa_imp,
        "metodo_dep":    int(inputs_economicos["metodo_dep"]),
        "periodo_dep":   int(inputs_economicos.get("periodo_dep", 10)),
        "tipo_macrs":    int(inputs_economicos.get("tipo_macrs", 0)),
        "tasa_interes":  tasa_int,
        "schedule":      schedule_raw,
    }


# ======================================================
# EJECUCIÓN END-TO-END
# ======================================================

def _resolver_factor_cepci(inputs_economicos):
    """Lee 'cepci_year_basis' y 'cepci_year_target' de los
    inputs y devuelve el multiplicador.  Defaults: ambos
    iguales (factor=1)."""
    basis  = int(inputs_economicos.get("cepci_year_basis",  cepci.AÑO_BASE_DEFAULT))
    target = int(inputs_economicos.get("cepci_year_target", cepci.AÑO_BASE_DEFAULT))
    return cepci.factor_cepci(basis, target), basis, target


def ejecutar_analisis(
        df_capital,
        df_fixed,
        df_internal,
        inputs_economicos,
        archivo_salida,
):
    """Corre todo el pipeline y exporta el reporte Excel.

    Devuelve dict con: data, costos, cf, params, npv, irr,
    dcfror, pbp_simple, pbp_descontado, roi, cepci_factor,
    npv_at_rates, indicadores.
    """

    if df_capital is None or df_capital.empty:
        raise ValueError("Capital Costs vacío")

    if df_fixed is None or df_fixed.empty:
        raise ValueError("Fixed Operating Costs vacío")

    if df_internal is None or df_internal.empty:
        raise ValueError("Variable Operating Costs vacío")

    factor, year_basis, year_target = _resolver_factor_cepci(inputs_economicos)

    # 1) data dict + costos (con ajuste CEPCI sobre ISBL)
    data = _construir_data(
        df_capital, df_fixed, df_internal,
        factor_cepci=factor,
    )
    costos = CostModel(data).calcular()

    # 2) params + schedule completo
    params = construir_params(inputs_economicos)

    reporte = ReportGenerator()
    schedule_full = reporte.construir_schedule(
        params["schedule"],
        params["vida"],
    )
    params["schedule"] = schedule_full

    # 3) cash flow
    cf = CashFlowModel(costos, params).calcular()

    # 4) indicadores
    tasa = params["tasa_interes"]
    t_start = params["schedule"]["t_start"]
    ind = indicators.resumen(
        cf["CF"], cf["años"], costos["FCI"],
        tasa_descuento=tasa, t_start=t_start,
    )

    # 5) exportar Excel base
    reporte.exportar_cashflow(
        archivo_salida,
        cf,
        params,
        costos,
        data,
    )

    # 6) hoja extra de indicadores
    _exportar_indicadores_excel(
        archivo_salida, ind, costos, data,
        cepci_info={
            "factor": factor,
            "year_basis": year_basis,
            "year_target": year_target,
        },
    )

    return {
        "data":            data,
        "costos":          costos,
        "cf":              cf,
        "params":          params,
        "npv":             ind["NPV"],
        "irr":             ind["IRR"],
        "dcfror":          ind["DCFROR"],
        "pbp_simple":      ind["PBP_simple"],
        "pbp_descontado":  ind["PBP_descontado"],
        "roi":             ind["ROI_promedio"],
        "npv_at_rates":    ind["NPV_at_rates"],
        "indicadores":     ind,
        "cepci_factor":    factor,
        "cepci_year_basis":  year_basis,
        "cepci_year_target": year_target,
    }


# ======================================================
# HOJA EXTRA "Profitability Indicators"
# ======================================================

def _exportar_indicadores_excel(archivo, ind, costos, data, cepci_info):
    """Agrega hoja 'Profitability Indicators' con la tabla
    completa de indicadores estándar de Turton §10."""

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = load_workbook(archivo)
    if "Profitability Indicators" in wb.sheetnames:
        del wb["Profitability Indicators"]
    ws = wb.create_sheet("Profitability Indicators")

    bold       = Font(bold=True)
    title      = Font(bold=True, size=13)
    fill_head  = PatternFill(start_color="D9D9D9", fill_type="solid")
    border_top = Border(top=Side(style="medium"))
    fmt_num    = "#,##0.00"
    fmt_pct    = "0.00%"
    fmt_years  = "0.00"

    row = 1

    # ---- header ----
    ws.cell(row, 1, "PROFITABILITY INDICATORS").font = title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # ---- CEPCI ----
    ws.cell(row, 1, "Capital Cost Inflation (CEPCI)").font = bold
    row += 1
    ws.cell(row, 1, "Year basis (estimation)")
    ws.cell(row, 2, cepci_info["year_basis"])
    row += 1
    ws.cell(row, 1, "Year target (analysis)")
    ws.cell(row, 2, cepci_info["year_target"])
    row += 1
    ws.cell(row, 1, "CEPCI factor (target / basis)")
    c = ws.cell(row, 2, cepci_info["factor"])
    c.number_format = "0.000"
    row += 1
    ws.cell(row, 1, "ISBL adjusted (MM USD)")
    c = ws.cell(row, 2, data["ISBL"])
    c.number_format = fmt_num
    row += 2

    # ---- summary table ----
    ws.cell(row, 1, "Headline metrics").font = bold
    row += 1
    ws.cell(row, 1, "Metric").font = bold
    ws.cell(row, 2, "Value").font = bold
    ws.cell(row, 3, "Units").font = bold
    for c in range(1, 4):
        ws.cell(row, c).fill = fill_head
    row += 1

    def emit(label, valor, fmt=fmt_num, units=""):
        nonlocal row
        ws.cell(row, 1, label)
        if valor is None:
            ws.cell(row, 2, "n/a")
        else:
            c = ws.cell(row, 2, valor)
            c.number_format = fmt
        ws.cell(row, 3, units)
        row += 1

    emit("NPV @ project discount rate", ind["NPV"], fmt_num, "MM USD")
    emit("IRR (= DCFROR)", ind["IRR"], fmt_pct, "")
    emit("DCFROR (alias)", ind["DCFROR"], fmt_pct, "")
    emit("Payback period (simple)", ind["PBP_simple"], fmt_years, "years")
    emit("Payback period (discounted)", ind["PBP_descontado"], fmt_years, "years")
    emit("ROI (mean annual CF / FCI)", ind["ROI_promedio"], fmt_pct, "")
    row += 1

    # ---- NPV @ varias tasas ----
    ws.cell(row, 1, "NPV @ alternative discount rates").font = bold
    row += 1
    ws.cell(row, 1, "Discount rate").font = bold
    ws.cell(row, 2, "NPV (MM USD)").font = bold
    for c in range(1, 3):
        ws.cell(row, c).fill = fill_head
    row += 1

    for tasa, val in sorted(ind["NPV_at_rates"].items()):
        c1 = ws.cell(row, 1, tasa); c1.number_format = fmt_pct
        c2 = ws.cell(row, 2, val);  c2.number_format = fmt_num
        # marcar la fila de la tasa "del proyecto"
        if abs(tasa - ind["tasa_descuento"]) < 1e-9:
            c1.font = bold
            c2.font = bold
        row += 1

    row += 1

    # ---- comparativa Payback simple vs descontado ----
    ws.cell(row, 1, "Comparison: simple vs discounted payback").font = bold
    row += 1
    pbs = ind["PBP_simple"]
    pbd = ind["PBP_descontado"]
    ws.cell(row, 1, "PBP simple (years)")
    ws.cell(row, 2, pbs if pbs is not None else "n/a")
    if pbs is not None:
        ws.cell(row, 2).number_format = fmt_years
    row += 1
    ws.cell(row, 1, "PBP discounted (years)")
    ws.cell(row, 2, pbd if pbd is not None else "n/a")
    if pbd is not None:
        ws.cell(row, 2).number_format = fmt_years
    row += 1
    if pbs is not None and pbd is not None:
        ws.cell(row, 1, "Discount penalty (Δ years)")
        c = ws.cell(row, 2, pbd - pbs)
        c.number_format = fmt_years
    elif pbs is not None and pbd is None:
        ws.cell(row, 1, "Discount penalty")
        ws.cell(row, 2, "project never pays back when discounted")

    wb.save(archivo)


# ======================================================
# IRR (Newton-Raphson simple sobre NPV) — legacy alias
# ======================================================


# ======================================================
# MONTE CARLO
# ======================================================

def construir_data_y_params(
        df_capital,
        df_fixed,
        df_internal,
        inputs_economicos,
):
    """Helper: construye data + params (con schedule
    expandido) sin correr el análisis.  Útil para correr
    Monte Carlo sin recomputar el schedule cada
    iteración."""

    if df_capital is None or df_capital.empty:
        raise ValueError("Capital Costs vacío")
    if df_fixed is None or df_fixed.empty:
        raise ValueError("Fixed Operating Costs vacío")
    if df_internal is None or df_internal.empty:
        raise ValueError("Variable Operating Costs vacío")

    factor, _, _ = _resolver_factor_cepci(inputs_economicos)

    data = _construir_data(
        df_capital, df_fixed, df_internal,
        factor_cepci=factor,
    )
    params = construir_params(inputs_economicos)

    reporte = ReportGenerator()
    params["schedule"] = reporte.construir_schedule(
        params["schedule"],
        params["vida"],
    )

    return data, params


def ejecutar_montecarlo(
        df_capital,
        df_fixed,
        df_internal,
        inputs_economicos,
        variables_inciertas,
        n_runs=5000,
        seed=None,
        correlacion=None,
        archivo_salida=None,
):
    """Corre N simulaciones Monte Carlo + tornado chart.

    correlacion: None | dict {(i,j): rho} | np.ndarray KxK.
        Define la matriz de correlación entre variables
        para Gaussian copula sampling.

    Si archivo_salida es dado, agrega hojas 'Monte Carlo'
    y 'Tornado' al .xlsx existente (típicamente el reporte
    económico ya generado por ejecutar_analisis).
    """

    from montecarlo import correr_montecarlo, correr_tornado, exportar_montecarlo_excel

    data, params = construir_data_y_params(
        df_capital, df_fixed, df_internal, inputs_economicos,
    )

    mc = correr_montecarlo(
        data, params, variables_inciertas,
        n_runs=n_runs, seed=seed, correlacion=correlacion,
    )

    tornado = correr_tornado(data, params, variables_inciertas)

    if archivo_salida is not None:
        exportar_montecarlo_excel(archivo_salida, mc, tornado)

    return {
        "mc":       mc,
        "tornado":  tornado,
        "archivo":  archivo_salida,
        "data":     data,
        "params":   params,
    }


# ======================================================
# IRR (Newton-Raphson simple sobre NPV)
# ======================================================

def _npv(cf, años, tasa):
    return sum(
        cf[i] / ((1 + tasa) ** años[i])
        for i in range(len(cf))
    )


def _calcular_irr(cf, años, tol=1e-6, max_iter=200):
    """IRR por bisección entre -0.99 y +10.  Devuelve None
    si no converge o si no hay cambio de signo en el CF."""

    if not cf or all(c >= 0 for c in cf) or all(c <= 0 for c in cf):
        return None

    lo, hi = -0.99, 10.0
    f_lo = _npv(cf, años, lo)
    f_hi = _npv(cf, años, hi)

    if f_lo * f_hi > 0:
        return None

    for _ in range(max_iter):
        mid = (lo + hi) / 2
        f_mid = _npv(cf, años, mid)

        if abs(f_mid) < tol:
            return mid

        if f_lo * f_mid < 0:
            hi, f_hi = mid, f_mid
        else:
            lo, f_lo = mid, f_mid

    return (lo + hi) / 2
