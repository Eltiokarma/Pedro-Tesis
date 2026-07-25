"""datasheet_export.py — la Ficha Técnica sale del programa (tandas 3 y 4).

Dos formatos, un solo contenido:

  · `write_datasheets_xlsx` — un libro con una hoja índice + una hoja por
    equipo.  Qt-free (openpyxl), sirve headless y desde el CLI.
  · `write_datasheets_pdf`  — un legajo: una ficha por página A4, con
    encabezado proyecto/tag/revisión/fecha y una página final de historial
    de revisiones △N (deuda 4d del ciclo 4).  Requiere PySide6.

LA REGLA DE ESTE MÓDULO: qué se imprime y en qué orden se decide UNA sola
vez, en `datasheet_rows()`.  Los dos renderers consumen esas filas y solo
deciden cómo pintarlas.  Si el XLSX y el PDF tuvieran cada uno su recorrido
del spec, divergirían al primer campo nuevo — y la ficha dejaría de ser el
mismo documento en dos formatos.

Este módulo NO calcula nada: todo sale de `datasheet.datasheet_spec`, que a
su vez solo lee resultados ya resueltos.  Exportar no puede mover la física.
"""
from typing import List, Optional, Tuple

import datasheet as ds

# (sección, etiqueta, valor, nota) — la nota es el texto tenue de la
# derecha: procedencia del dato, unidad ya incluida en el valor, o el
# detalle que el veredicto necesita para no afirmar de más.
Fila = Tuple[str, str, str, str]

_ORIGEN_TXT = {
    "calculado": "calculado",
    "declarado": "declarado (lock)",
    "tipico": "valor típico",
    "estimado": "estimado",
    "pendiente": "ingeniería de detalle",
}

_ESTADO_TXT = {
    "no_aplica": "sin catálogo comercial para este tipo",
    "sin_declarar": "hay catálogo; no se declaró modelo",
    "desconocido": "el modelo declarado no está en el catálogo",
    "escalar": "verificación escalar (requerido vs instalado)",
    "envolvente_modelo": "verificación por envolvente del modelo",
    "debil_familia": "envolvente de FAMILIA — no concluyente",
}


def _num(v, nd=3) -> str:
    """Número legible sin ceros de relleno; '—' si no hay dato."""
    if v is None or v == "":
        return "—"
    if isinstance(v, bool):
        return "sí" if v else "no"
    if isinstance(v, (int, float)):
        return f"{float(v):.{nd}f}".rstrip("0").rstrip(".") or "0"
    return str(v)


def _u(value, unit_canon: str) -> str:
    """Valor en la unidad ACTIVA del usuario.

    El spec habla en canónicas (°C/bar/kW/tm por año) porque es un
    contrato de datos; el papel muestra lo que el usuario eligió.  Si el
    sistema de unidades no cubre esa magnitud (m², m³/h, m), sale tal
    cual — ver `flowsheet_units.fmt_canonica`."""
    try:
        import flowsheet_units as funits
        return funits.fmt_canonica(value, unit_canon)
    except Exception:
        # El export tiene que salir aunque el módulo de unidades falle:
        # una ficha en unidades canónicas es peor que una en las del
        # usuario, pero muchísimo mejor que ninguna.
        u = f" {unit_canon}" if unit_canon else ""
        return f"{_num(value)}{u}"


def datasheet_rows(spec: dict) -> List[Fila]:
    """Aplana una ficha (`datasheet.datasheet_spec`) a filas imprimibles.

    Único lugar donde vive el orden y el vocabulario de la ficha: los dos
    exports leen de acá.  Nunca lanza — una sección vacía simplemente no
    aporta filas (el fallback genérico del agregador ya garantiza que
    identidad, condiciones y costos existan para los 60 tipos)."""
    filas: List[Fila] = []
    add = lambda s, e, v, n="": filas.append((s, e, v, n))

    ident = spec.get("identidad", {}) or {}
    add("Identidad", "Tag", ident.get("tag", "—"))
    add("Identidad", "Tipo de equipo", ident.get("eq_type", "—"))
    if ident.get("categoria"):
        add("Identidad", "Categoría", ident["categoria"])
    n_par = int(ident.get("n_paralelo", 1) or 1)
    if n_par > 1:
        add("Identidad", "Unidades en paralelo", str(n_par),
            "el costo del bloque ya multiplica por n")
    if ident.get("marca"):
        add("Identidad", "Equipo comercial",
            f"{ident['marca']} · {ident.get('modelo', '')}".strip(" ·"))

    cond = spec.get("condiciones", {}) or {}
    if cond.get("T_op_C") is not None:
        add("Condiciones de operación", "T operación",
            _u(cond["T_op_C"], "°C"))
    if cond.get("P_op_bar") is not None:
        add("Condiciones de operación", "P operación",
            _u(cond["P_op_bar"], "bar"))
    if cond.get("duty_kW"):
        add("Condiciones de operación", "Duty", _u(cond["duty_kW"], "kW"),
            "positivo aporta calor, negativo lo retira")
    if cond.get("delta_p_bar"):
        add("Condiciones de operación", "ΔP del equipo",
            _u(cond["delta_p_bar"], "bar"))
    if cond.get("fases"):
        add("Condiciones de operación", "Fases", " · ".join(cond["fases"]))

    corr = spec.get("corrientes", {}) or {}
    for sentido, clave in (("Entrada", "in"), ("Salida", "out")):
        for s in corr.get(clave, []) or []:
            etiqueta = f"{sentido} · {s.get('nombre', '')}"
            valor = " · ".join((
                _u(s.get("mass_flow_tm_a"), "tm/año"),
                _u(s.get("T_C"), "°C"),
                _u(s.get("P_bar"), "bar")))
            nota = " · ".join(x for x in (
                s.get("puerto", ""), s.get("fase", ""),
                s.get("comp_principal", ""),
                "auxiliar" if s.get("auto_aux") else "") if x)
            add("Corrientes", etiqueta, valor, nota)

    for c in spec.get("diseno", []) or []:
        add("Diseño", c.get("label", c.get("key", "")),
            _u(c.get("value"), c.get("unit", "")),
            _ORIGEN_TXT.get(c.get("origen", ""), c.get("origen", "")))

    mat = spec.get("materiales", {}) or {}
    if mat.get("material"):
        add("Materiales", "Material", mat["material"],
            f"FM {_num(mat.get('FM', 1.0))} · "
            f"{_ORIGEN_TXT.get(mat.get('origen', ''), '')}".strip(" ·"))

    for a in spec.get("auxiliares", []) or []:
        flecha = "→" if a.get("direccion") == "out" else "←"
        srv = a.get("utility") or a.get("fase") or ""
        lazo = f" · lazo {a['lazo']}" if a.get("lazo") else ""
        add("Servicios auxiliares", a.get("puerto", ""),
            f"{flecha} {srv}".strip(), (a.get("rol") or "") + lazo)

    cos = spec.get("costos", {}) or {}
    if cos:
        add("Costos (bare module)",
            f"Cp {cos.get('year_target', '')}".strip(),
            f"USD {_num(cos.get('Cp_target'), 0)}")
        add("Costos (bare module)", "CBM", f"USD {_num(cos.get('CBM'), 0)}",
            f"FBM {_num(cos.get('FBM'))} · FP {_num(cos.get('FP'))}")
        if cos.get("source"):
            add("Costos (bare module)", "Correlación",
                f"{cos.get('correlacion', '')} · {cos['source']}")
        if cos.get("fuera_rango"):
            add("Costos (bare module)", "Rango", "FUERA DE RANGO",
                "el costo es una extrapolación de la correlación")

    v = spec.get("verificacion", {}) or {}
    estado = v.get("estado", "")
    if estado:
        add("Verificación", "Estado", estado.replace("_", " "),
            _ESTADO_TXT.get(estado, ""))
    if v.get("marca"):
        add("Verificación", "Modelo declarado",
            f"{v['marca']} · {v.get('modelo', '')}".strip(" ·"))
    if v.get("apto") is not None:
        add("Verificación", "Veredicto",
            "APTO" if v["apto"] else "NO ENTRA")
    elif estado in ("envolvente_modelo", "debil_familia"):
        # El estado débil NUNCA afirma: se dice, no se deja en blanco.
        add("Verificación", "Veredicto", "sin afirmación",
            "la envolvente de familia no permite concluir"
            if estado == "debil_familia" else
            "sin ratio de utilización: solo AND de envolvente")
    if v.get("S_requerido") is not None:
        unidad = v.get("S_unit", "")
        add("Verificación", "Requerido",
            f"{_num(v['S_requerido'])} {unidad}".strip())
    if v.get("S_modelo") is not None:
        unidad = v.get("S_unit", "")
        add("Verificación", "Instalado",
            f"{_num(v['S_modelo'])} {unidad}".strip())
    if v.get("utilizacion") is not None:
        add("Verificación", "Utilización",
            f"{v['utilizacion'] * 100:.1f} %")
    for c in v.get("checks", []) or []:
        marca = "✓" if c.get("ok") else "✗"
        add("Verificación", f"{marca} {c.get('param', '')}",
            f"{_num(c.get('requerido'))} vs {_num(c.get('limite'))}",
            "límite de FAMILIA" if c.get("familia") else "límite del modelo")
    if v.get("fuente"):
        add("Verificación", "Fuente del fabricante", v["fuente"],
            f"consultado {v.get('fecha_consulta', '')}".strip())
    if v.get("mensaje"):
        add("Verificación", "Nota", v["mensaje"])

    notas = spec.get("notas", {}) or {}
    for w in notas.get("warnings", []) or []:
        add("Notas", "⚠", w)
    for f in notas.get("fuentes", []) or []:
        add("Notas", "Fuente de la correlación", f)
    if notas.get("pendiente_detalle"):
        add("Notas", "Ingeniería de detalle pendiente",
            ", ".join(notas["pendiente_detalle"]),
            "fuera del alcance de un diseño conceptual")
    return filas


def bloques_de(fs, blocks=None):
    """Los bloques que van a la ficha, en orden estable por tag.

    Los auxiliares (`auto_aux`) quedan FUERA: son el lazo de servicio que
    el simulador arma solo, no equipos que alguien vaya a comprar."""
    if blocks is None:
        blocks = [b for b in fs.blocks.values()
                  if not getattr(b, "auto_aux", False)]
    return sorted(blocks, key=lambda b: (b.name or ""))


# ─────────────────────────── XLSX (tanda 3) ───────────────────────────

def _nombre_hoja(tag: str, usados: set) -> str:
    """Excel no acepta []:*?/\\ ni más de 31 caracteres, y no admite dos
    hojas con el mismo nombre."""
    base = "".join("-" if ch in "[]:*?/\\" else ch
                   for ch in (tag or "equipo")).strip() or "equipo"
    base = base[:31]
    nombre, n = base, 2
    while nombre.lower() in usados:
        sufijo = f"~{n}"
        nombre = base[:31 - len(sufijo)] + sufijo
        n += 1
    usados.add(nombre.lower())
    return nombre


def write_datasheets_xlsx(path, fs, blocks=None, proyecto: str = "") -> int:
    """Escribe el libro de fichas. Devuelve cuántos equipos exportó.

    Hoja «Índice» con una fila por equipo (tag, tipo, marca/modelo, estado
    de verificación, CBM) y una hoja por equipo con la ficha completa."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:                      # pragma: no cover
        raise RuntimeError(
            "El export XLSX de fichas necesita openpyxl "
            "(pip install openpyxl)") from exc

    equipos = bloques_de(fs, blocks)
    wb = Workbook()

    f_titulo = Font(bold=True, size=13)
    f_seccion = Font(bold=True, size=10)
    f_hdr = Font(bold=True, color="FFFFFF")
    fill_seccion = PatternFill("solid", fgColor="EFEBE3")
    fill_hdr = PatternFill("solid", fgColor="4A4A4A")
    wrap = Alignment(vertical="top", wrap_text=True)

    idx = wb.active
    idx.title = "Índice"
    idx["A1"] = f"Fichas técnicas — {proyecto}".strip().rstrip("—").strip()
    idx["A1"].font = f_titulo
    encabezados = ("Tag", "Tipo", "Categoría", "Marca", "Modelo",
                   "Verificación", "CBM [USD]")
    for c, h in enumerate(encabezados, start=1):
        cel = idx.cell(row=3, column=c, value=h)
        cel.font = f_hdr
        cel.fill = fill_hdr
    for r, b in enumerate(equipos, start=4):
        spec = ds.datasheet_spec(b, fs)
        ident = spec.get("identidad", {}) or {}
        ver = spec.get("verificacion", {}) or {}
        cos = spec.get("costos", {}) or {}
        for c, val in enumerate((
                ident.get("tag", ""), ident.get("eq_type", ""),
                ident.get("categoria", ""), ident.get("marca", "") or "—",
                ident.get("modelo", "") or "—",
                (ver.get("estado", "") or "").replace("_", " "),
                cos.get("CBM", "")), start=1):
            idx.cell(row=r, column=c, value=val)
    for c, w in enumerate((14, 30, 18, 18, 30, 26, 14), start=1):
        idx.column_dimensions[get_column_letter(c)].width = w
    idx.freeze_panes = "A4"

    usados = set()
    for b in equipos:
        spec = ds.datasheet_spec(b, fs)
        ws = wb.create_sheet(_nombre_hoja(b.name, usados))
        ws["A1"] = f"Ficha técnica · {b.name}"
        ws["A1"].font = f_titulo
        if proyecto:
            ws["A2"] = proyecto
        fila = 4
        seccion_actual = None
        for seccion, etiqueta, valor, nota in datasheet_rows(spec):
            if seccion != seccion_actual:
                seccion_actual = seccion
                cel = ws.cell(row=fila, column=1, value=seccion.upper())
                cel.font = f_seccion
                cel.fill = fill_seccion
                ws.cell(row=fila, column=2).fill = fill_seccion
                ws.cell(row=fila, column=3).fill = fill_seccion
                fila += 1
            ws.cell(row=fila, column=1, value=etiqueta).alignment = wrap
            ws.cell(row=fila, column=2, value=valor).alignment = wrap
            ws.cell(row=fila, column=3, value=nota).alignment = wrap
            fila += 1
        for c, w in enumerate((34, 30, 46), start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

    wb.save(path)
    return len(equipos)


# ──────────────────────── PDF multipágina (tanda 4) ────────────────────

def _rev_actual(fs) -> str:
    revs = getattr(fs, "revisions", None) or []
    return revs[-1].get("rev", "") if revs else ""


# QPdfWriter pinta con QFont, y QFont necesita una QGuiApplication viva
# (si no: "Must construct a QGuiApplication before accessing
# QFontDatabase").  Desde la UI ya existe; desde el CLI o un test, no —
# así que se crea una y se retiene, porque si el intérprete la recolecta
# el siguiente export vuelve a fallar.
_qapp_propia = None


def _asegurar_qapp():
    from PySide6.QtGui import QGuiApplication
    global _qapp_propia
    app = QGuiApplication.instance()
    if app is None:
        _qapp_propia = app = QGuiApplication([])
    return app


def write_datasheets_pdf(path, fs, blocks=None, proyecto: str = "",
                         fecha: Optional[str] = None) -> int:
    """Escribe el legajo de fichas en PDF. Devuelve cuántas páginas de
    equipo emitió (el historial de revisiones no cuenta como ficha).

    Una ficha por página A4 vertical.  Encabezado con proyecto, tag,
    revisión vigente y fecha; pie con la numeración.  Si el flowsheet
    registró revisiones △N, se agrega una última página con el historial
    completo — el mismo cuadro que el Marco PFD dibuja en el plano, acá en
    su formato de legajo."""
    try:
        from PySide6.QtCore import QRectF, Qt
        from PySide6.QtGui import (QFont, QPageSize, QPainter, QPdfWriter,
                                   QPen, QColor)
    except ImportError as exc:                      # pragma: no cover
        raise RuntimeError(
            "El export PDF de fichas necesita PySide6 "
            "(pip install PySide6)") from exc
    _asegurar_qapp()

    if fecha is None:
        import datetime
        fecha = datetime.date.today().strftime("%d-%m-%Y")

    equipos = bloques_de(fs, blocks)
    if not equipos:
        raise ValueError("No hay equipos que exportar.")

    writer = QPdfWriter(str(path))
    writer.setPageSize(QPageSize(QPageSize.A4))
    writer.setResolution(300)
    writer.setTitle(f"Fichas técnicas — {proyecto}".strip().rstrip("—").strip()
                    or "Fichas técnicas")

    W = writer.width()
    H = writer.height()
    M = int(writer.resolution() * 0.6)          # margen ~15 mm

    # PAPEL SIEMPRE CLARO — decisión declarada, no accidente (auditoría
    # UI 2 §A.4.4).  La tinta sale de THEME_LIGHT y NO de TOK: TOK sigue
    # el tema activo de la app, y con tema oscuro este legajo se
    # imprimiría en gris sobre negro.  Un PFD y su ficha son documentos
    # imprimibles; el tema es de la pantalla, no del papel.
    import tokens as _tok
    _L = _tok.THEME_LIGHT
    TINTA = QColor(_L["label_ink"])
    SUAVE = QColor(_L["ink_mute"])
    REGLA = QColor(_L["ink_ghost"])

    # Tipografía del PROYECTO, no del PDF: el Marco PFD ya sale en
    # pfd_fonts (mismo fallback), y el plano y el legajo de fichas son dos
    # documentos del mismo trabajo — no pueden hablar dos tipografías.
    import pfd_fonts as _pf
    _SANS = _pf.SANS if _pf.available() else "Helvetica"
    _MONO = _pf.MONO if _pf.available() else "Courier"
    # Los TAMAÑOS son escala de documento impreso, no de UI (la misma
    # «excepción 2g» del Marco PFD): se miden en la hoja A4 a 300 dpi, no
    # en la pantalla, así que no salen de los tokens FONT_* — esos rigen
    # la interfaz.
    f_tag = QFont(_SANS, 15, QFont.Bold)
    f_proy = QFont(_SANS, 8)
    f_sec = QFont(_SANS, 8, QFont.Bold)
    f_lbl = QFont(_SANS, 8)
    f_val = QFont(_MONO, 8)
    f_pie = QFont(_SANS, 7)

    rev = _rev_actual(fs)
    total_paginas = len(equipos) + (1 if getattr(fs, "revisions", None) else 0)

    painter = QPainter(writer)
    painter.setRenderHint(QPainter.Antialiasing, True)
    painter.setRenderHint(QPainter.TextAntialiasing, True)
    try:
        col_lbl = int((W - 2 * M) * 0.30)
        col_val = int((W - 2 * M) * 0.28)
        alto_fila = int(writer.resolution() * 0.155)

        def encabezado(titulo: str, subtitulo: str) -> int:
            painter.setPen(QPen(TINTA))
            painter.setFont(f_tag)
            painter.drawText(M, M + alto_fila, titulo)
            painter.setFont(f_proy)
            painter.setPen(QPen(SUAVE))
            der = " · ".join(x for x in (
                proyecto, f"rev. △{rev}" if rev else "", fecha) if x)
            painter.drawText(QRectF(M, M + alto_fila * 1.2,
                                    W - 2 * M, alto_fila),
                             int(Qt.AlignRight | Qt.AlignVCenter), der)
            painter.drawText(M, int(M + alto_fila * 2.1), subtitulo)
            y = int(M + alto_fila * 2.6)
            painter.setPen(QPen(REGLA, 4))
            painter.drawLine(M, y, W - M, y)
            return y + alto_fila

        def pie(n: int):
            painter.setFont(f_pie)
            painter.setPen(QPen(SUAVE))
            painter.drawText(
                QRectF(M, H - M - alto_fila, W - 2 * M, alto_fila),
                int(Qt.AlignRight | Qt.AlignVCenter),
                f"{n} / {total_paginas}")
            painter.drawText(
                QRectF(M, H - M - alto_fila, W - 2 * M, alto_fila),
                int(Qt.AlignLeft | Qt.AlignVCenter),
                "Diseño conceptual — no apto para construcción")

        # Las tres celdas van con WORD WRAP y la fila crece hasta la más
        # alta.  Sin esto, un valor largo (la lista de ingeniería de
        # detalle, la URL de una fuente) se sale de su columna y se pinta
        # ENCIMA de la nota vecina — no lo recorta nadie, porque drawText
        # sin wrap ignora el ancho del rect.
        FLAGS = int(Qt.AlignLeft | Qt.AlignTop | Qt.TextWordWrap)

        def alto_celda(texto, ancho, fuente) -> float:
            if not texto:
                return 0.0
            painter.setFont(fuente)
            return painter.boundingRect(
                QRectF(0, 0, ancho, 1e6), FLAGS, str(texto)).height()

        col_nota = W - 2 * M - col_lbl - col_val
        for n, b in enumerate(equipos, start=1):
            if n > 1:
                writer.newPage()
            spec = ds.datasheet_spec(b, fs)
            ident = spec.get("identidad", {}) or {}
            y = encabezado(b.name or "—", ident.get("eq_type", ""))
            seccion_actual = None
            for seccion, etiqueta, valor, nota in datasheet_rows(spec):
                alto = max(alto_celda(etiqueta, col_lbl, f_lbl),
                           alto_celda(valor, col_val, f_val),
                           alto_celda(nota, col_nota, f_pie),
                           float(alto_fila))
                salto_seccion = (seccion != seccion_actual)
                extra = alto_fila * 1.4 if salto_seccion else 0
                if y + alto + extra > H - M - alto_fila * 2:
                    # La ficha no entra en una página: sigue en la
                    # siguiente, rotulada, en vez de recortarse en silencio.
                    pie(n)
                    writer.newPage()
                    y = encabezado(b.name or "—",
                                   f"{ident.get('eq_type', '')} (cont.)")
                    seccion_actual = None
                    salto_seccion = True
                if salto_seccion:
                    seccion_actual = seccion
                    y += int(alto_fila * 0.5)
                    painter.setFont(f_sec)
                    painter.setPen(QPen(TINTA))
                    painter.drawText(M, int(y + alto_fila * 0.8),
                                     seccion.upper())
                    y += int(alto_fila * 1.4)
                painter.setFont(f_lbl)
                painter.setPen(QPen(SUAVE))
                painter.drawText(QRectF(M, y, col_lbl, alto), FLAGS, etiqueta)
                painter.setFont(f_val)
                painter.setPen(QPen(TINTA))
                painter.drawText(QRectF(M + col_lbl, y, col_val, alto),
                                 FLAGS, valor)
                if nota:
                    painter.setFont(f_pie)
                    painter.setPen(QPen(SUAVE))
                    painter.drawText(
                        QRectF(M + col_lbl + col_val, y, col_nota, alto),
                        FLAGS, nota)
                y += alto
            pie(n)

        revs = getattr(fs, "revisions", None) or []
        if revs:
            writer.newPage()
            y = encabezado("Historial de revisiones", "cuadro △N del plano")
            cols = (int((W - 2 * M) * 0.08), int((W - 2 * M) * 0.62),
                    int((W - 2 * M) * 0.16))
            painter.setFont(f_sec)
            painter.setPen(QPen(TINTA))
            x = M
            for texto, ancho in zip(("REV", "DESCRIPCIÓN", "FECHA"), cols):
                painter.drawText(QRectF(x, y - alto_fila, ancho, alto_fila),
                                 int(Qt.AlignLeft | Qt.AlignVCenter), texto)
                x += ancho
            painter.drawText(QRectF(x, y - alto_fila, W - M - x, alto_fila),
                             int(Qt.AlignLeft | Qt.AlignVCenter), "POR")
            y += int(alto_fila * 0.4)
            painter.setPen(QPen(REGLA, 3))
            painter.drawLine(M, y, W - M, y)
            y += alto_fila
            for i, r in enumerate(revs):
                ultima = (i == len(revs) - 1)
                painter.setFont(f_val)
                painter.setPen(QPen(TINTA))
                x = M
                celdas = (f"△{r.get('rev', '')}" if ultima
                          else str(r.get("rev", "")),
                          str(r.get("desc", "")), str(r.get("date", "")))
                for texto, ancho in zip(celdas, cols):
                    painter.drawText(
                        QRectF(x, y - alto_fila, ancho, alto_fila),
                        int(Qt.AlignLeft | Qt.AlignVCenter), texto)
                    x += ancho
                painter.drawText(
                    QRectF(x, y - alto_fila, W - M - x, alto_fila),
                    int(Qt.AlignLeft | Qt.AlignVCenter), str(r.get("by", "")))
                y += alto_fila
            pie(total_paginas)
    finally:
        painter.end()
    return len(equipos)
