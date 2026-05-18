from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font


# =========================
# LECTOR DE INPUTS
# =========================
class InputReader:
    """Lee inputs del Excel legacy (data_entrada.xlsx).

    Asume layout fijo con celdas hardcodeadas:
        C3..C7  capital (ISBL, OSBL%, ENG%, CONT%, WC%)
        G3..G11 inputs de FCOP (labor + 8 porcentajes)
        I..N    streams desde fila 3
                (concept, unit, coef, flow, price, tipo)
                tipo: A=key product, B/C=byproduct/waste,
                      D=raw material, E=consumable, F=utility

    Usado solo por flujoflujo.py (script standalone).
    La GUI ANA.py no usa esta clase: arma el dict de
    inputs vía pipeline._construir_data, partiendo de los
    DataFrames editables del Treeview.
    """

    def __init__(self, ruta):
        self.ruta = ruta

    def leer(self):
        """Devuelve un dict consumible por CostModel."""
        wb = load_workbook(self.ruta, data_only=True)
        ws = wb.active

        data = {}  # ✅ FIX CLAVE

        # -------- FCOP DETALLADO ----------
        data["FCOP_inputs"] = {
            "labor": ws["G3"].value,
            "supervision_pct": ws["G4"].value / 100,
            "salary_overhead_pct": ws["G5"].value / 100,
            "maintenance_pct": ws["G6"].value / 100,
            "plant_overhead_pct": ws["G7"].value / 100,
            "tax_insurance_pct": ws["G8"].value / 100,
            "interest_pct": ws["G9"].value / 100,
            "general_expenses_pct": ws["G10"].value / 100,
            "royalties_pct": ws["G11"].value / 100
        }

        # CAPITAL
        data["ISBL"] = ws["C3"].value
        data["OSBL_pct"] = ws["C4"].value / 100
        data["ENG_pct"]  = ws["C5"].value / 100
        data["CONT_pct"] = ws["C6"].value / 100
        data["WC_pct"]   = ws["C7"].value / 100

        # LISTAS
        key, bp, rm, cons, uts = [], [], [], [], []

        r = 3
        while True:
            concept = ws[f"I{r}"].value
            if concept is None:
                break

            fila = {
                "concept": ws[f"I{r}"].value,
                "unit": ws[f"J{r}"].value,
                "coef": ws[f"K{r}"].value,
                "flow": ws[f"L{r}"].value,
                "price": ws[f"M{r}"].value
            }

            tipo = ws[f"N{r}"].value

            if tipo:
                tipo = str(tipo).upper()

                if tipo == "A": key.append(fila)
                elif tipo in ["B", "C"]: bp.append(fila)
                elif tipo == "D": rm.append(fila)
                elif tipo == "E": cons.append(fila)
                elif tipo == "F": uts.append(fila)

            r += 1

        data["key_products"] = key
        data["byproducts"] = bp
        data["raw_materials"] = rm
        data["consumables"] = cons
        data["utilities"] = uts


        return data

# =========================
# CASH FLOW
# =========================
class CostModel:
    """Calcula costos de capital y de producción a partir
    del dict data (formato InputReader o pipeline).

    Salidas en MM USD (asume que los inputs ya están en
    esa escala, salvo labor que pipeline.py convierte
    desde USD/yr).

    Fórmulas:
        ISBL = data["ISBL"]
        OSBL = OSBL% × ISBL
        ENG  = ENG% × (ISBL + OSBL)
        CONT = CONT% × (ISBL + OSBL)
        FCI  = ISBL + OSBL + ENG + CONT
        WC   = WC% × FCI
        REV  = Σ(flow × price) sobre key_products  /1e6
        BP   = Σ(flow × price) sobre byproducts    /1e6
        RM   = Σ(flow × price) sobre raw_materials /1e6
        CONS = Σ(coef × price × production)        /1e6
        UTS  = Σ(coef × price × production)        /1e6
        VCOP = RM − BP + CONS + UTS
        FCOP = (ver calcular_fcop_detallado)

    production = key_products[0]["flow"] o 1 si vacío.
    Limitación conocida: para plantas multi-producto solo
    se usa el primer key product como base de producción.
    """

    def __init__(self, data):
        self.data = data

    def calcular(self):
        """Ejecuta todos los cálculos.  Devuelve dict con
        ISBL, OSBL, ENG, CONT, FCI, WC, Revenue, Byproducts,
        RawMaterials, Consumables, Utilities, FCOP,
        FCOP_detalle, VCOP."""

        ISBL = self.data["ISBL"]
        OSBL = self.data["OSBL_pct"] * ISBL
        ENG  = self.data["ENG_pct"] * (ISBL + OSBL)
        CONT = self.data["CONT_pct"] * (ISBL + OSBL)

        FCI = ISBL + OSBL + ENG + CONT
        WC = self.data["WC_pct"] * FCI

        REV = sum(f["flow"] * f["price"] for f in self.data["key_products"]) / 1e6
        BP  = sum(f["flow"] * f["price"] for f in self.data["byproducts"]) / 1e6
        RM  = sum(f["flow"] * f["price"] for f in self.data["raw_materials"]) / 1e6

        production = self.data["key_products"][0]["flow"] if self.data["key_products"] else 1

        CONS = sum(f["coef"] * f["price"] * production for f in self.data["consumables"]) / 1e6
        UTS  = sum(f["coef"] * f["price"] * production for f in self.data["utilities"]) / 1e6

        VCOP = RM - BP + CONS + UTS

        # ✅ FCOP SIN RECURSIÓN
        fcop_det = self.calcular_fcop_detallado(ISBL, OSBL, FCI, WC)

        return {
            "ISBL": ISBL,
            "OSBL": OSBL,
            "ENG": ENG,
            "CONT": CONT,
            "FCI": FCI,
            "WC": WC,
            "Revenue": REV,
            "Byproducts": BP,
            "RawMaterials": RM,
            "Consumables": CONS,
            "Utilities": UTS,
            "FCOP": fcop_det["FCOP_total"],
            "FCOP_detalle": fcop_det,
            "VCOP": VCOP
        }
    def calcular_fcop_detallado(self, ISBL, OSBL, FCI, WC):
        """Desglose del Fixed Cost of Production.

        BUG conocido: este método mezcla escalas si labor
        viene en USD/yr crudo (lo demás está en MMUSD).
        El pipeline GUI corrige esto dividiendo labor por
        1e6 antes de llegar acá; el script flujoflujo.py
        original NO lo corrige.
        """

        f = self.data["FCOP_inputs"]

        labor = f["labor"]

        supervision = f["supervision_pct"] * labor

        salary_overhead = f["salary_overhead_pct"] * (labor + supervision)

        maintenance = f["maintenance_pct"] * FCI

        plant_overhead = f["plant_overhead_pct"] * (labor + maintenance)

        tax_insurance = f["tax_insurance_pct"] * (ISBL + OSBL)

        interest = f["interest_pct"] * FCI

        general = f["general_expenses_pct"] * WC

        total_fcop = (
            labor + supervision + salary_overhead + maintenance +
            plant_overhead + tax_insurance + interest + general
        )

        return {
            "Labor": labor,
            "Supervision": supervision,
            "Salary Overhead": salary_overhead,
            "Maintenance": maintenance,
            "Plant Overhead": plant_overhead,
            "Tax & Insurance": tax_insurance,
            "Interest": interest,
            "General Expenses": general,
            "FCOP_total": total_fcop,
            "royalties_pct": f["royalties_pct"]
        }


class CashFlowModel:
    """Construye el cash flow anual del proyecto.

    Parámetros esperados en params:
        vida           — años totales del análisis
        tasa_impuesto  — fracción (ej 0.35)
        metodo_dep     — 0=lineal, 1=MACRS
        periodo_dep    — años (solo si metodo_dep=0)
        tipo_macrs     — 0=MACRS5, 1=MACRS7, 2=MACRS15
        tasa_interes   — descuento para NPV
        schedule       — dict con FC, VCOP, FCOP, WL,
                         t_start, steady_start, cutoff,
                         años_display (ver
                         ReportGenerator.construir_schedule)

    Convenciones del modelo:
      - Royalties (versión activa): Revenue_base × pct
        cuando opera (NO escala con ramp-up VCOP).
      - Impuestos: se acumulan al cierre de cada año y se
        pagan con desfase de 1 año.
      - FCOP solo se carga cuando arranca operación
        (i ≥ t_start).
      - WC se invierte en t_start y se recupera en el
        último año.
    """

    def __init__(self, costos, params):
        self.costos = costos
        self.params = params

    # =========================
    # DEPRECIACIÓN
    # =========================
    def calcular_depreciacion(self):
        """Devuelve lista de depreciación anual (MM USD).

        Si metodo_dep=0: D constante = FCI / periodo_dep
        repetida periodo_dep veces.

        Si metodo_dep=1: tabla MACRS según tipo_macrs
        (MACRS 5/7/15).  Devuelve una lista del largo de
        la tabla; calcular() la alinea contra t_start.
        """

        FCI = self.costos["FCI"]
        metodo = self.params["metodo_dep"]

        depreciacion = []  # sin dummy

        # =========================
        # 🔵 LINEAL
        # =========================
        if metodo == 0:

            vida_dep = self.params["periodo_dep"]
            D = FCI / vida_dep

            depreciacion += [D] * vida_dep

        # =========================
        # 🔵 MACRS
        # =========================
        else:
            macrs_tablas = {
                0: [0.20, 0.32, 0.192, 0.1152, 0.1152, 0.0576],
                1: [0.1429, 0.2449, 0.1749, 0.1249, 0.0893, 0.0892, 0.0893, 0.0446],
                2: [0.05, 0.095, 0.0855, 0.077, 0.0693, 0.0623, 0.059, 0.059,
                    0.0591, 0.059, 0.0591, 0.059, 0.0591, 0.059, 0.0591, 0.0295]
            } 

            tipo_macrs = self.params["tipo_macrs"]
            tabla = macrs_tablas[tipo_macrs]

            for factor in tabla:
                depreciacion.append(FCI * factor)

        return depreciacion

    # =========================
    # CASH FLOW
    # =========================
    def calcular(self):
        """Loop principal de cash flow.

        Devuelve dict con listas paralelas (largo = vida):
            años, CapEx, Revenue, CCOP, GP, Dep, TI,
            Taxes, CF.
        """

        schedule = self.params["schedule"]

        # ✔ VIDA TOTAL DEL PROYECTO
        vida = len(schedule["FC"])

        tasa_impuesto = self.params["tasa_impuesto"]

        Revenue_base = self.costos["Revenue"]
        FCOP_base = self.costos["FCOP"]
        VCOP_base = self.costos["VCOP"]
        FCI = self.costos["FCI"]
        WC = self.costos["WC"]

        royalties_pct = self.costos["FCOP_detalle"]["royalties_pct"]

        #D = self.calcular_depreciacion()

        años = schedule["años_display"]

        Revenue, CCOP, GP, TI, taxes, CF = [], [], [], [], [], []
        CapEx = []

        # =========================
        # START OPERACIÓN
        # =========================
        t_start = schedule["WL"].index(1)

        # =========================
        # 🔵 DEPRECIACIÓN AJUSTADA
        # =========================
        D_base = self.calcular_depreciacion()

        D = [0] * vida

        for i in range(t_start, vida):

            idx_dep = i - t_start   #

            if idx_dep < len(D_base):
                D[i] = D_base[idx_dep]
        # =========================
        # LOOP PRINCIPAL
        # =========================
        taxes_accrued = [0] * vida
        taxes_paid = [0] * vida

        for i, t in enumerate(años):

            # =========================
            # CAPEX
            # =========================
            capex = FCI * (schedule["FC"][i])

            # ✔ WC cuando inicia operación
            if i == t_start:
                capex += WC

            # ✔ recuperación WC al final
            if i == vida - 1:
                capex -= WC

            CapEx.append(capex)

            # =========================
            # 🔵 OPERACIÓN
            # =========================

            factor = schedule["VCOP"][i] 

            # Revenue y VCOP dinámicos
            rev = Revenue_base * factor
            vcop = VCOP_base * factor

            # ✔ FCOP solo cuando opera
            if i < t_start:
                fcop = 0
            else:
                fcop = FCOP_base

            # =========================
            # 🔵 ROYALTIES
            # =========================

            # ✔ versión simplificada (ACTIVA)
            royalties = Revenue_base * royalties_pct if i >= t_start else 0

            # 🔵 versión real (déjala para luego)
            # royalties = rev * royalties_pct

            # =========================
            # CCOP TOTAL
            # =========================
            ccop = fcop + vcop + royalties

            # =========================
            # GROSS PROFIT
            # =========================
            if i < t_start:
                gp = 0
            else:
                gp = rev - ccop

            # =========================
            # TAXABLE INCOME
            # =========================
            ti = gp - D[i]

            # IMPUESTOS
            # =========================
            tax_accrued = max(ti * tasa_impuesto, 0)

            # ➕ AGREGAR LISTAS (si no las tienes arriba)
            taxes_accrued[i] = tax_accrued


            Revenue.append(rev)
            CCOP.append(ccop)
            GP.append(gp)
            TI.append(ti)

        # =========================
        # TAXES PAGADOS
        # =========================
        for i in range(1, vida):
            taxes_paid[i] = taxes_accrued[i - 1]

        Taxes = taxes_paid

        for i in range(vida):

            cf = (
                GP[i]
                - Taxes[i]
                - CapEx[i]
            )

            CF.append(cf)

        return {
            "años": años,
            "CapEx": CapEx,
            "Revenue": Revenue,
            "CCOP": CCOP,
            "GP": GP,
            "Dep": D,
            "TI": TI,
            "Taxes": Taxes,
            "CF": CF
        }


# =========================
# REPORTE EXCEL
# =========================
class ReportGenerator:
    """Genera el reporte Excel (2 hojas: Detailed Costs +
    Cash Flow) y, en su método construir_schedule, también
    arma el dict schedule completo a partir del input
    crudo {FC: [...], VCOP: [...]}.

    Limitación: exportar_cashflow contiene cálculos de
    presentación (royalties, salary_overheads agrupados,
    fcop_total recalculado) que mezclan presentación con
    lógica — un refactor futuro debería moverlos a
    CostModel.
    """

    def exportar_costos_detallado(self, wb, data, costos):
        """Escribe la hoja 'Detailed Costs' en wb."""

        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        ws = wb.create_sheet("Detailed Costs")

        # --------- Estilos ----------
        bold = Font(bold=True)
        titulo = Font(bold=True, size=14)
        right = Alignment(horizontal="right")

        fill = PatternFill(start_color="D9D9D9", fill_type="solid")

        thin = Side(style="thin")
        medium = Side(style="medium")

        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        num = "#,##0.00"

        row = 1

        # =========================================================
        # 🔵 CAPITAL COSTS
        # =========================================================
        ws.merge_cells("A1:D1")
        ws["A1"] = "CAPITAL COSTS"
        ws["A1"].font = titulo

        ws.append(["", "", "$MM", ""])

        ISBL = data["ISBL"]
        OSBL = data["OSBL_pct"] * ISBL
        ENG  = data["ENG_pct"] * (ISBL + OSBL)
        CONT = data["CONT_pct"] * (ISBL + OSBL)

        FCI = costos["FCI"]
        WC = costos["WC"]

        capital = [
            ["ISBL Capital Cost", ISBL],
            ["OSBL Capital Cost", OSBL],
            ["Engineering Costs", ENG],
            ["Contingency", CONT],
            ["Total Fixed Capital Cost", FCI],
            ["Working Capital", WC],
        ]

        for i, row_data in enumerate(capital, start=3):
            ws.cell(i,1,row_data[0])
            c = ws.cell(i,2,row_data[1])
            c.number_format = num
            c.alignment = right

        # =========================================================
        # 🔵 FIXED OPERATING COSTS (IDÉNTICO A TU EXCEL)
        # =========================================================
        offset = 10

        ws.merge_cells(f"A{offset}:D{offset}")
        ws[f"A{offset}"] = "FIXED OPERATING COSTS"
        ws[f"A{offset}"].font = titulo

        headers = ["Concept", "", "Basis", "$MM/yr"]

        for col in range(1,5):
            c = ws.cell(offset+1, col, headers[col-1])
            c.font = bold
            c.fill = fill

        start = offset + 2

        fc = costos["FCOP_detalle"]

        filas = [
            ("Labor", "$MM/yr", fc["Labor"]),
            ("Supervision", "% of operating Labor", fc["Supervision"]),
            ("Direct Salary Overhead", "% of Labor & Supervision", fc["Salary Overhead"]),
            ("Maintenance", "% of Total Fixed Capital Cost", fc["Maintenance"]),
            ("Plant Overhead", "% of Labor & Maintenance", fc["Plant Overhead"]),
            ("Tax & Insurance", "% of ISBL & OSBL", fc["Tax & Insurance"]),
            ("Interest on Debit Financing", "% of Fixed Capital", fc["Interest"]),
            ("", "% of Working Capital", fc["General Expenses"]),
        ]

        for i, (name, basis, value) in enumerate(filas):
            ws.cell(start+i, 1, name)
            ws.cell(start+i, 3, basis)

            c = ws.cell(start+i, 4, value)
            c.number_format = num
            c.alignment = right

        # TOTAL FCOP
        total_row = start + len(filas)

        ws.cell(total_row, 3, "Fixed Cost of Production (FCOP)").alignment = right
        ws.cell(total_row, 4, fc["FCOP_total"]).number_format = num
        ws.cell(total_row, 4).border = Border(top=medium)

        # =========================================================
        # 🔵 FUNCIONES AUXILIARES
        # =========================================================
        row = total_row + 4

        def titulo_tabla(texto):
            nonlocal row
            ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6)
            ws.cell(row,1,texto).font = titulo
            row += 1

        def encabezado():
            headers = ["Concept","Units","Units/Unit product","Units/yr","Price","$MM/yr"]
            for c in range(1,7):
                cell = ws.cell(row,c,headers[c-1])
                cell.font = bold
                cell.fill = fill

        # =========================================================
        # 🔵 REVENUES
        # =========================================================
        titulo_tabla("REVENUES AND RAW MATERIAL COSTS")
        encabezado()
        row += 1

        # KEY PRODUCTS
        ws.cell(row,1,"Key Products").font = bold
        row += 1

        total_key = 0

        for f in data["key_products"]:
            val = f["flow"] * f["price"] / 1e6
            total_key += val

            ws.cell(row,1,f["concept"])
            ws.cell(row,2,f["unit"])
            ws.cell(row,3,f["coef"])
            ws.cell(row,4,f["flow"])
            ws.cell(row,5,f["price"])
            ws.cell(row,6,val)

            row += 1

        ws.cell(row,1,"Total Key Product Revenues (REV)").alignment = right
        ws.cell(row,6,total_key)

        row += 2

        # BYPRODUCTS
        ws.cell(row,1,"By-products & Waste Streams").font = bold
        row += 1

        total_bp = 0

        for f in data["byproducts"]:
            val = f["flow"] * f["price"] / 1e6
            total_bp += val

            ws.cell(row,1,f["concept"])
            ws.cell(row,2,f["unit"])
            ws.cell(row,3,f["coef"])
            ws.cell(row,4,f["flow"])
            ws.cell(row,5,f["price"])
            ws.cell(row,6,val)

            row += 1

        ws.cell(row,1,"Total Byproducts and Wastes (BP)").alignment = right
        ws.cell(row,6,total_bp)

        row += 2

        # RAW MATERIALS
        ws.cell(row,1,"Raw Materials").font = bold
        row += 1

        total_rm = 0

        for f in data["raw_materials"]:
            val = f["flow"] * f["price"] / 1e6
            total_rm += val

            ws.cell(row,1,f["concept"])
            ws.cell(row,2,f["unit"])
            ws.cell(row,3,f["coef"])
            ws.cell(row,4,f["flow"])
            ws.cell(row,5,f["price"])
            ws.cell(row,6,val)

            row += 1

        ws.cell(row,1,"Total Raw Materials (RM)").alignment = right
        ws.cell(row,6,total_rm)

        row += 2

        production = data["key_products"][0]["flow"] if data["key_products"] else 1

        # =========================================================
        # 🔵 CONSUMABLES
        # =========================================================
        titulo_tabla("CONSUMABLES")
        encabezado()
        row += 1

        total_cons = 0

        for f in data["consumables"]:
            val = f["coef"] * f["price"] * production / 1e6
            total_cons += val

            ws.cell(row,1,f["concept"])
            ws.cell(row,2,f["unit"])
            ws.cell(row,3,f["coef"])
            ws.cell(row,4,"")
            ws.cell(row,5,f["price"])
            ws.cell(row,6,val)

            row += 1

        ws.cell(row,1,"Total Consumables (CONS)").alignment = right
        ws.cell(row,6,total_cons)

        row += 2

        # =========================================================
        # 🔵 UTILITIES
        # =========================================================
        titulo_tabla("UTILITIES")
        encabezado()
        row += 1

        total_uts = 0

        for f in data["utilities"]:
            val = f["coef"] * f["price"] * production / 1e6
            total_uts += val

            ws.cell(row,1,f["concept"])
            ws.cell(row,2,f["unit"])
            ws.cell(row,3,f["coef"])
            ws.cell(row,4,"")
            ws.cell(row,5,f["price"])
            ws.cell(row,6,val)

            row += 1

        ws.cell(row,1,"Total Utilities (UTS)").alignment = right
        ws.cell(row,6,total_uts)

        row += 2

        # =========================================================
        # 🔵 VCOP
        # =========================================================
        vcop = total_rm - total_bp + total_cons + total_uts

        ws.merge_cells(start_row=row,start_column=3,end_row=row,end_column=5)
        ws.cell(row,3,"Variable Cost of Production (VCOP = RM - BP + CONS + UTS)")
        ws.cell(row,6,vcop)

        row += 2

        # =========================================================
        # 🔵 SUMMARY
        # =========================================================
        titulo_tabla("SUMMARY")

        ws.cell(row,3,"Variable Cost of Production")
        ws.cell(row,6,vcop)
        row += 1

        ws.cell(row,3,"Fixed Cost of Production")
        ws.cell(row,6,costos["FCOP"])
        row += 1

        ws.cell(row,3,"Cash Cost of Production")
        ws.cell(row,6,vcop + costos["FCOP"])
        ws.cell(row,6).border = Border(top=medium)

        # =========================================================
        # 🔵 BORDES
        # =========================================================
        for r in ws.iter_rows(min_row=1,max_row=ws.max_row,min_col=1,max_col=6):
            for c in r:
                c.border = border

    def exportar_cashflow(self, nombre_archivo, cf, params, costos, data):
        """Genera el .xlsx completo con dos hojas (Detailed
        Costs + Cash Flow), NPV acumulado e IRR (fórmula
        Excel)."""

        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        wb.remove(wb.active)
        #ws.title = "Economic Analysis"
        # =========================
        # HOJA 1 → COSTOS
        # =========================
        self.exportar_costos_detallado(
            wb,
            data,
            costos
        )

        # =========================
        # HOJA 2 → CASH FLOW
        # =========================
        ws = wb.create_sheet("Cash Flow")

        bold = Font(bold=True)

        fila = 1

        # =========================
        # TITULOS SUPERIORES
        # =========================
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
        ws.cell(row=fila, column=1, value="REVENUES AND PRODUCTION COSTS").font = bold

        ws.merge_cells(start_row=fila, start_column=4, end_row=fila, end_column=6)
        ws.cell(row=fila, column=4, value="CAPITAL COSTS").font = bold

        ws.merge_cells(start_row=fila, start_column=8, end_row=fila, end_column=12)
        ws.cell(row=fila, column=8, value="CONSTRUCTION SCHEDULE").font = bold
        fila += 1

        # =========================
        # TABLAS SUPERIORES
        # =========================
        labels1 = [
            "Main product revenue",
            "Byproduct revenue",
            "Raw materials cost",
            "Utilities cost",
            "Consumables cost",
            "VCOP",
            "Salary and overheads",
            "Maintenance",
            "Interest",
            "Royalties",
            "FCOP"
        ]

        labels2 = [
            "ISBL Capital Cost",
            "OSBL Capital Cost",
            "Engineering Costs",
            "Contingency",
            "Total Fixed Capital Cost",
            "Working Capital"
        ]

        # =========================
        # SCHEDULE (YA VIENE CON STEADY_START)
        # =========================
        schedule = params["schedule"]

        steady_start = schedule["steady_start"]
        t_start = schedule["t_start"]

        headers_sched = ["Year", "% FC", "% WC", "% FCOP", "% VCOP"]

        # Encabezados schedule
        for j, h in enumerate(headers_sched):
            ws.cell(row=fila, column=8+j, value=h).font = bold

        # =========================
        # VALORES MODELO
        # =========================
        fc = costos["FCOP_detalle"]
        f_inputs = data["FCOP_inputs"]

        salary_overheads = (
            fc["Labor"] +
            fc["Supervision"] +
            fc["Salary Overhead"] +
            fc["Plant Overhead"] +
            fc["Tax & Insurance"]
        )

        maintenance = fc["Maintenance"]

        interest_total = (
            fc["Interest"] +
            fc["General Expenses"]
        )

        royalties = costos["Revenue"] * f_inputs["royalties_pct"]

        fcop_total = (
            salary_overheads +
            maintenance +
            interest_total +
            royalties
        )

        valores = [
            costos["Revenue"],
            costos["Byproducts"],
            costos["RawMaterials"],
            costos["Utilities"],
            costos["Consumables"],
            costos["VCOP"],
            salary_overheads,
            maintenance,
            interest_total,
            royalties,
            fcop_total
        ]

        valores_capital = [
            costos["ISBL"],
            costos["OSBL"],
            costos["ENG"],
            costos["CONT"],
            costos["FCI"],
            costos["WC"]
        ]

        # =========================
        # LLENADO TABLAS
        # =========================
        max_rows = max(len(labels1), len(labels2), schedule["cutoff"])
        instantaneo = (schedule["t_start"] == 1 and schedule["FC"][0] == 1)

        for i in range(max_rows):

            row = fila + i + 1

            # -------------------------
            # REVENUES & COSTS
            # -------------------------
            if i < len(labels1):
                ws.cell(row=row, column=1, value=labels1[i])
                ws.cell(row=row, column=2, value=valores[i])

            # -------------------------
            # CAPITAL COSTS
            # -------------------------
            if i < len(labels2):
                ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
                ws.cell(row=row, column=4, value=labels2[i])
                ws.cell(row=row, column=6, value=valores_capital[i])

            # -------------------------
            # SCHEDULE
            # -------------------------
            cutoff = schedule["cutoff"]

            if i < cutoff:

                if instantaneo:
                    year_label = i
                else:
                    year_label = i + 1

                if i == cutoff - 1:
                    if instantaneo:
                        year_label = f"{i}+"
                    else:
                        year_label = f"{i+1}+"

                ws.cell(row=row, column=8, value=year_label)
                c = ws.cell(
                    row=row,
                    column=9,
                    value=schedule["FC"][i]
                )

                c.number_format = "0%"
                c = ws.cell(
                    row=row,
                    column=10,
                    value=schedule["WL"][i]
                )

                c.number_format = "0%"
                c = ws.cell(
                    row=row,
                    column=11,
                    value=schedule["FCOP"][i]
                )

                c.number_format = "0%"
                c = ws.cell(
                    row=row,
                    column=12,
                    value=schedule["VCOP"][i]
                )

                c.number_format = "0%"

        fila = fila + max_rows + 3

        # =========================
        # CASH FLOW
        # =========================
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=12)
        ws.cell(row=fila, column=1, value="CASH FLOW ANALYSIS").font = bold

        fila += 1

        headers = [
            "Project Year", "Cap Ex", "Revenue", "CCOP", "Gr. Profit",
            "Deprcn", "Taxbl Inc", "Tax Paid",
            "Cash Flow", "PV of CF", "NPV"
        ]

        for col, h in enumerate(headers, start=1):
            ws.cell(row=fila, column=col, value=h).font = bold

        inicio_cf = fila
        fila += 1

        años = cf["años"]
        CapEx = cf["CapEx"]
        Revenue = cf["Revenue"]
        CCOP = cf["CCOP"]
        GP = cf["GP"]
        Dep = cf["Dep"]
        TI = cf["TI"]
        Taxes = cf["Taxes"]
        CF = cf["CF"]

        n = len(años)
        npv_acumulado = 0

        for i in range(n):

            row = fila

            # -------------------------
            # PROJECT YEAR
            # -------------------------
            ws.cell(row=row, column=1, value=años[i])

            # -------------------------
            # CAPEX
            # -------------------------
            # detectar t_start igual que en el modelo
            schedule = params["schedule"]

            # -------------------------
            # CAPEX DISPLAY CORRECTO
            # -------------------------
            '''fc_factor = params["schedule"]["FC"][i] if i < len(params["schedule"]["FC"]) else 0
            capex_value = costos["FCI"] * (fc_factor) #/100

            wc = costos["WC"]

            capex_display = capex_value

            # ✔ WC en t_start (NO en año 1)
            if i == t_start:
                capex_display += wc

            # ✔ recuperación al final
            if i == n - 1:
                capex_display -= wc

            ws.cell(row=row, column=2, value=capex_display)'''
            ws.cell(row=row, column=2, value=CapEx[i])

            # -------------------------
            # RESTO DE COLUMNAS
            # -------------------------
            ws.cell(row=row, column=3, value=Revenue[i])
            ws.cell(row=row, column=4, value=CCOP[i])
            ws.cell(row=row, column=5, value=GP[i])
            ws.cell(row=row, column=6, value=Dep[i])
            ws.cell(row=row, column=7, value=TI[i])
            ws.cell(row=row, column=8, value=Taxes[i])
            ws.cell(row=row, column=9, value=CF[i])

            pv_cf = (
                CF[i]
                / ((1 + params["tasa_interes"]) ** años[i])
            )

            npv_acumulado += pv_cf

            ws.cell(row=row, column=10, value=pv_cf)

            ws.cell(row=row, column=11, value=npv_acumulado)


            fila += 1

        # =========================
        # SUMMARY
        # =========================
        # =========================
        # ECONOMIC TABLES
        # =========================
        base = fila + 1

        # ---------------------------------
        # TITULOS
        # ---------------------------------
        ws.merge_cells(start_row=base, start_column=1, end_row=base, end_column=6)
        ws.cell(row=base, column=1, value="ECONOMIC ASSUMPTIONS").font = bold

        ws.merge_cells(start_row=base, start_column=7, end_row=base, end_column=12)
        ws.cell(row=base, column=7, value="ECONOMIC ANALYSIS").font = bold

        # ---------------------------------
        # ECONOMIC ASSUMPTIONS
        # ---------------------------------
        ws.cell(row=base+1, column=1, value="Tax rate")
        ws.cell(row=base+1, column=3, value=params["tasa_impuesto"])
        ws.cell(row=base+1, column=3).number_format = "0.00%"

        ws.cell(row=base+2, column=1, value="Depreciation method")

        if params["metodo_dep"] == 0:
            metodo_dep = "Straight-line"
            extra_label = "Depreciation period"
            extra_labelr = params["periodo_dep"]
            extra_value = "years"

        else:
            metodo_dep = "MACRS"
            extra_label = "Depreciation rate"

            if params["tipo_macrs"] == 0:
                extra_labelr = "MACRS 5"

            elif params["tipo_macrs"] == 1:
                extra_labelr = "MACRS 7"

            elif params["tipo_macrs"] == 2:
                extra_labelr = "MACRS 15"

            extra_value = ""

        ws.cell(row=base+2, column=3, value=metodo_dep)

        ws.cell(row=base+3, column=1, value=extra_label)
        ws.cell(row=base+3, column=3, value=extra_labelr)
        ws.cell(row=base+3, column=4, value=extra_value)

        ws.cell(row=base+4, column=1, value="Discount rate")
        ws.cell(row=base+4, column=3, value=params["tasa_interes"])
        ws.cell(row=base+4, column=3).number_format = "0.00%"

        # ---------------------------------
        # ECONOMIC ANALYSIS
        # ---------------------------------
        ws.cell(row=base+1, column=7, value="NPV")
        ws.cell(row=base+1, column=9, value=npv_acumulado)

        ws.cell(row=base+2, column=7, value="IRR")
        irr_cell = ws.cell(
            row=base+2,
            column=9,
            value=f"=IRR(I{inicio_cf+1}:I{inicio_cf+n})"
        )

        irr_cell.number_format = "0.00%"

        wb.save(nombre_archivo)

    def construir_schedule(self, schedule, vida_operacion):
        """Arma el schedule completo del proyecto.

        Args:
            schedule: dict con FC (lista de fracciones de
                CapEx por año de construcción) y VCOP
                (lista de fracciones de capacidad por año
                de operación durante el ramp-up).
                Convención FC=[0]: planta instantánea
                (0 años de construcción, opera desde el
                año 1).
                Convención Opción B (modo batch): clave
                opcional schedule["batch_recipe"] de tipo
                batch_schedule.BatchRecipe activa el
                puente Capa 3.
            vida_operacion: años de operación de planta
                (la construcción NO cuenta).

        Devuelve dict:
            FC, VCOP, FCOP, WL: listas de largo `vida`
                (vida = construcción + operación).
            t_start: índice del primer año de operación.
            steady_start: primer índice estacionario
                (sin cambios hasta el final).
            cutoff: corte para reporte (steady_start + 3).
            años_display: etiquetas de año para reporte.
            batch: (OPCIONAL — solo si batch_recipe pasó por
                el puente) metadata del ciclo batch:
                cycle_time_s, batches_per_year,
                annual_production_kg, utility_peaks, etc.
                Los consumidores existentes (CashFlowModel,
                results_ui, reporte Excel) NO leen esta
                clave — lectura defensiva con .get().
        """

        # =========================
        # INPUTS
        # =========================
        fc_input = schedule["FC"]
        vcop_input = schedule["VCOP"]
        # Opción B — modo batch (clave opcional, retrocompatible).
        # Si schedule trae 'batch_recipe' (BatchRecipe de Capa 1),
        # activamos el puente.  Caso contrario, comportamiento
        # legacy intacto byte-idéntico.
        _batch_recipe = schedule.get("batch_recipe")
        _batch_availability = schedule.get("batch_availability", 0.90)

        # =========================
        # DETECTAR TIPO
        # =========================
        instantaneo = (fc_input == [0])

        # =========================
        # START OPERACIÓN
        # =========================
        if instantaneo:

            # Año 0 = inversión
            # Año 1 = inicia operación
            t_start = 1

            años_display = list(
                range(0, vida_operacion + 1)
            )

        else:

            # cantidad de años de construcción
            t_start = len(fc_input)

            años_display = list(
                range(1, t_start + vida_operacion + 1)
            )

        # =========================
        # VIDA TOTAL
        # =========================
        vida = len(años_display)

        # =========================
        # TABLAS
        # =========================
        fc = []
        vcop_pct = []

        # =========================
        # LOOP
        # =========================
        for i in range(vida):

            # =========================
            # FC
            # =========================
            if instantaneo:

                if i == 0:
                    fc.append(1)

                else:
                    fc.append(0)

            else:

                if i < len(fc_input):
                    fc.append(fc_input[i])

                else:
                    fc.append(0)

            # =========================
            # VCOP
            # =========================
            if i < t_start:

                vcop_pct.append(0)

            else:

                idx_vcop = i - t_start

                if idx_vcop < len(vcop_input):

                    vcop_pct.append(
                        vcop_input[idx_vcop]
                    )

                else:

                    vcop_pct.append(
                        vcop_input[-1]
                    )

        # =========================
        # FCOP Y WL
        # =========================
        fcop = []
        wl = []

        for i in range(vida):

            # FCOP
            if i >= t_start:
                fcop.append(1)
            else:
                fcop.append(0)

            # WL
            if i == t_start:
                wl.append(1)
            else:
                wl.append(0)

        # =========================
        # DETECTAR ESTADO ESTACIONARIO
        # =========================
        steady_start = None

        for i in range(vida - 1):

            estado_actual = (
                fc[i],
                vcop_pct[i],
                fcop[i],
                wl[i]
            )

            restante_estable = True

            for j in range(i + 1, vida):

                estado_j = (
                    fc[j],
                    vcop_pct[j],
                    fcop[j],
                    wl[j]
                )

                if estado_j != estado_actual:
                    restante_estable = False
                    break

            if restante_estable:
                steady_start = i
                break

        if steady_start is None:
            steady_start = vida - 1

        # =========================
        # CORTE REPORTE
        # =========================
        cutoff = steady_start + 3 #2

        # =========================
        # PUENTE BATCH (Opción B — Capa 3)
        # Estrictamente aditivo: si schedule trae 'batch_recipe',
        # calculamos producción anual equivalente vía Capa 1 y la
        # exponemos como clave 'batch' del dict de retorno.  El
        # cash flow sigue usando VCOP / FCOP / WL con su semántica
        # actual (la capacidad nominal de la planta es la base, y
        # el batch alimenta esa capacidad implícitamente via la
        # producción anual equivalente que ya está en el modelo
        # económico — el factor de utilización VCOP no cambia
        # porque la vida del proyecto es la misma).  Esta clave
        # extra la lee el dashboard de forma defensiva.
        # =========================
        batch_block = None
        if _batch_recipe is not None:
            try:
                import batch_schedule as _bs
                batch_block = _bs.to_schedule_block(
                    _batch_recipe, availability=_batch_availability
                )
            except (ImportError, ValueError) as _e:
                # Falla en Capa 1 (módulo ausente o receta inválida):
                # NO romper el cash flow estacionario — solo logueamos
                # el fallo de forma silenciosa.  Sin batch_block, los
                # consumers ven schedule sin la clave 'batch' y operan
                # como pre-batch.
                batch_block = {"error": f"{type(_e).__name__}: {_e}"}

        # =========================
        # OUTPUT
        # =========================
        result = {
            "FC": fc,
            "VCOP": vcop_pct,
            "FCOP": fcop,
            "WL": wl,
            "steady_start": steady_start,
            "cutoff": cutoff,
            "t_start": t_start,
            "años_display": años_display
        }
        if batch_block is not None:
            result["batch"] = batch_block
        return result