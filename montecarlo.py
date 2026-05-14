# ======================================================
# MONTE CARLO — análisis de incertidumbre sobre NPV/IRR
# ======================================================
# Soporta:
#   - Distribuciones por variable: triangular, normal,
#     uniforme.
#   - Correlación entre variables: matriz NxN opcional,
#     muestreo por Gaussian copula (Cholesky de la matriz
#     + Φ + F⁻¹).
#   - Fast path numpy: precomputa todo lo que NO cambia
#     entre corridas (BP, CONS, UTS, FCOP, schedule,
#     depreciación base) y vectoriza el cash flow.
#
# Variables soportadas:
#   - precio de cada Key Product
#   - precio de cada Raw Material
#   - ISBL Capital Cost
#
# Output:
#   - distribución NPV/IRR sobre N corridas
#   - estadísticos (mean, std, percentiles, P(NPV<0))
#   - tornado determinístico (vary-one-at-a-time)
#
# Refs: Towler & Sinnott §9, Turton §10.
# Copula: Nelsen (2006) "An Introduction to Copulas",
# Iman-Conover (1982).
# ======================================================

import copy
import math
import statistics

from dataclasses import dataclass, field
from typing import List, Optional

import numpy as np

from flujoflujoclass import CostModel, CashFlowModel


# ======================================================
# CONSTANTES
# ======================================================

KIND_PRODUCT_PRICE      = "product_price"
KIND_RAW_MATERIAL_PRICE = "raw_material_price"
KIND_ISBL               = "isbl"

DIST_TRIANGULAR = "triangular"
DIST_NORMAL     = "normal"
DIST_UNIFORM    = "uniform"

VALID_DISTS = (DIST_TRIANGULAR, DIST_NORMAL, DIST_UNIFORM)


# ======================================================
# VARIABLE INCIERTA
# ======================================================

@dataclass
class VariableIncierta:
    """Spec de una variable a someter a Monte Carlo.

    kind: KIND_PRODUCT_PRICE | KIND_RAW_MATERIAL_PRICE
          | KIND_ISBL
    indice: posición dentro de data["key_products"] o
            data["raw_materials"].  Ignorado para ISBL.
    nombre: para mostrar en outputs.
    dist:  DIST_TRIANGULAR | DIST_NORMAL | DIST_UNIFORM

    Convenciones de los tres parámetros valor_min,
    valor_mode, valor_max:

        triangular: a=min, c=mode, b=max  (estricto)
        uniform:    a=min, b=max  (mode ignorado)
        normal:     mean=mode, std=(max-min)/4
                    (con max-min ≈ 4σ ≈ 95% del rango)
    """

    kind: str
    indice: int
    nombre: str
    valor_min: float
    valor_mode: float
    valor_max: float
    dist: str = DIST_TRIANGULAR

    def __post_init__(self):
        if self.dist not in VALID_DISTS:
            raise ValueError(f"dist inválida: {self.dist!r}")
        if self.dist in (DIST_TRIANGULAR, DIST_UNIFORM):
            if not (self.valor_min <= self.valor_max):
                raise ValueError(
                    f"'{self.nombre}': min ≤ max requerido"
                )
        if self.dist == DIST_TRIANGULAR:
            if not (self.valor_min <= self.valor_mode <= self.valor_max):
                raise ValueError(
                    f"'{self.nombre}': min ≤ mode ≤ max requerido"
                )

    # ---- inverse CDF, recibe u ∈ (0,1) ----
    def inverse_cdf(self, u):
        u = np.clip(u, 1e-10, 1 - 1e-10)

        if self.dist == DIST_UNIFORM:
            return self.valor_min + (self.valor_max - self.valor_min) * u

        if self.dist == DIST_NORMAL:
            mean = self.valor_mode
            std = max((self.valor_max - self.valor_min) / 4.0, 1e-12)
            return mean + std * _norm_ppf_vec(u)

        # triangular
        a, b, c = self.valor_min, self.valor_max, self.valor_mode
        Fc = (c - a) / (b - a) if b > a else 0.5
        # split
        result = np.where(
            u <= Fc,
            a + np.sqrt(u * (b - a) * (c - a)),
            b - np.sqrt((1 - u) * (b - a) * (b - c)),
        )
        return result

    # ---- sample escalar (sin correlación) ----
    def sample(self, rng):
        u = rng.random()
        return float(self.inverse_cdf(np.array([u]))[0])


# ======================================================
# INVERSE CDF NORMAL — vectorizada
# ======================================================
# statistics.NormalDist().inv_cdf es escalar; vectorizamos.

_norm_inv_cdf_scalar = statistics.NormalDist().inv_cdf
_norm_ppf_vec = np.vectorize(_norm_inv_cdf_scalar)


# ======================================================
# SAMPLING CORRELACIONADO (Gaussian copula)
# ======================================================

def _matriz_correlacion_a_array(K, correlacion):
    """Acepta correlacion como:
        - None  → identidad
        - dict {(i,j): rho}  → simétrica, diagonal 1
        - np.ndarray KxK     → tal cual

    Devuelve np.ndarray KxK validado (simétrica, diag 1).
    """
    if correlacion is None:
        return np.eye(K)

    if isinstance(correlacion, dict):
        M = np.eye(K)
        for (i, j), rho in correlacion.items():
            if i == j:
                continue
            M[i, j] = rho
            M[j, i] = rho
        return M

    M = np.asarray(correlacion, dtype=float)
    if M.shape != (K, K):
        raise ValueError(
            f"Matriz correlación con shape {M.shape}, "
            f"se esperaba ({K},{K})"
        )
    # asegurar simetría y diagonal 1
    M = (M + M.T) / 2
    np.fill_diagonal(M, 1.0)
    return M


def _muestrear_correlacionado(variables, n_runs, correlacion, seed):
    """Devuelve array (n_runs, K) de muestras según las dist
    marginales de `variables` y la matriz `correlacion`.

    Si correlacion es None → muestreo independiente.
    """
    K = len(variables)
    rng = np.random.default_rng(seed)

    M = _matriz_correlacion_a_array(K, correlacion)

    if np.allclose(M, np.eye(K)):
        # Path rápido independiente
        U = rng.random((n_runs, K))
    else:
        # Gaussian copula
        try:
            L = np.linalg.cholesky(M)
        except np.linalg.LinAlgError as e:
            raise ValueError(
                "Matriz de correlación no es positive semi-definite. "
                "Revisá los coeficientes (|ρ|<1 entre pares; "
                "para 3+ variables hay restricciones adicionales)."
            ) from e

        Z = rng.standard_normal((n_runs, K))
        Z_corr = Z @ L.T  # muestras N(0,M)
        # Φ(z) → U(0,1) marginales
        U = 0.5 * (1.0 + np.vectorize(math.erf)(Z_corr / math.sqrt(2)))

    # Aplicar F⁻¹ por columna
    X = np.empty_like(U)
    for k, v in enumerate(variables):
        X[:, k] = v.inverse_cdf(U[:, k])

    return X


# ======================================================
# APLICAR MUESTRAS AL data dict (camino lento, exacto)
# ======================================================

def _aplicar_muestras(data, variables, muestras_fila):
    nuevo = copy.deepcopy(data)
    for var, valor in zip(variables, muestras_fila):
        if var.kind == KIND_PRODUCT_PRICE:
            nuevo["key_products"][var.indice]["price"] = float(valor)
        elif var.kind == KIND_RAW_MATERIAL_PRICE:
            nuevo["raw_materials"][var.indice]["price"] = float(valor)
        elif var.kind == KIND_ISBL:
            nuevo["ISBL"] = float(valor)
        else:
            raise ValueError(f"Kind desconocido: {var.kind}")
    return nuevo


# ======================================================
# NPV / IRR
# ======================================================

def _npv(cf, años, tasa):
    return sum(
        cf[i] / ((1 + tasa) ** años[i])
        for i in range(len(cf))
    )


def _irr_biseccion(cf, años, tol=1e-6, max_iter=200):
    if not cf or all(c >= 0 for c in cf) or all(c <= 0 for c in cf):
        return None
    lo, hi = -0.99, 10.0
    f_lo = _npv(cf, años, lo)
    f_hi = _npv(cf, años, hi)
    if f_lo * f_hi > 0:
        return None
    for _ in range(max_iter):
        mid = (lo + hi) / 2
        f_mid = _npv(cf, años, mid)
        if abs(f_mid) < tol:
            return mid
        if f_lo * f_mid < 0:
            hi, f_hi = mid, f_mid
        else:
            lo, f_lo = mid, f_mid
    return (lo + hi) / 2


def _correr_un_escenario_lento(data, params):
    costos = CostModel(data).calcular()
    cf = CashFlowModel(costos, params).calcular()
    tasa = params["tasa_interes"]
    npv = _npv(cf["CF"], cf["años"], tasa)
    irr = _irr_biseccion(cf["CF"], cf["años"])
    return npv, irr


# ======================================================
# FAST PATH NUMPY — vectorizado sobre N corridas
# ======================================================
#
# Idea: para cada corrida solo cambian:
#   - precios de key_products  (vector de length K_p)
#   - precios de raw_materials (vector de length K_rm)
#   - ISBL                     (escalar)
#
# Todo lo demás (BP, CONS, UTS, FCOP, schedule, dep base,
# OSBL/ENG/CONT %s, tax rate, royalties_pct, WC%) es
# invariante.  Lo precomputamos UNA vez y vectorizamos el
# cash flow en numpy.
#
# Cuidado: replicamos exactamente la lógica de CostModel
# y CashFlowModel para que los números coincidan.
# ======================================================


def _precomputar_invariantes(data, params):
    """Calcula lo que no depende de las variables inciertas.

    Devuelve dict con vectores/escalares listos para el
    fast path:
        BP, CONS, UTS, OSBL_pct, ENG_pct, CONT_pct, WC_pct,
        royalties_pct, tax_rate, discount,
        FCOP_per_FCI (lineal en FCI)  — coef que multiplica
                                        a FCI dentro de FCOP
        FCOP_const   — parte fija (labor + supervision +
                       salary_overhead + plant_overhead +
                       tax_insurance_const + general)
        FCOP_per_ISBL_OSBL — coef * (ISBL+OSBL)
        D_base_template — depreciación según método
                           (función de FCI; se escalará)
        años_arr, t_start, vida, FC_arr, VCOP_arr, WL_arr
        Para REV: lista de flows de key_products
        Para RM:  lista de flows de raw_materials
    """

    # --- shares estáticas de capital ---
    OSBL_pct = data["OSBL_pct"]
    ENG_pct  = data["ENG_pct"]
    CONT_pct = data["CONT_pct"]
    WC_pct   = data["WC_pct"]
    # FCI = ISBL * (1 + OSBL_pct) * (1 + ENG_pct + CONT_pct)
    # WC = WC_pct * FCI
    FCI_multiplier = (1.0 + OSBL_pct) * (1.0 + ENG_pct + CONT_pct)

    # --- FCOP descompuesto en (constante) + (k_isbl_osbl)*ISBL_OSBL + (k_fci)*FCI ---
    fc = data["FCOP_inputs"]
    labor = fc["labor"]
    supervision = fc["supervision_pct"] * labor
    salary_overhead = fc["salary_overhead_pct"] * (labor + supervision)
    # plant_overhead depende de maintenance que depende de FCI:
    # plant_overhead = plant_pct * (labor + maintenance)
    # maintenance = maint_pct * FCI
    # tax_insurance = tax_pct * (ISBL + OSBL) = tax_pct * ISBL * (1+OSBL_pct)
    # interest = int_pct * FCI
    # general = gen_pct * WC = gen_pct * WC_pct * FCI

    FCOP_const = labor + supervision + salary_overhead + fc["plant_overhead_pct"] * labor

    FCOP_k_isbl_osbl = fc["tax_insurance_pct"]  # multiplica a (ISBL+OSBL) = ISBL*(1+OSBL%)

    FCOP_k_fci = (
        fc["maintenance_pct"]
        + fc["plant_overhead_pct"] * fc["maintenance_pct"]
        + fc["interest_pct"]
        + fc["general_expenses_pct"] * WC_pct
    )

    # --- BP, CONS, UTS (no dependen de variables inciertas) ---
    production = (
        data["key_products"][0]["flow"]
        if data["key_products"] else 1.0
    )
    BP = sum(f["flow"] * f["price"] for f in data["byproducts"]) / 1e6
    CONS = sum(f["coef"] * f["price"] * production for f in data["consumables"]) / 1e6
    UTS  = sum(f["coef"] * f["price"] * production for f in data["utilities"])  / 1e6

    # --- flows e índices de RM/KP (precios serán variables) ---
    kp_flows = np.array([f["flow"] for f in data["key_products"]], dtype=float)
    rm_flows = np.array([f["flow"] for f in data["raw_materials"]], dtype=float)

    # precios base (caso sin sampling, o variables no inciertas)
    kp_prices_base = np.array([f["price"] for f in data["key_products"]], dtype=float)
    rm_prices_base = np.array([f["price"] for f in data["raw_materials"]], dtype=float)

    # --- schedule ya expandido ---
    sched = params["schedule"]
    FC_arr   = np.array(sched["FC"],   dtype=float)
    VCOP_arr = np.array(sched["VCOP"], dtype=float)
    WL_arr   = np.array(sched["WL"],   dtype=float)
    años_arr = np.array(sched["años_display"], dtype=float)
    t_start  = sched["WL"].index(1)
    vida = len(FC_arr)

    # FCOP solo cuando opera (i >= t_start)
    op_mask = np.zeros(vida)
    op_mask[t_start:] = 1.0

    # --- depreciación: tabla base "por unidad de FCI"
    # luego se multiplica por FCI en el cálculo ---
    if params["metodo_dep"] == 0:
        vida_dep = int(params["periodo_dep"])
        d_per_unit = np.zeros(vida)
        for i in range(t_start, min(t_start + vida_dep, vida)):
            d_per_unit[i] = 1.0 / vida_dep
    else:
        macrs_tablas = {
            0: [0.20, 0.32, 0.192, 0.1152, 0.1152, 0.0576],
            1: [0.1429, 0.2449, 0.1749, 0.1249, 0.0893, 0.0892, 0.0893, 0.0446],
            2: [0.05, 0.095, 0.0855, 0.077, 0.0693, 0.0623, 0.059, 0.059,
                0.0591, 0.059, 0.0591, 0.059, 0.0591, 0.059, 0.0591, 0.0295],
        }
        tabla = macrs_tablas[params["tipo_macrs"]]
        d_per_unit = np.zeros(vida)
        for i in range(t_start, vida):
            idx = i - t_start
            if idx < len(tabla):
                d_per_unit[i] = tabla[idx]

    return {
        "FCI_multiplier":  FCI_multiplier,
        "WC_pct":          WC_pct,
        "OSBL_pct":        OSBL_pct,
        "FCOP_const":      FCOP_const,
        "FCOP_k_isbl_osbl":FCOP_k_isbl_osbl,
        "FCOP_k_fci":      FCOP_k_fci,
        "BP":              BP,
        "CONS":            CONS,
        "UTS":             UTS,
        "kp_flows":        kp_flows,
        "rm_flows":        rm_flows,
        "kp_prices_base":  kp_prices_base,
        "rm_prices_base":  rm_prices_base,
        "FC_arr":          FC_arr,
        "VCOP_arr":        VCOP_arr,
        "op_mask":         op_mask,
        "años_arr":        años_arr,
        "t_start":         t_start,
        "vida":            vida,
        "tasa_impuesto":   params["tasa_impuesto"],
        "tasa_interes":    params["tasa_interes"],
        "royalties_pct":   data["FCOP_inputs"]["royalties_pct"],
        "d_per_unit":      d_per_unit,
    }


def _cash_flow_vectorizado(inv, ISBL_arr, REV_arr, RM_arr):
    """Calcula NPV e IRR (None si no calculable) para cada
    corrida.

    Args:
        inv: dict de precomputar_invariantes.
        ISBL_arr: (N,) array de ISBL por corrida.
        REV_arr:  (N,) array de Revenue base por corrida.
        RM_arr:   (N,) array de RawMaterials por corrida.

    Devuelve: npvs (N,), irrs (lista N de float o None).
    """

    N = len(ISBL_arr)
    vida = inv["vida"]
    t_start = inv["t_start"]

    FCI = ISBL_arr * inv["FCI_multiplier"]  # (N,)
    WC  = inv["WC_pct"] * FCI               # (N,)
    OSBL = ISBL_arr * inv["OSBL_pct"]
    ISBL_plus_OSBL = ISBL_arr + OSBL

    FCOP = (
        inv["FCOP_const"]
        + inv["FCOP_k_isbl_osbl"] * ISBL_plus_OSBL
        + inv["FCOP_k_fci"] * FCI
    )  # (N,)

    VCOP = RM_arr - inv["BP"] + inv["CONS"] + inv["UTS"]  # (N,)

    # CapEx por año (vida,)
    capex_per_year = np.outer(FCI, inv["FC_arr"])  # (N, vida)
    # WC en t_start
    capex_per_year[:, t_start] += WC
    # Recuperación WC al final
    capex_per_year[:, -1] -= WC

    # Revenue/VCOP por año = base * factor_vcop[i] (solo cuando opera)
    factor = inv["VCOP_arr"]  # (vida,)
    rev_per_year = np.outer(REV_arr, factor)  # (N, vida)
    vcop_per_year = np.outer(VCOP, factor)

    # FCOP solo cuando opera
    fcop_per_year = np.outer(FCOP, inv["op_mask"])

    # Royalties simplificada (versión activa del modelo):
    #   Revenue_base * royalties_pct cuando i >= t_start
    royalties_per_year = np.outer(
        REV_arr * inv["royalties_pct"],
        inv["op_mask"],
    )

    ccop_per_year = fcop_per_year + vcop_per_year + royalties_per_year

    gp_per_year = rev_per_year - ccop_per_year
    # GP = 0 antes de t_start
    gp_per_year[:, :t_start] = 0.0

    # Depreciación: D[i] = FCI * d_per_unit[i]
    dep_per_year = np.outer(FCI, inv["d_per_unit"])

    ti_per_year = gp_per_year - dep_per_year

    # Tax accrued = max(TI*tasa, 0)
    tax_accrued = np.maximum(ti_per_year * inv["tasa_impuesto"], 0.0)

    # Tax paid: desfase 1 año.  taxes_paid[i] = taxes_accrued[i-1].
    taxes_paid = np.zeros_like(tax_accrued)
    taxes_paid[:, 1:] = tax_accrued[:, :-1]

    cf_per_year = gp_per_year - taxes_paid - capex_per_year

    # NPV
    años = inv["años_arr"]
    discount = (1.0 + inv["tasa_interes"]) ** años  # (vida,)
    npvs = (cf_per_year / discount).sum(axis=1)     # (N,)

    # IRR por corrida (no vectorizable trivialmente)
    años_lst = años.tolist()
    irrs = []
    for n in range(N):
        cf_lst = cf_per_year[n].tolist()
        irrs.append(_irr_biseccion(cf_lst, años_lst))

    return npvs, irrs


# ======================================================
# CORRIDA MONTE CARLO (API pública)
# ======================================================

def correr_montecarlo(
        data_base,
        params,
        variables: List[VariableIncierta],
        n_runs: int = 5000,
        seed: Optional[int] = None,
        correlacion=None,
        usar_fast_path: bool = True,
):
    """Ejecuta MC con sampling correlacionado opcional.

    correlacion:
        - None: variables independientes
        - dict {(i,j): rho}: define pares; resto diagonal
        - np.ndarray KxK: matriz completa

    usar_fast_path: si True (default), vectoriza en numpy;
    requiere que las variables sean solo de los tres kinds
    soportados.  Si False, hace deepcopy + CostModel cada
    iteración (más lento pero exacto frente a cambios de
    estructura del data dict).
    """

    # Sampling correlacionado
    samples_arr = _muestrear_correlacionado(
        variables, n_runs, correlacion, seed,
    )

    if usar_fast_path:
        npvs, irrs = _correr_fast_path(
            data_base, params, variables, samples_arr,
        )
    else:
        npvs, irrs = _correr_slow_path(
            data_base, params, variables, samples_arr,
        )

    return {
        "npvs":      list(npvs),
        "irrs":      list(irrs),
        "samples":   samples_arr.tolist(),
        "variables": variables,
        "stats":     _estadisticos(list(npvs), list(irrs)),
    }


def _correr_fast_path(data_base, params, variables, samples_arr):
    inv = _precomputar_invariantes(data_base, params)
    N, K = samples_arr.shape

    # Construir REV / RM / ISBL por corrida.
    # Empezamos con los valores base y sobreescribimos según
    # cada variable.
    kp_prices_runs = np.tile(inv["kp_prices_base"], (N, 1))
    rm_prices_runs = np.tile(inv["rm_prices_base"], (N, 1))
    isbl_runs = np.full(N, data_base["ISBL"], dtype=float)

    for k, var in enumerate(variables):
        col = samples_arr[:, k]
        if var.kind == KIND_PRODUCT_PRICE:
            kp_prices_runs[:, var.indice] = col
        elif var.kind == KIND_RAW_MATERIAL_PRICE:
            rm_prices_runs[:, var.indice] = col
        elif var.kind == KIND_ISBL:
            isbl_runs = col

    # Revenue = Σ(flow * price) / 1e6 sobre key_products
    REV_runs = (kp_prices_runs * inv["kp_flows"]).sum(axis=1) / 1e6
    RM_runs  = (rm_prices_runs * inv["rm_flows"]).sum(axis=1) / 1e6

    return _cash_flow_vectorizado(inv, isbl_runs, REV_runs, RM_runs)


def _correr_slow_path(data_base, params, variables, samples_arr):
    N, _ = samples_arr.shape
    npvs = np.empty(N)
    irrs = []
    for n in range(N):
        data_run = _aplicar_muestras(data_base, variables, samples_arr[n])
        npv, irr = _correr_un_escenario_lento(data_run, params)
        npvs[n] = npv
        irrs.append(irr)
    return npvs, irrs


# ======================================================
# ESTADÍSTICOS
# ======================================================

def _percentil(valores_ordenados, p):
    n = len(valores_ordenados)
    if n == 0:
        return None
    if n == 1:
        return valores_ordenados[0]
    k = (n - 1) * p / 100.0
    f = int(k)
    c = min(f + 1, n - 1)
    return (
        valores_ordenados[f]
        + (k - f) * (valores_ordenados[c] - valores_ordenados[f])
    )


def _estadisticos(npvs, irrs):
    npvs_ord = sorted(npvs)
    n = len(npvs)
    mean = sum(npvs) / n if n else 0.0
    var = (
        sum((x - mean) ** 2 for x in npvs) / (n - 1)
        if n > 1 else 0.0
    )
    std = var ** 0.5
    irrs_validas = [r for r in irrs if r is not None]

    return {
        "n":           n,
        "npv_mean":    mean,
        "npv_std":     std,
        "npv_min":     npvs_ord[0] if n else None,
        "npv_max":     npvs_ord[-1] if n else None,
        "npv_p10":     _percentil(npvs_ord, 10),
        "npv_p50":     _percentil(npvs_ord, 50),
        "npv_p90":     _percentil(npvs_ord, 90),
        "p_npv_neg":   sum(1 for x in npvs if x < 0) / n if n else 0.0,
        "irr_mean":    (
            sum(irrs_validas) / len(irrs_validas)
            if irrs_validas else None
        ),
        "irr_p10":     (
            _percentil(sorted(irrs_validas), 10)
            if irrs_validas else None
        ),
        "irr_p50":     (
            _percentil(sorted(irrs_validas), 50)
            if irrs_validas else None
        ),
        "irr_p90":     (
            _percentil(sorted(irrs_validas), 90)
            if irrs_validas else None
        ),
        "irr_n_valid": len(irrs_validas),
    }


# ======================================================
# TORNADO CHART (determinístico)
# ======================================================

def correr_tornado(
        data_base,
        params,
        variables: List[VariableIncierta],
):
    """Vary-one-at-a-time entre min y max manteniendo el
    resto en mode.  Ordena por |swing| descendente."""

    data_base_mode = copy.deepcopy(data_base)
    for v in variables:
        _set_var(data_base_mode, v, v.valor_mode)

    npv_base, _ = _correr_un_escenario_lento(data_base_mode, params)

    resultados = []
    for v in variables:
        # min
        data_lo = copy.deepcopy(data_base_mode)
        _set_var(data_lo, v, v.valor_min)
        npv_lo, _ = _correr_un_escenario_lento(data_lo, params)
        # max
        data_hi = copy.deepcopy(data_base_mode)
        _set_var(data_hi, v, v.valor_max)
        npv_hi, _ = _correr_un_escenario_lento(data_hi, params)

        resultados.append({
            "nombre":     v.nombre,
            "kind":       v.kind,
            "npv_base":   npv_base,
            "npv_low":    npv_lo,
            "npv_high":   npv_hi,
            "delta_low":  npv_lo - npv_base,
            "delta_high": npv_hi - npv_base,
            "swing":      abs(npv_hi - npv_lo),
        })

    resultados.sort(key=lambda r: r["swing"], reverse=True)
    return resultados


def _set_var(data, var, valor):
    if var.kind == KIND_PRODUCT_PRICE:
        data["key_products"][var.indice]["price"] = float(valor)
    elif var.kind == KIND_RAW_MATERIAL_PRICE:
        data["raw_materials"][var.indice]["price"] = float(valor)
    elif var.kind == KIND_ISBL:
        data["ISBL"] = float(valor)


# ======================================================
# EXPORTAR A EXCEL (hoja extra)
# ======================================================

def exportar_montecarlo_excel(
        archivo,
        resultado_mc,
        resultado_tornado,
):
    """Agrega hojas 'Monte Carlo' y 'Tornado' a un .xlsx
    existente (o crea uno nuevo)."""

    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font
    import os

    if os.path.exists(archivo):
        wb = load_workbook(archivo)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Monte Carlo" in wb.sheetnames:
        del wb["Monte Carlo"]
    ws = wb.create_sheet("Monte Carlo")

    bold = Font(bold=True)
    stats = resultado_mc["stats"]

    ws.cell(1, 1, "MONTE CARLO SUMMARY").font = bold
    summary = [
        ("Runs",        stats["n"]),
        ("NPV mean",    stats["npv_mean"]),
        ("NPV std",     stats["npv_std"]),
        ("NPV min",     stats["npv_min"]),
        ("NPV P10",     stats["npv_p10"]),
        ("NPV P50",     stats["npv_p50"]),
        ("NPV P90",     stats["npv_p90"]),
        ("NPV max",     stats["npv_max"]),
        ("P(NPV<0)",    stats["p_npv_neg"]),
        ("IRR P10",     stats["irr_p10"]),
        ("IRR P50",     stats["irr_p50"]),
        ("IRR P90",     stats["irr_p90"]),
        ("IRR valid #", stats["irr_n_valid"]),
    ]
    for i, (k, v) in enumerate(summary, start=2):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)

    fila_raw = len(summary) + 4
    ws.cell(fila_raw, 1, "RAW RUNS").font = bold

    headers = ["Run", "NPV", "IRR"] + [v.nombre for v in resultado_mc["variables"]]
    for col, h in enumerate(headers, start=1):
        ws.cell(fila_raw + 1, col, h).font = bold

    for i, (npv, irr, muestras) in enumerate(zip(
        resultado_mc["npvs"],
        resultado_mc["irrs"],
        resultado_mc["samples"],
    )):
        ws.cell(fila_raw + 2 + i, 1, i + 1)
        ws.cell(fila_raw + 2 + i, 2, npv)
        ws.cell(fila_raw + 2 + i, 3, irr if irr is not None else "")
        for j, m in enumerate(muestras):
            ws.cell(fila_raw + 2 + i, 4 + j, m)

    if "Tornado" in wb.sheetnames:
        del wb["Tornado"]
    ws_t = wb.create_sheet("Tornado")

    ws_t.cell(1, 1, "TORNADO CHART (deterministic)").font = bold
    headers_t = ["Variable", "NPV @ min", "NPV @ base", "NPV @ max",
                 "Δ low", "Δ high", "Swing"]
    for col, h in enumerate(headers_t, start=1):
        ws_t.cell(2, col, h).font = bold

    for i, r in enumerate(resultado_tornado, start=3):
        ws_t.cell(i, 1, r["nombre"])
        ws_t.cell(i, 2, r["npv_low"])
        ws_t.cell(i, 3, r["npv_base"])
        ws_t.cell(i, 4, r["npv_high"])
        ws_t.cell(i, 5, r["delta_low"])
        ws_t.cell(i, 6, r["delta_high"])
        ws_t.cell(i, 7, r["swing"])

    wb.save(archivo)
